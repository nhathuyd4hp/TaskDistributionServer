import argparse
import logging
import os
import re
import shutil
import stat
import sys
import time
from pathlib import Path

import openpyxl
import win32com.client
from config_access_token import token_file  # noqa
from Nasiwak import *  # noqa
from openpyxl import load_workbook
from pywinauto import Desktop
from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)


parser = argparse.ArgumentParser()
parser.add_argument("--task-id", required=True)
args = parser.parse_args()
task_id = args.task_id
log_path = Path(Path(__file__).resolve().parent).parent.parent.parent / "logs" / f"{task_id}.log"
log_path.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s  - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(str(log_path))],
)
logging.info(f"Log file created: {log_path}")

# Replace with your actual file path
file_path = os.path.join(os.getcwd(), "Access_token", "Access_token.txt")
logging.info(f"file path for text file is: {file_path}")
# Open and read the file
with open(file_path, "r", encoding="utf-8") as file:
    content = file.read()
logging.info(f"Extracted text from .txt file is: {content}")

REPO_OWNER = "Nasiwak"  # Your GitHub username
REPO_NAME = "Hajime_shinki"  # your repo name
CURRENT_VERSION = "v1.0.4"  # this bot version
ACCESS_TOKEN = content  # main token
Bot_Update(REPO_OWNER, REPO_NAME, CURRENT_VERSION, ACCESS_TOKEN)  # noqa


andpad_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/andpad.json"
webaccess_json_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/webaccess.json"
sharepoint_json_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/sharepoint.json"
maildealer_json_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/MailDealer.json"

andpad_config = create_json_config(andpad_url, ACCESS_TOKEN)  # noqa
webaccess_config = create_json_config(webaccess_json_url, ACCESS_TOKEN)  # noqa
sharepoint_config = create_json_config(sharepoint_json_url, ACCESS_TOKEN)  # noqa
maildealer_config = create_json_config(maildealer_json_url, ACCESS_TOKEN)  # noqa

excelfile = "Hajime_案件化.xlsm"
text_file = "重要.txt"
folder_path = r"Ankens"
files_folder = "Files"  # to create Files folder
Accessurl = "https://webaccess.nsk-cad.com/"
Maildealerurl = "https://md29.maildealer.jp/index.php"
Andpadurl = "https://andpad.jp/login?iss=https%3A%2F%2Fauth.andpad.jp%2F"
macro_file = os.path.join(os.getcwd(), excelfile)
kansai = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E9%96%A2%E8%A5%BF%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
shikoku = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E4%B8%AD%E5%9B%BD%E3%83%BB%E5%9B%9B%E5%9B%BD%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
toukai = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E6%9D%B1%E6%B5%B7%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
kantou = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E9%96%A2%E6%9D%B1%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
kyuushuu = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E4%B9%9D%E5%B7%9E%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
koriyama = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E9%83%A1%E5%B1%B1%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
sendai = "https://nskkogyo.sharepoint.com/sites/%E9%A3%AF%E7%94%B0/DocLib4/Forms/AllItems.aspx?isAscending=false&id=%2Fsites%2F%E9%A3%AF%E7%94%B0%2FDocLib4%2F%E4%B8%80%E5%BB%BA%E8%A8%AD%EF%BC%88%E4%BB%99%E5%8F%B0%EF%BC%89&sortField=Modified&viewid=9cdc060d%2Da101%2D4bc3%2Dae21%2De789da53314e"
logging.info(f"macro file path is: {macro_file}")

# if not os.path.exists(f'{excelfile}'):
if not os.path.exists(excelfile):
    logging.info("案件化 ファイル見つかりません")
    sys.exit()
else:
    logging.info("案件化 File found, moving to the next step")


def remove_readonly(func, path, exc_info):
    os.chmod(path, stat.S_IWRITE)
    func(path)


for attempt in range(5):  # noqa
    try:
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path, onerror=remove_readonly)
            logging.info(f"Deleted contents of {folder_path} folder.")
        break
    except PermissionError:
        logging.warning("Folder is busy. Retrying...")
        time.sleep(1)
else:
    logging.error("Failed to delete folder after multiple attempts.")


if os.path.exists(text_file):
    logging.info(f"found {text_file}, moving to process the ankens..")
else:
    logging.info(f"Didn't find {text_file}, exiting....")
    sys.exit()


def create_download_directory(folder_path):

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        logging.info(f"{folder_path} created")
    else:
        logging.info(f"{folder_path} exists")


def wait_for_download(download_folder):
    seconds = 0
    dl_wait = True
    while dl_wait and seconds < 40:
        time.sleep(1)
        dl_wait = False
        for fname in os.listdir(download_folder):
            if fname.endswith(".crdownload") or fname.endswith(".part"):
                dl_wait = True
        seconds += 1
    if seconds == 40:
        logging.info("Download timed out")
    else:
        logging.info("Download completed")


def transform_value(value):
    # Regular expression to extract the main part and the fraction (e.g., "可児市土田第４期", "1/4")
    match = re.match(r"(.*)\s(\d+)/(\d+)", value)
    if not match:
        return value  # Return the original value if the format doesn't match

    main_part, current, total = match.groups()

    if current == "1" and total == "1":
        # Special case for 1/1
        return f"{main_part} (1棟)"
    else:
        # General case for other fractions like 1/4
        return f"{main_part} {current}号棟({total}棟)"


def mail_dealer_login():

    site = maildealer_config["MailDealer_url"]
    driver.get(site)
    time.sleep(2)

    # locate logid and pass boxes
    logid = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(("name", maildealer_config["MailDealer_name"]["MailDealer_Username"]))
    )
    logid.clear()
    # logid.send_keys("Nasiwakロボ")
    logid.send_keys("aman")
    time.sleep(0.5)

    log_pass = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(("name", maildealer_config["MailDealer_name"]["MailDealer_Password"]))
    )
    log_pass.clear()
    # log_pass.send_keys("ouocf68l")
    log_pass.send_keys("8iod3vqx")
    time.sleep(0.5)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Login_submit"]))
    ).click()
    # logid.send_keys(Keys.ENTER)
    time.sleep(3)
    logging.info("Maildealer Login successful")

    # switch to side frame
    sidemenu = driver.find_element(By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Sidemenu"])
    driver.switch_to.frame(sidemenu)
    time.sleep(0.5)

    # click on Folder
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//span[@title='≪ベトナム納期≫一建設(FAX・メール)']"))
    ).click()
    logging.info("一建設 Folder opened")
    time.sleep(2)


def Andpad_login():

    # Open a new tab using JavaScript
    driver.execute_script("window.open('about:blank','_blank');")

    # Switch to the newly opened tab
    driver.switch_to.window(driver.window_handles[-1])

    # Open ANDPAD
    driver.get(Andpadurl)
    time.sleep(2)

    button = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, andpad_config["xpaths"]["andpad_ログイン画面へ"]))
    )
    button.click()

    # Wait for the username and password fields to be present
    username_field = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_メールアドレス"]))
    )
    password_field = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_パスワード"]))
    )

    # Input your credentials
    username_field.send_keys("ighd@nsk-cad.com")
    password_field.send_keys("nsk159753")

    # Locate and click the login button
    login_button = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, andpad_config["xpaths"]["andpad_ログイン"]))
    )
    login_button.click()
    time.sleep(2)
    logging.info("Andpad Login successful")

    # Using home to
    # driver.close()
    # logging.info(f"Closed Andpad Tab after logging")


def sharepoint_login():

    url = "https://nskkogyo.sharepoint.com/sites/2021"
    time.sleep(1)

    # open a new tab
    # driver.execute_script("window.open('about:blank','_blank');")

    # Switch to the newly opened tab
    # driver.switch_to.window(driver.window_handles[-1])

    # Assuming the login page has input fields with IDs 'username' and 'password'
    driver.get(url)

    # username = "kushalnasiwak@nskkogyo.onmicrosoft.com"
    # password = "D&059794748972ot"
    username = sharepoint_config["username2"]
    password = sharepoint_config["password2"]
    logging.info(f"ID: {username}\nPass:{password}")
    time.sleep(4)

    # Find the username input field on the login page
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["email"]))
    ).clear()
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["email"]))
    ).send_keys(username)
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["loggin_button"]))
    ).click()
    time.sleep(1)
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["password"]))
    ).clear()
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["password"]))
    ).send_keys(password)
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["loggin_button"]))
    ).click()
    time.sleep(1)
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["checkbox"]))
    ).click()
    time.sleep(0.5)
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, sharepoint_config["login_xpaths"]["yes_button"]))
    ).click()
    time.sleep(3)

    logging.info("Logged in to Sharepoint")
    # driver.close()
    # logging.info('Successfully closed Sharepoint windows')


def webaccess_login():

    # open a new tab
    driver.execute_script("window.open('about:blank','_blank');")

    # Switch to the newly opened tab
    driver.switch_to.window(driver.window_handles[-1])

    # Open WebAccess
    driver.get(Accessurl)
    time.sleep(2)

    # Enter login id
    logid = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["ログイン_xpaths"]["ログインID"]))
    )
    logid.clear()
    logid.send_keys("NasiwakRobot")

    # Enter password
    logpassword = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["ログイン_xpaths"]["パスワード"]))
    )
    logpassword.clear()
    logpassword.send_keys("159753")

    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["ログイン_xpaths"]["ログイン"]))
    ).submit()
    time.sleep(1)

    # Locate and click the 受注一覧 button
    受注一覧_button = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧"]))
    )
    受注一覧_button.click()

    # click on reset button
    reset_xpath_button = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["リセット"]))
    )
    reset_xpath_button.click()
    time.sleep(2)

    logging.info("Succesfully logged in to WebAccess")


def check_box():
    driver.switch_to.default_content()
    main_frame = driver.find_element(By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Mainmenu"])
    driver.switch_to.frame(main_frame)
    time.sleep(0.5)
    try:
        # click on blue tick
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Tickmark"]))
        ).click()
        time.sleep(1)

        # select the checkbox below kanrenzuke
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "(//*[@class='checkbox__label' and contains(text(),'このメールと同じ親番号のメールをすべて関連付ける')])[1]",  # noqa
                )
            )
        ).click()
        time.sleep(1)

        # (//*[@class="checkbox__indicator"])[3]
        # (//*[@class='checkbox__label' and contains(text(),'このメールと同じ親番号のメールをすべて関連付ける')])[1]

        # driver.execute_script("arguments[1].click();", kanrenzuke)
        logging.info("Clicked one checkbox below kanrenzuke successfully ")

    except Exception as e:
        logging.error(f"Error when clicking on checkbox below kanrenzuke\n{e}")
        return False


def kanren(bangou):
    driver.switch_to.default_content()
    main_frame = driver.find_element(By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Mainmenu"])
    driver.switch_to.frame(main_frame)
    time.sleep(0.5)
    try:
        # click on blue tick
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Tickmark"]))
        ).click()
        time.sleep(1)

        # select the input box and send ankenbango
        input_box = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, "fMatterID")))
        input_box.click()
        input_box.send_keys(bangou)
        time.sleep(0.5)

        # checking how many places have anken_id text
        anken_id_xpath = driver.find_elements(By.XPATH, "//*[contains(text(),'案件ID')]")
        anken_id = len(anken_id_xpath)
        logging.info(f"Len of anken_id is: {anken_id}")

        if anken_id == 1:
            # (//*[@class="checkbox__indicator"])[3]
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located(
                        (
                            By.XPATH,
                            "(//*[@class='checkbox__label' and contains(text(),'このメールと同じ親番号のメールをすべて関連付ける')])[1]",  # noqa
                        )
                    )
                ).click()
                logging.info("Clicked on checkbox below kanrenzuke")
            except Exception as e:
                logging.info(f"Error when clicking on checkbox below kanrenzuke, error is:\n{e}")
        else:
            logging.info("only 1 anken id_found")
        time.sleep(0.5)

        # Click on the kanrenzuke button
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Kanren_関連付ける"])
            )
        ).click()
        time.sleep(1)

        # driver.execute_script("arguments[1].click();", kanrenzuke)
        logging.info("案件紐付け Successful")

    except Exception as e:
        logging.error(f"Error when doing 案件紐付け.....\n{e}")
        return False


def ankenka(excellinenumber, ankenmei, builder_name):
    driver.switch_to.window(driver.window_handles[0])

    button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, maildealer_config["Maildealer_ankenka"]["Maildealer_three_dot"]))
    )
    button.click()

    案件管理 = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, maildealer_config["Maildealer_ankenka"]["Maildealer_案件管理"]))
    )
    案件管理.click()

    time.sleep(2)

    driver.switch_to.default_content()
    # check anken register or not
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(0.5)

    # register anken
    driver.find_element(By.XPATH, maildealer_config["Maildealer_ankenka"]["案件を登録する_button"]).click()
    logging.info("clicked on 案件を登録する")
    time.sleep(3)
    WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, maildealer_config["Maildealer_ankenka"]["input_builder"]))
    ).send_keys({builder_name})
    logging.info("Builder sent")
    WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, maildealer_config["Maildealer_ankenka"]["input_ankenmei"]))
    ).send_keys(ankenmei)
    logging.info("Ankenmei sent")
    time.sleep(2)

    # Click on 登録 button
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, maildealer_config["Maildealer_ankenka"]["登録_button"]))
    ).click()
    logging.info("Clicked on 登録 button")
    time.sleep(1)

    # copy anken bangou
    anken_bangou = driver.find_element(By.XPATH, maildealer_config["Maildealer_ankenka"]["copy_anken_bangou"]).text
    time.sleep(0.5)
    sheet[f"B{excellinenumber}"].value = anken_bangou
    wb.save(excelfile)
    logging.info(f"Wrote 案件番号:{anken_bangou} to excel and saved successfully")
    time.sleep(0.5)

    driver.close()
    time.sleep(0.5)
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(1)
    return anken_bangou


def webaccess_check(koujibangou):
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(0.5)

    # click on reset
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["リセット"]))
    ).click()
    time.sleep(2)

    # remove the date
    remove_date = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["確定納品日_1"]))
    )
    remove_date.clear()
    remove_date.send_keys(Keys.RETURN)
    time.sleep(0.5)

    # send kouji bangou
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["工事番号"]))
    ).send_keys(koujibangou)
    time.sleep(0.5)

    # click on search
    search = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["検索"]))
    )
    driver.execute_script("window.scrollTo(1, 1);")
    time.sleep(0.5)
    search.click()
    time.sleep(2)

    try:
        # check if any anken is there or not after seraching
        WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["参照"]))
        )
        logging.info("Anken already registered, moving to the next anken")
        return True
    except Exception:
        logging.info("参照 element not found, moving to the register the anken in mail dealer")
        return False


def clear_cells(excelfile):

    try:
        # Create an instance of Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Keep Excel hidden

        # Suppress Save As dialog or alerts
        excel.DisplayAlerts = False

        # Open the workbook as read-only
        workbook = excel.Workbooks.Open(excelfile)

        logging.info("案件化 file opened and now running the macros")

        # Run the macro
        workbook.Sheets("シート")
        time.sleep(0.5)
        excel.Application.Run("ClearSheet1Data")
        time.sleep(2)
        logging.info("Ran ClearSheet1Data to clear cells from previous run")

        # Close the workbook
        workbook.Close(SaveChanges=True)

    except Exception as e:
        logging.error(f"Error running ClearSheet1Data macro: {e}")

    finally:
        # Quit Excel and clean up
        excel.Quit()
        del excel
        logging.info("Excel closed successfully")
        time.sleep(1)


def address_macro(excelfile):
    try:
        # Create an instance of Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Keep Excel hidden

        # Suppress Save As dialog or alerts
        excel.DisplayAlerts = False

        # Open the workbook as read-only
        workbook = excel.Workbooks.Open(excelfile)

        logging.info("案件化 file opened and now running the macros")

        # Run the macro
        workbook.Sheets("シート")
        time.sleep(0.5)
        excel.Application.Run("ExtractAndUpdate")
        time.sleep(2)
        logging.info("Ran ExtractAndUpdate for Builder name and Builder code")

        # Close the workbook
        workbook.Close(SaveChanges=True)

    except Exception as e:
        logging.error(f"Error running the macro: {e}")

    finally:
        # Quit Excel and clean up
        excel.Quit()
        del excel
        logging.info("Excel closed successfully")
        time.sleep(1)


def address(excellinenumber):
    time.sleep(0.5)
    driver.switch_to.window(driver.window_handles[-1])
    logging.info("switched to Andpad")
    time.sleep(1)

    # get address
    # type of address:
    # type1: 福岡県筑後市大字長浜楮原1522番2
    # type2: 〒866-0813 熊本県八代市上片町1675-1
    ######## put conditions for both type of address ##########

    try:
        xpath_ad = f"{andpad_config["xpaths"]["andpad_概要_xpaths"]["address"]}"
        print(f"address_xpath is: {xpath_ad}")
        juusho = (
            WebDriverWait(driver, 30)
            .until(
                EC.visibility_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_概要_xpaths"]["address"]))
            )
            .text
        )
        time.sleep(0.5)
        logging.info("juusho extracted from andpad")

        address_1 = juusho
        logging.info(f"address is: {address_1}")

        # will take this address in case some error after this
        new_address = address_1

        try:
            address_2 = address_1.split(" ")[1]
            logging.info(f"address2 is: {address_2}")
            for i in range(len(address_2)):
                if address_2[i].isdigit():
                    new_address = address_2[:i]
                    break
        except Exception as e:
            logging.error(e)
            logging.info("different address type, try method 2")

            for i in range(len(address_1)):
                if address_1[i].isdigit():
                    new_address = address_1[:i]
                    break

        time.sleep(1)
        logging.info(f"Final address is: {new_address}")
        sheet[f"K{excellinenumber}"].value = new_address
        logging.info(f"Wrote {new_address}")
        wb.save(excelfile)
        time.sleep(1)

    except Exception as e:
        logging.info(f"Error occurred in getting address: {e}")
        # sheet[f'K{excellinenumber}'].value = "NG, 住所読み取る時エラー"


# //div/p[@class="single-header__title"] ## oya name yaha se fetch kr
# //table/tr/td[@class="table-row__date--first-planned"][1] ## koujibangou ke liye
# //div/a[contains(text(),'oya_folder_name')] ## oya page pe jane ke liye


def process_mail(excellinenumber):
    try:
        extract_koujibangou = (
            WebDriverWait(driver, 10)
            .until(
                EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_概要_xpaths"]["案件管理ID"]))
            )
            .text
        )
        logging.info(f"Fetched koujibangou:\n{extract_koujibangou}")
        koujibangou = extract_koujibangou
        sheet[f"V{excellinenumber}"].value = koujibangou
        logging.info(f"Wrote {koujibangou}")
        wb.save(excelfile)
        time.sleep(1)
    except Exception as e:
        logging.info(f"Error fetching koujibangou, error is:\n{e}")
        sheet[f"V{excellinenumber}"].value = "NG, 工事番号読み取る時エラー"
        wb.save(excelfile)
        # excellinenumber += 1
        return False

    # Perform web access check
    # try:
    # if not webaccess_check(koujibangou):
    # if webaccess_check(koujibangou) == True:
    #     sheet[f'B{excellinenumber}'].value = "NG、アクセス登録有"
    #     logging.info(f"Wrote NG、アクセス登録有")
    #     wb.save(excelfile)
    #     time.sleep(1)
    #     # excellinenumber += 1
    #     return False
    # elif webaccess_check(koujibangou) == None:
    #     sheet[f'B{excellinenumber}'].value = "NG、アクセスエラー"
    #     logging.info(f"Wrote NG、アクセスエラー")
    #     wb.save(excelfile)
    #     time.sleep(1)
    #     return False
    # elif webaccess_check(koujibangou) == False:
    #     logging.info(f"Anken not found in access, fetching info from Andpad & registering it now")
    # else:
    #     logging.info(f"(just else) Anken not found in access, fetching info from Andpad & registering it now")

    access_result = webaccess_check(koujibangou)  # Call the function only once

    if access_result == True:  # noqa
        sheet[f"B{excellinenumber}"].value = "NG、アクセス登録有"
        logging.info("Wrote NG、アクセス登録有")
        wb.save(excelfile)
        time.sleep(1)
        return False
    elif access_result is None:  # Check if the result is None # noqa
        sheet[f"B{excellinenumber}"].value = "NG、アクセスエラー"
        logging.info("Wrote NG、アクセスエラー")
        wb.save(excelfile)
        time.sleep(1)
        return False
    elif access_result == False:  # noqa
        logging.info("Anken not found in access, fetching info from Andpad & registering it now")
    else:
        logging.error(f"Unexpected result from webaccess_check: {result}")

    # if not address(excellinenumber):
    try:
        address(excellinenumber)
    except Exception:
        logging.info("Error when fetching address")
        # Click on the folder
        sheet[f"K{excellinenumber}"].value = "NG, 住所読み取る時エラー"
        wb.save(excelfile)
        # excellinenumber += 1
        return False

    try:
        extract_ankenmei = (
            WebDriverWait(driver, 10)
            .until(
                EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_概要_xpaths"]["anken_text"]))
            )
            .text
        )
        logging.info(f"Fetched Ankenmei:\n{extract_ankenmei}")
        # ankenmei = extract_ankenmei.split(" 　")[1]

        if "土地売り" in extract_ankenmei:
            logging.info("Found 土地売り in Ankenmei, SKIPPING....")
            sheet[f"D{excellinenumber}"].value = "NG, 土地売り...案件"
            wb.save(excelfile)
            return False

        match = re.search(r"\d{2,}-\d{1,}\s+[　](.+)", extract_ankenmei)
        match_2 = re.search(r"^\d{2}-\d{4}-\d{4}　(.+?様邸|.+)", extract_ankenmei)
        if match:
            logging.info("Ankemei format is match")
            ankenmei = match.group(1)
            logging.info(f"New ankenmei is: {ankenmei}")
            ankenmei_new = ankenmei
            try:
                ankenmei_new = transform_value(ankenmei)
                logging.info(f"Transformed ankenmei is: {ankenmei_new}")
            except Exception as e:
                logging.error(e)
                logging.info("Error in transform_value(value)")
        elif match_2:
            logging.info("Ankemei format is match_2")
            ankenmei_new = match_2.group(1)
            logging.info(f"New ankenmei is: {ankenmei_new}")
        else:
            logging.info("Ankenmei format didn't match regex pattern, moving to next anken ")
            sheet[f"D{excellinenumber}"].value = "NG, 案件名フォマーとエラー"
            wb.save(excelfile)
            return False

        time.sleep(0.5)
        sheet[f"D{excellinenumber}"].value = ankenmei_new
        logging.info(f"Wrote {ankenmei_new}")
        wb.save(excelfile)
        time.sleep(1)
    except Exception as e:
        logging.info(f"Error fetching ankenmei, error is:\n{e}")
        sheet[f"D{excellinenumber}"].value = "NG, 案件名読み取る時エラー"
        wb.save(excelfile)
        # excellinenumber += 1
        return False

    try:
        address_macro(macro_file)
    except Exception as e:
        logging.info(f"Error in address_macro, error is: {e}")

    # input('a')
    try:
        read_range = f"C{excellinenumber}"
        logging.info(f"Read_range is: {read_range}")
        builder_name = read_data_from_excel(excelfile, read_range)
        logging.info(f"Builder name fetched from excel is: {builder_name}")
        # if builder_name == "No Match" or builder_name == "一建設(登録無し)":
        if builder_name == "確認必要" or builder_name == "住所無し":
            logging.info("Builder name issue, process manually")
            return False
    except Exception as e:
        logging.info(f"Error in read_data_from_excel, error is:\n{e}")
        return False

    # Perform ankenka
    try:
        anken_bangou = ankenka(excellinenumber, ankenmei_new, builder_name)
    except Exception as e:
        logging.info(f"Error in ankenka, error is:\n{e}")
        sheet[f"B{excellinenumber}"].value = "NG,案件化エラー"
        wb.save(excelfile)
        # excellinenumber += 1
        return False

    try:
        kanren(anken_bangou)
    except Exception as e:
        logging.info(f"Error in 案件紐付け, error is:\n{e}")
        sheet[f"G{excellinenumber}"].value = "NG,案件紐付けエラー"
        wb.save(excelfile)
        time.sleep(0.5)
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Kanren_popupclose"])
                )
            ).click()
        except Exception as e:
            logging.info("Anken kanren pop up not found")

    wb.save(excelfile)
    time.sleep(0.5)
    try:
        fileUpload(folder_path, ankenmei_new, anken_bangou, builder_name)
    except Exception as e:
        logging.info(f"Error in fileUpload, error is:\n{e}")

    return True


def oya(excellinenumber):
    try:
        topp = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_トップ"]))
        )
        topp.click()
        time.sleep(3)
        logging.info("Found and clicked on トップ\nPattern is oya, calling BAAP!!")

        try:
            extract_oya_name = (
                WebDriverWait(driver, 10)
                .until(
                    EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_project_page_heading"]))
                )
                .text
            )
            logging.info(f"Fetched Oya Name:{extract_oya_name}")
            oya_name = extract_oya_name
            logging.info(f"oya name is: {oya_name}")
            time.sleep(1)
        except Exception as e:
            logging.info(f"Error in extract_oya_name, error is:\n{e}")
            return excellinenumber, False

        try:
            table = andpad_config["xpaths"]["andpad_トップ_tr"]
            logging.info(f"table is: {table}")
            time.sleep(2)
            # Find all elements matching the XPath
            elements = driver.find_elements(By.XPATH, table)
            time.sleep(0.5)

            # Check the length of elements found
            num_elements = len(elements)
            logging.info(f"Number of elements found: {num_elements}")

            for element_number in range(num_elements):
                try:
                    # Click the element dynamically using its position in the list
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located(
                            (By.XPATH, f"({andpad_config['xpaths']['andpad_トップ_tr']})[{element_number + 1}]")
                        )
                    ).click()
                    logging.info("fetching no of tr's")

                    result = process_mail(excellinenumber)
                    logging.info(f"process_mail result is:{result}")
                    if not result:
                        excellinenumber += 1
                        logging.info(f"BAAP excellinenumber when False:{excellinenumber}")
                        driver.switch_to.window(driver.window_handles[-1])
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, andpad_config["xpaths"]["anpad_project_heading"].format(keyword=oya_name))
                            )
                        ).click()
                        logging.info("Clicked on oya folder")
                        time.sleep(2)
                        continue

                    # Increment after successful processing
                    excellinenumber += 1
                    logging.info(f"BAAP excellinenumber when True:{excellinenumber}")
                    driver.switch_to.window(driver.window_handles[-1])
                    time.sleep(0.5)
                    # Click on the folder
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, andpad_config["xpaths"]["anpad_project_heading"].format(keyword=oya_name))
                        )
                    ).click()
                    logging.info("Clicked on oya folder")
                    time.sleep(2)
                except Exception as e:
                    logging.info(f"Error processing element number {element_number}, error is:\n{e}")
                    if element_number == num_elements - 1:
                        logging.info("Element number = num_elements1")
                        sheet[f"B{excellinenumber}"].value = "NG, アクセス検査エラー"
                        logging.info("Wrote NG, アクセス検査エラー")
                        wb.save(excelfile)
                        excellinenumber += 1
                        driver.switch_to.window(driver.window_handles[-1])
                        driver.close()
                        time.sleep(0.5)
                        return excellinenumber, False

            if element_number == num_elements - 1:
                # if True in result:
                if result:
                    logging.info("Element number = num_elements2")
                    driver.switch_to.window(driver.window_handles[-1])
                    driver.close()
                    time.sleep(0.5)
                    return excellinenumber, True
                else:
                    logging.info("Element number = num_elements3")
                    driver.switch_to.window(driver.window_handles[-1])
                    driver.close()
                    time.sleep(0.5)
                    return excellinenumber, False
                # return excellinenumber, 'all_elements_read'

        except Exception as e:
            logging.info(f"Error in last Exception of oya, error is:\n{e}")
            driver.switch_to.window(driver.window_handles[-1])
            driver.close()
            time.sleep(0.5)
            excellinenumber += 1
            return excellinenumber, False  ###### added this

    except Exception as e:
        logging.error(e)
        logging.info("Pattern is kodomo, calling BACCHA!!")
        gaiyo = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, andpad_config["xpaths"]["andpad_概要"]))
        )
        gaiyo.click()
        time.sleep(3)
        logging.info("Found and clicked on 概要")
        result = process_mail(excellinenumber)
        if not result:
            logging.info(f"Baccha excellinenumber when False:{excellinenumber}")
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(0.5)
            excellinenumber += 1
            driver.close()
            time.sleep(1)
            return excellinenumber, False
        time.sleep(0.5)
        # Increment after successful processing
        driver.switch_to.window(driver.window_handles[-1])
        driver.close()
        time.sleep(0.5)
        excellinenumber += 1
        logging.info(f"Baccha excellinenumber when True:{excellinenumber}")
    time.sleep(1)
    return excellinenumber, True


def read_data_from_excel(file_path, read_range):
    # Read updated data from Excel
    wb = openpyxl.load_workbook(file_path)
    # sheet = wb.active  # Adjust to your sheet name if needed
    sheet = wb["シート"]
    data = sheet[read_range].value  # Read the cell value
    wb.close()
    return data


def move_to_folder(result):

    # switch back to the main tab (mail dealer tab)
    driver.switch_to.window(driver.window_handles[0])

    driver.switch_to.default_content()
    mainmenu = driver.find_element("id", "ifmMain")
    driver.switch_to.frame(mainmenu)
    time.sleep(3)
    print(f"Result in move_to_folder: {result}")

    # click on 1stfolder box
    # driver.find_element("xpath", '//*[@id='form-olv-p-viewmail']/div[1]/div[2]/div[1]/div/div[1]/div").click()
    driver.find_element(By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Folder"]).click()
    logging.info("folder box clicked")
    time.sleep(0.5)

    # input('a')

    # select the folder box (whole list)
    WebDriverWait(driver, 10).until(
        EC.visibility_of_all_elements_located(
            (By.XPATH, "/html/body/div/div/form/div[1]/div[2]/div[1]/div/div[1]/div[2]/ul")
        )
    )
    WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.XPATH, "//*[@class='list has-scroll']")))

    zumenshiryou = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, maildealer_config["MailDealer_xpaths"]["folder_list"].format(keyword="図面資料待"))
        )
    )
    zumenshiryou.click()
    # driver.find_element(By.XPATH,"//*[contains(text(), '図面対応済<ﾍﾞﾄﾅﾑ>')]").click()
    logging.info("successfully moved to 図面資料待 folder")
    time.sleep(4)  ## give some time after sending click, so that folder can move

    try:
        # click on tantousha box
        tantousha_box = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//*[@class='dropdown is-text is-focus-visible'][2]"))
        )
        tantousha_box.click()
        time.sleep(1)

        # select 自分/jiubn
        jibun = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//*[@class='list__item is-bold has-value']"))
        )
        jibun.click()
        time.sleep(4)
        logging.info("successfully put 自分/jibun label")
    except Exception as e:
        logging.error(e)
        logging.info("Error selecting 自分/jibun, probably already selected")

    # if result == "True":
    # if True in result:
    if result is True:
        try:
            driver.find_element(By.XPATH, "//*[@class='olv-c-dropdown olv-c-dropdown--multi-line']").click()
            logging.info("folder box clicked")
            time.sleep(0.5)

            WebDriverWait(driver, 10).until(
                EC.visibility_of_all_elements_located((By.XPATH, "//*[@class='list has-scroll']"))
            )

            # select the 案件化済
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Label_案件化済"])
                )
            ).click()
            logging.info("successfully put 案件化済 label")
            time.sleep(4)  ## give some time after sending click, so that folder can move
        except Exception as e:
            logging.error(e)
            logging.info("Error selecting 案件化済 label")
    # input('a')


def fileUpload(folder_path, ankenmei, ProjectNumber, builder):

    folder_name = f"{ProjectNumber} {ankenmei}"
    shiryou = os.path.join(folder_path, folder_name, "資料")

    # create shirou folder
    create_download_directory(shiryou)

    ###### Creating a copy of text file to shiryou folder #####
    if not os.path.exists(f"{shiryou}/{text_file}"):
        try:
            shutil.copy(f"{text_file}", f"{shiryou}/{text_file}")
            logging.info("text File copy created in 資料 Folder")
        except Exception as e:
            logging.error(e)
            logging.info("text File not found")
    # input('a')

    upload_path = os.path.join(os.getcwd(), folder_path, folder_name)
    logging.info(f"Upload path: {upload_path}")

    # open a new tab
    driver.execute_script("window.open('about:blank','_blank');")

    # Switch to the newly opened tab
    driver.switch_to.window(driver.window_handles[-1])
    # driver.switch_to.window(driver.window_handles[2])
    time.sleep(0.5)

    if builder == "一建設(関西)":
        logging.info(f"Builder is: {builder}")
        driver.get(kansai)
    elif builder == "一建設(中国・四国)":
        logging.info(f"Builder is: {builder}")
        driver.get(shikoku)
        time.sleep(5)
    elif builder == "一建設(東海)":
        logging.info(f"Builder is: {builder}")
        driver.get(toukai)
        time.sleep(5)
    elif builder == "一建設(関東)":
        logging.info(f"Builder is: {builder}")
        driver.get(kantou)
        time.sleep(5)
    elif builder == "一建設(九州)":
        logging.info(f"Builder is: {builder}")
        driver.get(kyuushuu)
        time.sleep(5)
    elif builder == "一建設(郡山)":
        logging.info(f"Builder is: {builder}")
        driver.get(koriyama)
        time.sleep(5)
    elif builder == "一建設(仙台)":
        logging.info(f"Builder is: {builder}")
        driver.get(sendai)
        time.sleep(5)

    for win in Desktop(backend="win32").windows():
        if win.window_text() == "Select Folder to Upload":
            win.post_message(0x0010)
            break
    time.sleep(2)
    try:
        # Click on upload
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, sharepoint_config["xpaths"]["upload"]))
        ).click()
        time.sleep(2)

        # Click on folder
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, sharepoint_config["xpaths"]["folder"]))
        ).click()
        time.sleep(2)

        for win in Desktop(backend="win32").windows():
            if win.window_text() == "Select Folder to Upload":
                win.post_message(0x0010)
                break
        time.sleep(2)

        # Upload
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
        ).send_keys(upload_path)
        time.sleep(15)

    except Exception as e:
        logging.error(f"Upload probably failed: {e}")

    driver.close()
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[0])
    logging.info("switched to mail dealer")
    time.sleep(1)


# Set Chrome options for downloading files to the specified folder
chrome_options = Options()

prefs = {
    "credentials_enable_service": False,
    "profile.password_manager_enabled": False,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": False,
}
chrome_options.add_experimental_option("prefs", prefs)

chrome_options.add_argument("--guest")  # prevents profile sync
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-notifications")

driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
time.sleep(0.5)


try:
    # window 0
    mail_dealer_login()
    time.sleep(1)
except Exception as e:
    logging.info(f"Error in mail_dealer_login, error is:\n{e}")

try:
    # window 1
    webaccess_login()
except Exception as e:
    logging.info(f"Error in webaccess_login, error is:\n{e}")

try:
    # window 2
    Andpad_login()
    time.sleep(1)
except Exception as e:
    logging.info(f"Error in Andapd_login, error is:\n{e}")

try:
    # window 3
    sharepoint_login()
    time.sleep(1)
except Exception as e:
    logging.info(f"Error in sharepoint_login, error is:\n{e}")

try:
    # clear previous data from the file
    clear_cells(macro_file)
    time.sleep(0.5)
except Exception as e:
    logging.info(f"Error in clear_cells, error is:\n{e}")

wb = load_workbook(excelfile, keep_vba=True)
# sheet = wb.active
sheet = wb["シート"]
time.sleep(1)
try:
    excellinenumber = 2

    try:
        # switch back to mail_dealer
        driver.switch_to.window(driver.window_handles[0])
        logging.info("switched back to mail dealer")
        time.sleep(2)
        driver.switch_to.default_content()
        main_frame = driver.find_element(By.XPATH, maildealer_config["MailDealer_xpaths"]["MailDealer_Mainmenu"])
        driver.switch_to.frame(main_frame)
        time.sleep(0.5)

        # try:
        # while True:
        total_mails = 300
        # total_mails = 20
        while total_mails > 0:

            try:  # xpath for the table having all the mails
                # mails  = driver.find_elements(By.XPATH,"//div[2]/table/tbody")
                mails = driver.find_elements(By.XPATH, maildealer_config["MailDealer_xpaths"]["All_Mails_table"])

                if not mails:
                    logging.info("No more mails to process.")
                    break

                # tantosha asked to make bot deal with those mails with label 資料のみ in 新着 tab
                # only process mails having Andpad in subject
                # remember that if there is tantosha written, ignore that mail
                # just deal with mails having 資料のみ label and tantosha --
                # Skip if 土地売り in ankenmei

                # for subject in mail:
                # for index, subject in enumerate(mails):
                for index in range(len(mails)):
                    try:
                        # Re-fetch the element in case of stale reference
                        # subject = driver.find_elements(By.XPATH, "//div[2]/table/tbody")[index]
                        subject = driver.find_elements(
                            By.XPATH, maildealer_config["MailDealer_xpaths"]["All_Mails_table"]
                        )[index]
                        mail_text = subject.text  # Extract text from the current mail
                        logging.info(f"Mail content: {mail_text}")
                        # input('a')

                        # mail_text = subject.text  # Extract text from the current mail
                        # logging.info(f'Mail content: {mail_text}')
                        # # input('a')

                        keywords = ["参加", "招待", "資料のみ"]
                        # if any(keyword in mail_text for keyword in keywords):
                        if any([keyword in mail_text for keyword in keywords]):
                            logging.info("Mail contains one of the keywords: 参加 or 招待 or 資料のみ")
                            time.sleep(3)

                            # Step 2: Check if the mail contains '資料のみ'
                            # if "資料のみ" in mail_text and "--" not in mail_text:
                            # if "資料のみ" in mail_text and "--" not in mail_text:
                            if re.search(r"\d+-\d+\s*資料のみ", mail_text) and "--" not in mail_text:
                                logging.info("Mail contains '資料のみ' + Tantousha  SKIPPING this mail...")
                                # input('a')
                                total_mails -= 1  # Decrement the mail counter
                                continue  # Skip to the next mail
                            elif (
                                re.search(r"\d+-\d+\s*資料のみ", mail_text)
                                and "--" in mail_text
                                and "【ANDPAD】" in mail_text
                            ):
                                logging.info("Mail contains '資料のみ', processing this mail")
                            elif ("参加" in mail_text or "招待" in mail_text) and "案件化済" not in mail_text:
                                logging.info("Mail contains '参加' or'招待', processing this mail")
                            else:
                                logging.info(
                                    "Mail DOES NOT contain just '資料のみ' or'参加' or'招待', SKIPPING this mail..."
                                )
                                # input('a')
                                total_mails -= 1  # Decrement the mail counter
                                continue  # Skip to the next mail

                            # input('a')
                            # continue

                            subject.click()
                            logging.info("Clicked on Mail")
                            time.sleep(3)
                            # click on the andpad link in the mail
                            WebDriverWait(driver, 10).until(
                                EC.visibility_of_element_located(
                                    (By.XPATH, maildealer_config["MailDealer_xpaths"]["Mailbody_1stlink"])
                                )
                            ).click()
                            logging.info("clicked on the 1st link in the mail")
                            time.sleep(1)
                            # switch to Andpad
                            driver.switch_to.window(driver.window_handles[-1])

                            if "お知らせ" in mail_text:
                                try:
                                    anken_shousai = "//a[@class='message-header__textlink' and text()='案件詳細へ']"
                                    anken_shousai = WebDriverWait(driver, 10).until(
                                        EC.presence_of_element_located(
                                            (By.XPATH, andpad_config["andpad_チャット_xpath"]["案件詳細へ2"])
                                        )
                                    )
                                    anken_shousai.click()
                                    time.sleep(3)
                                    logging.info("Found and clicked on 案件詳細へ")
                                except Exception as e:
                                    logging.info(f"Error in clicking or DIDN'T FIND 案件詳細へ, error is:\n{e}")

                            ## Before going to oya function, redirect to the top page or gaiyou page

                            try:
                                excellinenumber, result = oya(excellinenumber)
                                logging.info(f"oya result is:{result}")
                                # input('a')
                                logging.info(f"Excel line in after exiting oya function is: {excellinenumber}")
                                if not result:
                                    # if False in result:
                                    driver.switch_to.window(driver.window_handles[0])
                                    time.sleep(0.5)
                                    try:
                                        move_to_folder(result)
                                    except Exception as e:
                                        logging.info(f"Error in move_to_folder, error is:\n{e}")
                                    WebDriverWait(driver, 10).until(
                                        EC.presence_of_element_located(
                                            (By.XPATH, maildealer_config["MailDealer_xpaths"]["back_button"])
                                        )
                                    ).click()
                                    logging.info("Clicked on back button")
                                    time.sleep(3)
                                    total_mails -= 1  # Decrement the mail counter
                                    continue
                                # elif 'all_elements_read' in result:
                                #     logging.info(f"all_elements_read, moving to the next step")

                                time.sleep(0.5)
                                try:
                                    move_to_folder(result)
                                except Exception as e:
                                    logging.info(f"Error in move_to_folder, error is:\n{e}")
                                time.sleep(0.5)
                                WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, maildealer_config["MailDealer_xpaths"]["back_button"])
                                    )
                                ).click()
                                logging.info("Clicked on back button")
                                # input('a')
                                time.sleep(4)
                                total_mails -= 1  # Decrement the mail counter
                                # input('a')
                            except Exception as e:
                                logging.info(f"Error in oya, error is:\n{e}")
                    except StaleElementReferenceException as e:
                        logging.warning(f"Stale element encountered, re-fetching mail: {e}")
                        total_mails -= 1
                        continue

                    else:
                        logging.info("'参加' or '招待' not found in the mail, moving to the next mail")
                        total_mails -= 1  # Decrement the mail counter
                        continue
            except Exception as e:
                logging.info(f"Error in opening mail, error is:\n{e}")
    except Exception as e:
        logging.info(f"Error in 2nd opening mail, error is:\n{e}")

    driver.switch_to.window(driver.window_handles[0])

except Exception as e:
    logging.info(f"Error in very last exception, error is:\n{e}")
finally:
    driver.quit()
    logging.info("ALL MAILS READ!!!!")

try:
    address_macro(macro_file)
    logging.info("last instance of address_macro ran successfuly")

except Exception as e:
    logging.info(f"Error in running last instance of address_macro, error is: {e}")
