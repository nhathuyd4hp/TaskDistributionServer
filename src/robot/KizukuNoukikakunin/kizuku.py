import logging
import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    ElementNotInteractableException,
    NoSuchElementException,
    NoSuchWindowException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Setup logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logging.info("Loading Excel data...")

# Load Excel data
file_path = r"kizuku 納期確認.xlsx"
workbook = load_workbook(filename=file_path)
sheet = workbook["Sheet1"]


# Extract data and filter out rows with missing values
data = [
    (
        sheet[f"B{row}"].value,
        sheet[f"C{row}"].value,
        sheet[f"D{row}"].value,
        sheet[f"F{row}"].value,
        sheet[f"G{row}"].value,
        sheet[f"H{row}"].value,
    )
    for row in range(2, sheet.max_row + 1)
    if sheet[f"B{row}"].value
    and sheet[f"C{row}"].value
    and sheet[f"D{row}"].value
    and sheet[f"F{row}"].value
    and sheet[f"G{row}"].value
    and sheet[f"H{row}"].value is not None
]

# Separate the filtered data into individual lists
ビルダー名, 案件名, 納期, 担当者, 担当者2, 不足 = zip(*data) if data else ([], [], [], [], [], [])

logging.info("Data loaded successfully")
logging.info(f"ビルダー名: {ビルダー名}")
logging.info(f"案件名: {案件名}")
logging.info(f"納期: {納期}")
logging.info(f"担当者: {担当者}")
logging.info(f"担当者2: {担当者2}")
logging.info(f"不足: {不足}")

logging.info("Starting iteration over the extracted data...")


def safe_interact_with_element(driver, by, value, interaction, retries=3):
    """
    Helper function to safely interact with an element using retries.
    """
    for attempt in range(retries):
        try:
            element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((by, value)))
            driver.execute_script("arguments[0].scrollIntoView(true);", element)
            interaction(element)
            return True
        except (
            ElementClickInterceptedException,
            ElementNotInteractableException,
            TimeoutException,
            StaleElementReferenceException,
        ) as e:
            logging.warning(f"Attempt {attempt + 1} failed: {e}")
            if isinstance(e, StaleElementReferenceException):
                continue  # Re-find the element on the next attempt
            time.sleep(1)  # Wait briefly before retrying
    logging.error(f"Failed to interact with element by {by} using value {value} after {retries} retries.")
    return False


def safe_click_with_js(driver, by, value):
    """
    Function to click an element using JavaScript.
    """
    try:
        element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((by, value)))
        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        driver.execute_script("arguments[0].click();", element)
        logging.info(f"Clicked element by {by} using JavaScript")
        return True
    except TimeoutException as e:
        logging.error(f"Failed to click element by {by} using JavaScript: {e}")
        return False


# Iterate over the extracted data and perform actions
for row_index, (案件, ビルダー, 日付, 担当, 担当者2, 不足) in enumerate(  # noqa
    zip(案件名, ビルダー名, 納期, 担当者, 担当者2, 不足), start=2
):
    # logging.info(f"Processing builder: {ビルダー}")
    try:

        logging.info(f"Setting up WebDriver for builder: {案件}")
        # Webdriver setup
        kizukuurl_webaccess = "https://kizuku2.ctx.co.jp/"
        chrome_options = Options()
        chrome_options.add_argument("--disable-notifications")
        driver = webdriver.Chrome(options=chrome_options)
        driver.maximize_window()
        driver.get(kizukuurl_webaccess)

        # Wait setup
        wait = WebDriverWait(driver, 15)
        logging.info(f"Logging in for builder: {案件}")
        # Login process based on ビルダー value
        login_id = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[2]/form/div[1]/input"))
        )
        password_id = driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[2]/form/div[2]/input")

        if ビルダー in ["ケイアイスター不動産", "ケイアイスター不動産(準耐火)", "旭ハウジング"]:
            login_id.send_keys("keiai@nsk-cad.com")
            password_id.send_keys("nskkantou")
        elif ビルダー == "ケイアイプランニング株式会社":
            login_id.send_keys("kip@nsk-cad.com")
            password_id.send_keys("343nqmun")

        driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[2]/form/input[4]").click()
        logging.info("Login form submitted")
        wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "/html/body/div/div[1]/div/section/div[1]/div[1]/div/div[1]/div[1]/div/button[1]")
            )
        )
        logging.info("Page has loaded completely")

        # Open a new tab
        driver.execute_script("window.open('');")

        # Switch to the new tab (it will be the last in the list)
        driver.switch_to.window(driver.window_handles[-1])
        # Navigate to the specific URL after login
        talk_list_url = "https://kizuku2.ctx.co.jp/plans/bukken/index"
        driver.get(talk_list_url)
        logging.info("Navigated to PlanBukken")
        wait = WebDriverWait(driver, 15)

        # Ensure the loader or modal is not present
        try:
            wait.until(EC.invisibility_of_element((By.CSS_SELECTOR, "div.modal-loader")))
        except TimeoutException:
            logging.warning("Loader modal did not disappear in time.")

        # Improved interaction logic
        if not safe_interact_with_element(
            driver, By.XPATH, "/html/body/div/div[1]/div/section/div[1]/div/div[2]/div/div/button", lambda e: e.click()
        ):
            logging.info("Retrying with JavaScript click")
            if safe_click_with_js(
                driver, By.XPATH, "/html/body/div/div[1]/div/section/div[1]/div/div[2]/div/div/button"
            ):
                logging.info("Search button clicked using JavaScript")
            else:
                logging.error(f"Failed to click the search button for {ビルダー} even with JavaScript")
                sheet[f"I{row_index}"] = "Failed to click search button"
                workbook.save(file_path)

        # Interact with the input field to send 案件
        if safe_interact_with_element(
            driver,
            By.XPATH,
            "/html/body/div[1]/div[1]/div/section/div[3]/div/div/form/div/div[2]/div/input",
            lambda e: e.send_keys(案件),  # noqa
        ):
            logging.info(f"Entered 案件名 for {ビルダー}: {案件}")
            # Click the search button after entering 案件
            if safe_interact_with_element(
                driver,
                By.XPATH,
                "/html/body/div[1]/div[1]/div/section/div[3]/div/div/div[2]/button[3]",
                lambda e: e.click(),
            ):
                logging.info(f"Clicked the search button after entering 案件名 for {ビルダー}")
                # Wait for the search results to appear and click the specific 案件
                if safe_interact_with_element(
                    driver,
                    By.XPATH,
                    "/html/body/div/div[1]/div/section/div[2]/table/tbody/tr/td[2]/div/button[2]/img",
                    lambda e: e.click(),
                ):
                    logging.info(f"Clicked on 案件 in the search results for {ビルダー}")
                    # Click the send button to choose the tantosha
                    attention_button = "/html/body/div[2]/div/div[1]/div/div/div[1]/div[5]/div[1]/ul/li[6]"
                    if safe_interact_with_element(driver, By.XPATH, attention_button, lambda e: e.click()):
                        logging.info("Clicked the attention button")
                    else:
                        logging.error("Failed to click the attention button")
                        sheet[f"I{row_index}"] = "Failed to click the attention button"
                        workbook.save(file_path)

                    # Send the value of "担当者"
                    if safe_interact_with_element(
                        driver,
                        By.XPATH,
                        "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[1]/input",
                        lambda e: e.send_keys(担当),  # noqa
                    ):
                        logging.info(f"Sent the value of 担当者: {担当}")
                        # Click the associated button
                        if safe_interact_with_element(
                            driver,
                            By.XPATH,
                            "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[1]/span/button",
                            lambda e: e.click(),
                        ):
                            logging.info("Clicked the button after sending 担当者 value")

                            # Try to click the checkbox for the result
                            checkbox_clicked = safe_interact_with_element(
                                driver,
                                By.XPATH,
                                "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[6]/div/table/tbody/tr/td/div[1]",
                                lambda e: e.click(),
                            )

                            if checkbox_clicked:
                                logging.info("Clicked the checkbox for the result")
                            else:
                                logging.info("No checkbox found, proceeding to the next step")

                            # Click to choose the tantosha
                            if safe_interact_with_element(
                                driver,
                                By.XPATH,
                                "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[4]/button",
                                lambda e: e.click(),
                            ):
                                logging.info("Chose the tantosha")
                            else:
                                logging.error("Failed to click the tantosha")
                                sheet[f"I{row_index}"] = "Failed to choose the tantosha"
                                workbook.save(file_path)
                        else:
                            logging.error("Failed to click the button after sending 担当者 value")
                            sheet[f"I{row_index}"] = "Failed to click button after sending 担当者"
                            workbook.save(file_path)
                    else:
                        logging.error("Failed to send the value of 担当者")
                        sheet[f"I{row_index}"] = "Failed to send 担当者 value"
                        workbook.save(file_path)
                else:
                    logging.error(f"Failed to click on 案件 in the search results for {ビルダー}")
                    sheet[f"I{row_index}"] = "Failed to click on 案件"
            else:
                logging.error(f"Failed to click the search button after entering 案件名 for {ビルダー}")
                sheet[f"I{row_index}"] = "Failed to click search button after entering 案件名"
        else:
            logging.error(f"Failed to enter 案件名 for {ビルダー}: {案件}")
            sheet[f"I{row_index}"] = "Failed to enter 案件名"

        # If 担当者2 has a value, choose 担当者2
        if 担当者2 != "なし":
            if safe_interact_with_element(driver, By.XPATH, attention_button, lambda e: e.click()):
                logging.info("Clicked the attention button")
            else:
                logging.error("Failed to click the attention button")
                sheet[f"I{row_index}"] = "Failed to click the attention button"
                workbook.save(file_path)

            # Send the value of "担当者2"
            if safe_interact_with_element(
                driver,
                By.XPATH,
                "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[1]/input",
                lambda e: e.send_keys(担当者2),  # noqa
            ):
                logging.info(f"Sent the value of 担当者2: {担当者2}")
                # Click the associated button
                if safe_interact_with_element(
                    driver,
                    By.XPATH,
                    "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[1]/span/button",
                    lambda e: e.click(),
                ):
                    logging.info("Clicked the button after sending 担当者2 value")

                    # Try to click the checkbox for the result
                    checkbox_clicked = safe_interact_with_element(
                        driver,
                        By.XPATH,
                        "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[6]/div/table/tbody/tr/td/div[1]",
                        lambda e: e.click(),
                    )

                    if checkbox_clicked:
                        logging.info("Clicked the checkbox for the result")
                    else:
                        logging.info("No checkbox found, proceeding to the next step")

                    # Click to choose the 担当者2
                    if safe_interact_with_element(
                        driver,
                        By.XPATH,
                        "/html/body/div[2]/div/div[1]/div/div/div[3]/div[31]/div[2]/div/div[2]/div[4]/button",
                        lambda e: e.click(),
                    ):
                        logging.info("Chose the 担当者2")
                    else:
                        logging.error("Failed to click the 担当者2")
                        sheet[f"I{row_index}"] = "Failed to choose the 担当者2"
                        workbook.save(file_path)
                else:
                    logging.error("Failed to click the button after sending 担当者2 value")
                    sheet[f"I{row_index}"] = "Failed to click button after sending 担当者2"
                    workbook.save(file_path)
            else:
                logging.error("Failed to send the value of 担当者2")
                sheet[f"I{row_index}"] = "Failed to send 担当者2 value"
                workbook.save(file_path)

        # Send the appropriate message based on the value of 不足
        if 不足 != "なし":
            message = f"""いつも大変お世話になっております。
ご依頼頂いております、軽天材の納材日確認となります。

【納材日：{日付}】

★{不足}が不足しておりますので、大至急送付をお願い致します。

変更等御座いましたら、５日(営業日)前までにご連絡をお願い致します。
※納材日2日前(2営業日前)以降の納期変更に関しては別途費用を頂戴しております。
ご注意の上、納期をご確認下さい。

──────＜ご案内とご注意事項＞──────
▼中入れについて
建物内への納材はお受けしておりません。
予めご了承のほど宜しくお願い致します。

▼納材場所の確保について
本メールが届きましたら、納材場所をご検討いただき、
当日、大工さん不在の際は納材場所の確保をお願い致します。
───────────────────────
ご連絡の行違いが御座いましたら、お詫び申し上げます。
よろしくお願い致します。

エヌ・エス・ケー工業株式会社　"""
        else:
            message = f"""いつも大変お世話になっております。
ご依頼頂いております、軽天材の納材日確認となります。

【納材日：{日付}】

変更等御座いましたら、５日(営業日)前までにご連絡をお願い致します。
※納材日2日前(2営業日前)以降の納期変更に関しては別途費用を頂戴しております。
ご注意の上、納期をご確認下さい。

──────＜ご案内とご注意事項＞──────
▼中入れについて
建物内への納材はお受けしておりません。
予めご了承のほど宜しくお願い致します。

▼納材場所の確保について
本メールが届きましたら、納材場所をご検討いただき、
当日、大工さん不在の際は納材場所の確保をお願い致します。
───────────────────────
ご連絡の行違いが御座いましたら、お詫び申し上げます。
よろしくお願い致します。

エヌ・エス・ケー工業株式会社　"""

        # Embed the value of 担当者 at the beginning of the message if no checkbox was clicked
        if not checkbox_clicked:
            message = f"{担当}様\n" + message

        ## Send the message using JavaScript to avoid triggering the file input pop-up
        message_xpath = "/html/body/div[2]/div/div[1]/div/div/div[1]/div[4]/div/textarea"
        message2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, message_xpath)))
        message2.send_keys(message)

        send_button1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sendMsgBtn"]')))
        send_button1.click()
        logging.info("Clicked the send button")
        time.sleep(5)
        sheet[f"I{row_index}"] = "Message sent successfully"
        workbook.save(file_path)
        logging.info(f"Updated the sheet for row {row_index} with 'Message sent successfully'")

    except (NoSuchElementException, TimeoutException, NoSuchWindowException) as e:
        logging.error(f"An error occurred during processing builder: {ビルダー} - {e}")
        sheet[f"I{row_index}"] = f"Error during processing: {e}"
        workbook.save(file_path)

    finally:
        driver.quit()
        logging.info(f"WebDriver closed for builder: {ビルダー}")

logging.info("Iteration over the data completed.")
