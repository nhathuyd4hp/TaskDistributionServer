# === progress_reporter.py ===

import datetime
import logging
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# In-memory progress data
_progress_data = []


def clear_report():
    """Clear the progress memory."""
    global _progress_data
    _progress_data = []


def add_result(anken_number, builder_id, status):
    """Add a download result record."""
    _progress_data.append({"案件番号": anken_number, "ビルダーID": builder_id, "ステータス": status})


def _beautify_excel(output_path):
    """Beautify the Excel formatting."""
    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Freeze header
        ws.freeze_panes = "A2"

        # Bold Headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # get the column letter (e.g., A, B, C)

            for cell in column_cells:
                if cell.value:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))

            adjusted_width = (max_length + 2) * 1.2  # slightly bigger
            ws.column_dimensions[column].width = adjusted_width

        # Center align "ステータス" (Status Column - 3rd Column)
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Color Rows based on Status
        for row in ws.iter_rows(min_row=2):
            status = row[2].value  # 3rd column ("ステータス")
            if status == "Success":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green
            elif status == "Failed":
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red
            else:
                fill = None

            if fill:
                for cell in row:
                    cell.fill = fill

    wb.save(output_path)
    logging.info(f"🎨 Beautified Progress Report: {output_path}")


def save_report():
    """Save all collected progress to a beautiful timestamped Excel."""
    try:
        if not _progress_data:
            logging.warning("⚠️ No progress data to save!")
            return

        df = pd.DataFrame(_progress_data)

        # Separate sheets
        downloaded = df[df["ステータス"] == "Success"]
        failed = df[df["ステータス"] == "Failed"]

        progress_folder = os.path.join(os.getcwd(), "ProgressReports")
        os.makedirs(progress_folder, exist_ok=True)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        output_path = os.path.join(progress_folder, f"Progress_Report_{timestamp}.xlsx")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            downloaded.to_excel(writer, sheet_name="Downloaded", index=False)
            failed.to_excel(writer, sheet_name="Failed", index=False)

        _beautify_excel(output_path)

    except Exception as e:
        logging.error(f"❌ Failed to save Progress Report: {e}")


def get_report():
    """Return the in-memory progress data."""
    return _progress_data
