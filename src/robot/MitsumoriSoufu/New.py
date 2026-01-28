# === Final New.py for 割付図_見積書送付 ===
import argparse
import locale
import logging
import os
import shutil
import threading
import time
import tkinter as tk
from datetime import datetime
from tkinter import CENTER, Label

import pandas as pd
from config_access_token import token_file  # noqa
from graph_downloader import graph_download_and_save_files
from logging_setup import setup_logging
from Nasiwak import Bot_Update, create_json_config
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from tkcalendar import DateEntry

# Set the locale to Japanese
locale.setlocale(locale.LC_ALL, "ja_JP.UTF-8")

# Setup logging
setup_logging()


# === Main Class ===
class 割付図_見積書送付:
    version = "1.4"

    def __init__(self, from_date, to_date):
        self.from_date = from_date
        self.to_date = to_date

    def process_data(self):

        # token

        # Replace with your actual file path
        file_path = os.path.join(os.getcwd(), "Access_token", "Access_token.txt")
        # logging.info(f"file path for text file is: {file_path}")
        # Open and read the file
        with open(file_path, "r", encoding="utf-8") as file:
            content = file.read()
        logging.info(f"Extracted text from .txt file is: {content}")

        # ✨ Setup
        maildealer_json_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/MailDealer.json"
        webaccess_json_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/webaccess.json"
        access_token = content

        Maildealer_Data = create_json_config(maildealer_json_url, access_token)
        Webaccess_Data = create_json_config(webaccess_json_url, access_token)
        Bot_Update("Nasiwak", "", "v1.3", access_token)

        # ✨ WebDriver Setup
        chrome_options = Options()
        prefs = {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": False,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_argument("--guest")  # prevents profile sync
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        driver = webdriver.Chrome(options=chrome_options)
        driver.maximize_window()

        # 🧹 Step 1: Clear old files
        if os.path.exists("Ankens"):
            shutil.rmtree("Ankens")
        os.makedirs("Ankens", exist_ok=True)

        # 🛜 Step 2: Login to WebAccess
        driver.get(Webaccess_Data["webaccess_url"])
        self.webaccess_login(driver, Webaccess_Data)
        time.sleep(2)

        # 📥 Step 3: Search Ankengos
        self.access_search(driver, Webaccess_Data)
        time.sleep(2)

        # 🛜 Step 4: Login to MailDealer
        self.maildealer_login(driver, Maildealer_Data)
        time.sleep(2)

        # 📄 Step 5: Process CSV to Excel
        csv_files = [f for f in os.listdir() if f.endswith(".csv")]
        if not csv_files:
            logging.error("❌ No CSV found after download.")
            return
        csv_file = csv_files[0]
        excel_file = csv_file.replace(".csv", ".xlsx")
        pd.read_csv(csv_file, encoding="utf-8-sig").to_excel(excel_file, index=False, engine="openpyxl")
        os.remove(csv_file)

        # ✨ Clean only 住協建設㈱
        df = pd.read_excel(excel_file, dtype=str)
        df = df[df["得意先名"] == "住協建設㈱"]
        df.to_excel(excel_file, index=False, engine="openpyxl")
        logging.info(f"✅ Cleaned to only 住協建設㈱: {len(df)} rows.")

        # 📋 Step 6: Download PDFs
        案件番号_list = df["案件番号"].tolist()
        案件名_list = df["物件名"].tolist()
        ビルダー名_list = df["得意先名"].tolist()
        納期_list = df["確定納期"].tolist()

        wb = load_workbook(excel_file)
        sheet = wb.active
        line = 2

        for ankenbango, ankenname, buildername, 納期 in zip(案件番号_list, 案件名_list, ビルダー名_list, 納期_list):
            logging.info(f"📋 Processing {ankenbango} - {ankenname}")

            # 📥 1. Download 割付図・見積 PDFs
            success = graph_download_and_save_files(ankenbango, "Ankens", buildername, ankenname, 納期)

            sheet[f"A{line}"] = ankenbango
            sheet[f"B{line}"] = buildername
            sheet[f"C{line}"] = ankenname
            sheet[f"E{line}"] = 納期

            if not success:
                sheet[f"D{line}"] = "NG"
                wb.save(excel_file)
                line += 1
                continue

            # ✉️ 2. Create Draft MailDealer Mail
            try:
                self.create_maildealer_draft(
                    driver, Maildealer_Data, Webaccess_Data, ankenname, buildername, 納期, ankenbango
                )
                sheet[f"D{line}"] = "OK"
            except Exception as e:
                logging.error(f"❌ MailDealer draft failed: {e}")
                sheet[f"D{line}"] = "NG"

            wb.save(excel_file)
            line += 1
            time.sleep(1)

        self.format_excel(excel_file)

        driver.quit()
        logging.info("🎯 All Finished Successfully!")

    def webaccess_login(self, driver, Webaccess_Data):
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, Webaccess_Data["xpaths"]["ログイン_xpaths"]["ログインID"]))
            ).send_keys("NasiwakRobot")
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, Webaccess_Data["xpaths"]["ログイン_xpaths"]["パスワード"]))
            ).send_keys("159753")
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["ログイン_xpaths"]["ログイン"]))
            ).click()
            time.sleep(5)
        except Exception as e:
            logging.error(f"❌ WebAccess login failed: {e}")

    def access_search(self, driver, Webaccess_Data):
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧"]))
            ).click()
            time.sleep(2)
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["リセット"]))
            ).click()
            time.sleep(3)

            from_date = self.from_date.strftime("%Y/%m/%d")
            to_date = self.to_date.strftime("%Y/%m/%d")

            # Handle FROM date
            from_date_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["確定納品日_1"]))
            )
            driver.execute_script("arguments[0].removeAttribute('readonly')", from_date_element)
            from_date_element.clear()
            from_date_element.send_keys(from_date)
            time.sleep(1)

            # Handle TO date
            to_date_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["確定納品日_2"]))
            )
            driver.execute_script("arguments[0].removeAttribute('readonly')", to_date_element)
            to_date_element.clear()
            to_date_element.send_keys(to_date)
            time.sleep(1)

            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["案件名_物件名"]))
            ).click()
            time.sleep(2)

            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["出荷区分"]))
            ).click()
            time.sleep(2)

            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["出荷区分_options"]["新規"])
                )
            ).click()
            time.sleep(2)

            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["案件名_物件名"]))
            ).click()
            time.sleep(1)

            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["検索"]))
            ).click()
            time.sleep(2)

            driver.execute_cdp_cmd("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": os.getcwd()})
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["案件一覧のダウンロード"])
                )
            ).click()
            time.sleep(5)
        except Exception as e:
            logging.error(f"❌ Access search failed: {e}")

    def maildealer_login(self, driver, Maildealer_Data):
        try:
            driver.execute_script(f"window.open('{Maildealer_Data['MailDealer_url']}', '_blank');")
            driver.switch_to.window(driver.window_handles[-1])
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, Maildealer_Data["MailDealer_name"]["MailDealer_Username"]))
            ).send_keys("チラント")
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, Maildealer_Data["MailDealer_name"]["MailDealer_Password"]))
            ).send_keys("7iww6vqp")
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Maildealer_Data["MailDealer_xpaths"]["MailDealer_Login_submit"]))
            ).click()
        except Exception as e:
            logging.error(f"❌ MailDealer login failed: {e}")

    def create_maildealer_draft(self, driver, Maildealer_Data, Webaccess_Data, ankenname, buildername, 納期, 案件番号):
        """Create and save a MailDealer draft email with attached PDFs."""

        def remove_non_bmp(text):
            return "".join(c for c in text if ord(c) < 0x10000)

        # ✉️ Mail Content
        mail_body = remove_non_bmp(
            f"""
    野原グループ
    森様

    いつもお世話になっております。

    {buildername} {納期}納品分
    軽天割付図面と見積書になります。

    ご査収の程宜しくお願い致します。

    {buildername}
    現場名：{ankenname}

    ★★★★★★エヌ・エス・ケー工業株式会社★★★★★★
    """
        )
        try:
            # 1. Go to MailDealer compose
            driver.switch_to.window(driver.window_handles[-1])
            WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifmSide")))
            WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, Maildealer_Data["MailDealer_xpaths"]["MailDealer_createmail"]))
            ).click()
            time.sleep(1)

            driver.switch_to.default_content()
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(2)

            # 2. Click Next
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Maildealer_Data["MailDealer_xpaths"]["MailDealer_createmail_next"])
                )
            ).click()
            time.sleep(1)

            # 3. Fill From/To/Subject/Body
            FromMail = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.NAME, Maildealer_Data["MailDealer_name"]["MailDealer_createmail_From"])
                )
            )
            FromMail.clear()
            FromMail.send_keys("kantou@nsk-cad.com")
            logging.info("From: kantou@nsk-cad.com")

            ToMail = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.NAME, Maildealer_Data["MailDealer_name"]["MailDealer_createmail_To"])
                )
            )
            ToMail.clear()
            ToMail.send_keys("juken1@nohara-inc.co.jp")
            logging.info("To: juken1@nohara-inc.co.jp")
            time.sleep(2)

            Subject = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.ID, Maildealer_Data["MailDealer_ID"]["MailDealer_createmail_Subject"])
                )
            )
            Subject.clear()
            Subject.send_keys(f"【軽天割付図面・御見積書送付】{buildername} {納期}納品分 {ankenname} ")
            logging.info(f"Subject: 【軽天割付図面・御見積書送付】{buildername} {納期}納品分 {ankenname} ")
            time.sleep(2)

            actions = ActionChains(driver)
            actions.send_keys(Keys.TAB).send_keys(Keys.TAB).send_keys(mail_body)
            actions.perform()
            time.sleep(2)

            # 4. Attach PDFs
            attach = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (By.XPATH, Maildealer_Data["MailDealer_xpaths"]["MailDealer_createmail_attach"])
                )
            )

            folder_path = os.path.join(os.getcwd(), "Ankens", ankenname)
            pdfs = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]

            if pdfs:
                attach.send_keys("\n".join(pdfs))
                time.sleep(2)
            else:
                logging.warning(f"⚠️ No PDFs found for {ankenname}")

            # 5. Send mail
            save_dropdown = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Maildealer_Data["MailDealer_xpaths"]["MailDealer_createmail_confirm"])
                )
            )
            save_dropdown.click()
            time.sleep(1)

            save_as_draft = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Maildealer_Data["MailDealer_xpaths"]["MailDealer_createmail_send"])
                )
            )
            save_as_draft.click()
            logging.info(f"✅ Sent email for {ankenname}")

            self.Access2(driver, Webaccess_Data, 案件番号)

        except Exception as e:
            logging.error(f"❌ Failed to send email: {e}")
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

    def Access2(self, driver, Webaccess_Data, 案件番号):
        try:
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(2)

            # 1. 受注一覧 button
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧"]))
            ).click()
            logging.info("✅ Clicked 受注一覧")

            # 2. リセット button
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["リセット"]))
            ).click()
            logging.info("✅ Clicked リセット")

            # 3. 案件番号 input
            anken_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["案件番号"]))
            )
            anken_input.clear()
            anken_input.send_keys(案件番号)
            logging.info(f"✅ Inputted 案件番号: {案件番号}")

            # 4. 検索 button
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["検索"]))
            ).click()
            logging.info("✅ Clicked 検索")
            time.sleep(2)

            # 5. 参照 button
            try:
                WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧_xpaths"]["参照"]))
                ).click()
                logging.info("✅ Clicked 参照")
            except Exception:
                logging.warning("⚠️ 参照 button not found, skipping this 案件.")
                return

            time.sleep(2)

            # 6. Get selected option
            project_drawing_select = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "project_drawing"))
            )
            selected_option = project_drawing_select.get_attribute("value")
            logging.info(f"📋 Current selected drawing status: {selected_option}")

            # 7. 図面 button
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["案件詳細_xpaths"]["図面"]))
            ).click()
            logging.info("✅ Clicked 図面")
            time.sleep(2)

            # 8. Choose new status
            if selected_option == "作図済":
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, Webaccess_Data["xpaths"]["案件詳細_xpaths"]["図面_options"]["送付済"])
                    )
                ).click()
                logging.info("✅ Selected 送付済み")
            elif selected_option == "CBUP済":
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, Webaccess_Data["xpaths"]["案件詳細_xpaths"]["図面_options"]["CB送付済"])
                    )
                ).click()
                logging.info("✅ Selected CB送付済み")

            time.sleep(2)

            # 9. 見積 button
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["案件詳細_xpaths"]["見積書"]))
            ).click()
            logging.info("✅ Clicked 見積書")
            time.sleep(2)

            # 10. Choose new status
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Webaccess_Data["xpaths"]["案件詳細_xpaths"]["見積書_options"]["送付済"])
                )
            ).click()
            logging.info("✅ Selected 見積書 送付済み")

            time.sleep(2)

            # 11. Save the Project Info
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, Webaccess_Data["xpaths"]["案件詳細_xpaths"]["案件情報を更新する"])
                )
            ).click()
            logging.info("✅ Clicked 保存 (Save)")

            time.sleep(2)

            # 12. Return back to 受注一覧 page
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, Webaccess_Data["xpaths"]["受注一覧"]))
            ).click()
            logging.info("✅ Returned to 受注一覧")

        except Exception as e:
            logging.error(f"❌ Failed during Access2 update: {e}")

    def format_excel(self, excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active

        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 25

        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(excel_file)


# === App Launcher ===
class App(tk.Tk):
    def __init__(
        self,
        from_date: datetime,
        to_date: datetime,
    ):
        super().__init__()
        self.title("割付図_見積書送付 Bot")
        self.geometry("600x400")

        Label(self, text="割付図_見積書送付 Bot", font=("Roboto", -18, "bold")).pack(pady=10)
        Label(self, text="From Date:").place(relx=0.3, rely=0.4, anchor=CENTER)
        self.from_date_entry = DateEntry(self, date_pattern="yyyy/mm/dd")
        self.from_date_entry.place(relx=0.5, rely=0.4, anchor=CENTER)

        Label(self, text="To Date:").place(relx=0.3, rely=0.5, anchor=CENTER)
        self.to_date_entry = DateEntry(self, date_pattern="yyyy/mm/dd")
        self.to_date_entry.place(relx=0.5, rely=0.5, anchor=CENTER)

        # SET DEFAULT VALUE
        self.from_date_entry.set_date(from_date)
        self.to_date_entry.set_date(to_date)

        # Button(self, text="Start Bot", command=self.start_bot, bg="#3290db", fg="white").place(
        #     relx=0.5, rely=0.6, anchor=CENTER
        # )

        self.after(5000, self.start_bot)

    def start_bot(self):
        from_date = self.from_date_entry.get_date()
        to_date = self.to_date_entry.get_date()
        threading.Thread(target=self.run_bot, args=(from_date, to_date), daemon=True).start()

    def run_bot(self, from_date, to_date):
        try:
            bot = 割付図_見積書送付(from_date, to_date)
            bot.process_data()
        finally:
            self.after(0, self.finish_app)

    def finish_app(self):
        self.quit()
        self.destroy()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-date", required=True)
    parser.add_argument("--to-date", required=True)
    args = parser.parse_args()
    App(
        from_date=datetime.fromisoformat(args.from_date),
        to_date=datetime.fromisoformat(args.to_date),
    ).mainloop()
