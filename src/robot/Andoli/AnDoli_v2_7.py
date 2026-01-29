#
# import pyperclip
import logging
import os
import shutil
import time
from datetime import datetime

import pandas as pd

# import pyautogui
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# List of error's in Andoli bot -

# select_builder エラー  = Error when selecting builder
# 案件名見つかりません = Ankenmei not found (no result came after searching)
# 案件名見つかりませんでした = Ankmei not found (check bot log for more details)
# 現場コメントや登録ボタンエラー = Error in comment or 登録ボタン (check bot log for more details)
# 監督名エラー = Kantokumei not found
# メッセージ送信出来ませんでした = Error sending msg (check bot log for more details)
# 監督名一つ以上 = more than 1 kantoku found
# 選択ボタンエラー = error when selecting 選択 button
# システムエラー = System name doesn't match (check the name)


# icon = 'C:\\Users\\Amank\\Downloads\\Stuff\\Images\\ico file\\5.ico'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("Andoli_bot_log.log"), logging.StreamHandler()],
)

Andpad_url = "https://andpad.jp/login?iss=https%3A%2F%2Fauth.andpad.jp%2F"
Dandoli_url = "https://www.dandoli.jp/login?nextPage=/grandyhouse/sites/1468846/info"
Files = "Files"
excelFile = "案件情報.xlsx"
excel_path = Files + "\\" + excelFile
Write_data = "Andoli納期確認送付.xlsx"
pdf_file_path = "Andoli納期確認送付.pdf"

# delete Result.pdf everytime
if os.path.exists(f"{pdf_file_path}"):
    os.remove(pdf_file_path)
    logging.info("old pdf file removed")
else:
    logging.info("pdf file doesn't exist")

# Create Files folder if it doesn't exist
if not os.path.exists(f"{Files}"):
    os.makedirs(f"{Files}")
    logging.info("Files Folder Created")
else:
    logging.info("Folder Exists")

# move the 案件情報.xlsx file to the files folder
if not os.path.exists(f"{Files}/{excelFile}"):
    try:
        shutil.move(excelFile, f"{Files}/{excelFile}")
        logging.info("案件化 File Moved to FIles Folder")
    except Exception:
        logging.info("案件化 ファイル見つかりません")
        exit()

# Remove everything apart from Excel from Files folder
for f in os.listdir(Files):
    fpath = os.path.join(Files, f)
    if f != excelFile:
        if os.path.isfile(fpath) or os.path.islink(fpath):
            os.remove(fpath)
        elif os.path.isdir(fpath):
            shutil.rmtree(fpath)
        logging.info(f"Removed: {fpath}")
    else:
        logging.info(f"Kept: {fpath}")


# Set Chrome options for downloading files to the specified folder
chrome_options = Options()
prefs = {
    "credentials_enable_service": False,
    "profile.password_manager_enabled": False,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": False,
}
chrome_options.add_experimental_option("prefs", prefs)

chrome_options.add_argument("--guest")  # prevents profile sync
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-notifications")


# Initialize the WebDriver before the loop
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()


def Andpad_login():

    # # Open a new tab using JavaScript
    # driver.execute_script("window.open('about:blank','_blank');")

    # Switch to the newly opened tab
    driver.switch_to.window(driver.window_handles[1])

    driver.get(Andpad_url)

    if 得意先名 == "ヤマト住建㈱(アイワ)":
        logging.info("builder is Yamato")
        id = "yamatojk-andpad@aiwa-st.co.jp"
        password = "aiwa159753"

    elif 得意先名 == "㈱秀光ビルド":
        logging.info("builder is Shuuko")
        id = "nskhome@nsk-cad.com"
        password = "nsk159753"

    else:
        logging.info("builder is not Yamato")
        id = "haga@nsk-cad.com"
        password = "nsk000"

    time.sleep(5)

    # send click on login button(initial page)
    # driver.find_element(By.XPATH,"/html/body/div[1]/div[1]/div[2]/form/input[1]").click()
    # driver.find_element(By.XPATH,'//input[@ value="ログイン画面へ"]').click()
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//input[@ value="ログイン画面へ"]'))
    ).click()
    time.sleep(1)

    # enter login ID
    # driver.find_element(By.XPATH,"/html/body/div/div[2]/form/div[1]/input").send_keys(id) #("5ih@nsk-cad.com")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//input[@type="email"]'))).send_keys(id)
    time.sleep(1)

    # enter password
    # driver.find_element(By.XPATH,"/html/body/div/div[2]/form/div[2]/input").send_keys(password) #("nsk159753")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//input[@type="password"]'))).send_keys(
        password
    )
    time.sleep(1)  #

    # send click on login button
    # driver.find_element(By.XPATH,"/html/body/div/div[2]/form/button").click()
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//button[contains(text(),'ログイン')]"))
    ).click()
    time.sleep(3)

    logging.info("login successful")
    time.sleep(6)

    # pyautogui.press ('esc')


def andpad_logout():

    # switch to ANDPAD tab
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(1)

    # click on logout
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'ログアウト')]"))
    ).click()
    logging.info("Logout successful")
    time.sleep(1)

    # driver.switch_to.window(driver.window_handles[1])


def dandoli_login():

    driver.get(Dandoli_url)

    # Optionally, switch back to the original tab if needed
    # driver.switch_to.window(driver.window_handles[0])
    time.sleep(5)

    # click on username and send keys
    driver.find_element(By.XPATH, "//div/input[@name = 'username']").send_keys("kantou@nsk-cad.com")
    time.sleep(2)

    # click on Password and send keys
    driver.find_element(By.XPATH, "//div/input[@name = 'password']").send_keys("nsk00426")

    # send click on login button
    driver.find_element(By.XPATH, "//div/button[@class='btn login-btn']").click()
    time.sleep(3)

    logging.info("login successful")
    time.sleep(5)
    # pyautogui.press ('esc')


def export_to_pdf(Write_data, pdf_file_path):
    try:
        logging.info("Creating PDF...")

        # Convert Excel to PDF
        app = xw.App()
        app.visible = False
        app.display_alerts = False
        excel = app.books.open(rf"{os.getcwd()}\{Write_data}")
        excel.to_pdf(rf"{pdf_file_path}", include="Sheet1")
        time.sleep(2)
        # excel.save()
        # time.sleep(2)
        # excel.close()
        app.quit()

        logging.info(f"Created {Write_data} and saved as PDF: {pdf_file_path}")

    except Exception as e:
        logging.info(f"{e}")
        logging.info("Error Converting Excel to PDF")


def select_builder(name, 得意先名):

    # click on bird icon
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[1]/header/div[1]/a"))
    ).click()
    time.sleep(2)
    # wait for the page to load completely
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))

    # click on builder option
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div/span/a/div[1]/div"))
    )
    logging.info("builder option found")
    option = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div/span/a/div[2]/i[2]"))
    )
    driver.execute_script("arguments[0].click();", option)
    time.sleep(3)

    # wait for the page to load completely
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))

    # select builder option dropdown
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[7]/div[2]/ul")))
    logging.info("builder option dropdown found")
    time.sleep(3)

    try:
        # wait for the page to load completely
        WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))
    except Exception:
        logging.info("page didn't load completely")

    # if "ヤマダホームズ" in 得意先名 or "ﾔﾏﾀﾞﾎｰﾑｽﾞ" in 得意先名:
    #     得意先名  = 支店名

    try:
        # logging.info('kenshin found')
        time.sleep(1)
        select_builder = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable(
                (By.XPATH, f"//div[contains(text(), '{得意先名}')]/div[contains(text(), '{name}')]")
            )
        )
        # driver.execute_script("arguments[1].click();", select_builder)
        time.sleep(0.5)
        select_builder.click()
        logging.info(f"{得意先名} clicked")
        time.sleep(3)
    except Exception:
        # logging.info(f"Error occured, in selecting builder\n{e}")
        logging.info("Error occured, in selecting builder")
        return False

    time.sleep(0.5)

    # wait for the page to load completely
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))

    # click on genbakanri
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[1]/header/div[1]/div[1]/ul/li[2]/div/a"))
    ).click()
    time.sleep(1)
    logging.info("genbakanri clicked")

    # wait for the page to load completely
    WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))

    ######################################
    # select 詳細検索
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[2]/div[1]/div/form/div[1]/div[2]/div/label[2]")
        )
    ).click()

    # click on 現場名 column and send keys
    send_anken = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[2]/div[1]/div/form/div[1]/div[1]/div[3]/div[5]/input")
        )
    )
    time.sleep(2)
    send_anken.send_keys(物件名)
    time.sleep(1)
    logging.info("search bar clicked")

    # wait for the page to load completely
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))

    # click on search button
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[2]/div[1]/div/form/div[2]/div[1]/button")
        )
    ).click()
    logging.info("search button clicked")
    time.sleep(1)

    try:
        # wait for the page to load completely
        WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))
    except Exception as e:
        logging.info(f"page didn't load properly, error is: {e}")


dandoli_login()
time.sleep(1)

# Open a new tab using JavaScript ###### using this later to open ANDPAD
driver.execute_script("window.open('about:blank','_blank');")


# Andpad_login()

# wb = load_workbook(excel_path)

ankenmei = []
type = []
kantoku = []
nouki = []
ankenmeir = []
result = []
shiryou = []
builder = []
system = []
bangou = []
shiten = []


df = pd.read_excel(
    excel_path,
    sheet_name="Sheet1",
    usecols=["支店名", "案件番号", "システム", "物件名", "確定納期", "追加不足", "担当者", "不足", "得意先名"],
    dtype=str,
)
df["確定納期"] = df["確定納期"].astype(str).str.split(" ", expand=True)[0]
# df_col =  df.iloc[:,0] # : means column is constant, : is column and 0 is row
try:

    # Iterate over each row in the DataFrame
    for _, row in df.iterrows():
        # Extract values from the respective columns
        支店名 = row["支店名"]
        システム = row["システム"]
        物件名 = row["物件名"]
        案件番号 = row["案件番号"]
        追加不足 = row["追加不足"]
        確定納期 = row["確定納期"]
        担当者 = row["担当者"]
        不足 = row["不足"]
        得意先名 = row["得意先名"]

        if pd.isna(物件名):
            logging.info("案件が見つかりません")
            break

        logging.info(f"OG date is: {確定納期}")

        # Check if 不足 is NaN and handle it accordingly
        # 不足 = '' if pd.isna(不足) else str(不足)=
        # 追加不足 = '' if pd.isna(追加不足) else str(追加不足)
        得意先名 = "" if pd.isna(得意先名) else str(得意先名)
        支店名 = "" if pd.isna(支店名) else str(支店名)

        date1 = str(確定納期)
        try:
            date2 = datetime.strptime(date1, "%Y-%m-%d")
            date3 = date2.strftime("%m月%d日")
            logging.info(f"date format 1: {date3}")
        except Exception:
            logging.info("date format different, taking date as it is")
            date3 = date1
            logging.info(f"date format 2: {date3}")

        name = str(物件名)

        # write in Result excel file
        shiten.append(支店名)
        ankenmei.append(物件名)
        type.append(追加不足)
        # type.append(f"{追加不足 if 追加不足 else ''}")
        kantoku.append(担当者)
        nouki.append(date3)
        shiryou.append(不足)
        builder.append(得意先名)
        system.append(システム)
        bangou.append(案件番号)
        # shiryou.append(f"{不足 if 不足 else ''}")
        time.sleep(2)

        abc = 得意先名
        logging.info(f"OG builder is: {abc}")

        logging.info(
            f"「{支店名}」,「{物件名}」, 「{追加不足}」, 「{担当者}」, 「{date3}」, 「{不足}」, 「{システム}」, 「{得意先名}」, 「{案件番号}」"  # noqa
        )

        if "ヤマダホームズ" in 得意先名 or "ﾔﾏﾀﾞﾎｰﾑｽﾞ" in 得意先名:
            得意先名 = 支店名
            logging.info(f"Builder/Shiten is: {得意先名}")
        elif "ネクストワンインターナショナル㈱" in 得意先名:
            得意先名 = "ネクストワンインターナショナル"
        elif "建新" in 得意先名:
            得意先名 = "建新の現場管理（建築）"
        else:
            得意先名 = 得意先名

        # if 不足:
        #     text2 = f'★ {不足}が不足しておりますので、大至急送付をお願い致します。'
        # else:
        #     text2 = ""

        if pd.isna(不足):
            text2 = f"""
いつも大変お世話になっております。
ご依頼頂いております、軽天材の納材日確認となります。


【納材日：{date3}】


変更等御座いましたら、５日(営業日)前までに
ご連絡をお願い致します。

※納材日2日前(2営業日前)以降の納期変更に関しては別途費用を頂戴しております。
ご注意の上、納期をご確認下さい。


──────＜ご案内とご注意事項＞──────
▼中入れについて
建物内への納材はお受けしておりません。
予めご了承のほど宜しくお願い致します。

▼納材場所の確保について
本メールが届きましたら、納材場所をご検討いただき、
当日、大工さん不在の際は納材場所の確保をお願い致します。
───────────────────────

ご連絡の行違いが御座いましたら、
お詫び申し上げます。
よろしくお願い致します。

"""

        else:
            text2 = f"""
いつも大変お世話になっております。
★ {不足}が不足しておりますので、大至急送付をお願い致します。"""

        time.sleep(0.5)
        text = text2
        logging.info(f"{text}")
        time.sleep(2)

        # pyautogui.press ('esc')
        time.sleep(1)

        jap_name = "エヌ・エス・ケー工業"
        eng_name = "NSK工業(株)"

        # if 'Dandoli' in system:
        if システム == "Dandoli" or システム == "dandoli" or システム == "DANDOLI":

            logging.info("system is Dandoli")
            # Switch to the newly opened tab
            driver.switch_to.window(driver.window_handles[0])

            # selecting the builder
            try:
                # wait for the page to load completely
                WebDriverWait(driver, 30).until(
                    EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div"))
                )
            except Exception:
                logging.info("page didn't load properly or already loaded")

            try:
                select_builder(jap_name, 得意先名)
            except Exception:
                # logging.info(f"Error in select_builder, error is: {e}")
                logging.info("Error in select_builder")
                # continue
            time.sleep(1)

            try:
                try:
                    try:
                        select_ankenmei = f"//div[@data-title='{物件名}' and span[@class='highlight']]"
                        time.sleep(0.5)
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, select_ankenmei)))
                        time.sleep(1)
                        logging.info("found ankenmei in method 1 (@data-title)")
                    except Exception:
                        logging.info(
                            "Search ankenmei method 1 (@data-title) failed, trying method 2 with new_bukkenmei, 1st instance"  # noqa
                        )
                        new_bukkenmei = 物件名.replace("･", "・")
                        logging.info(f"New bukkenmei is: {new_bukkenmei}")
                        time.sleep(0.5)
                        # select_ankenmei= f"//div[@data-title='{new_bukkenmei}']"
                        select_ankenmei = f"//div[@data-title='{new_bukkenmei}' and span[@class='highlight']]"
                        time.sleep(0.5)
                        # Wait for the elements to be present in the DOM
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, select_ankenmei)))
                        logging.info("found ankenmei in method 2 (@data-title) with new bukkenmei, 1st instance")
                        logging.info("Not found with Japanese name, trying with @class now")

                except Exception:
                    try:
                        select_ankenmei = f"//div[@data-title='{物件名}']"
                        # select_ankenmei= f"//div[@data-title='{物件名}' and span[@class='highlight']]"
                        time.sleep(0.5)
                        # Wait for the elements to be present in the DOM
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, select_ankenmei)))
                        time.sleep(1)
                        logging.info("found ankenmei in method 1(@class)")
                    except Exception:
                        logging.info(
                            "Search ankenmei method 1(@class) failed, trying method 2 with new_bukkenmei, 1st instance"
                        )
                        new_bukkenmei = 物件名.replace("･", "・")
                        logging.info(f"New bukkenmei is: {new_bukkenmei}")
                        time.sleep(0.5)
                        select_ankenmei = f"//div[@data-title='{new_bukkenmei}']"
                        # select_ankenmei= f"//div[@data-title='{new_bukkenmei}' and span[@class='highlight']]"
                        time.sleep(0.5)
                        # Wait for the elements to be present in the DOM
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, select_ankenmei)))
                        logging.info("found ankenmei in method 2 (@class) with new bukkenmei, 1st instance")
                        logging.info("Not found with Japanese name, trying with English name now")

            except Exception:
                # logging.info(f"Error in locating ankenmei after searching, error is: {e}")
                logging.info("Error in locating ankenmei when searching")
                if "ヤマダホームズ" in abc:
                    try:
                        logging.info("trying with eng_name")
                        select_builder(eng_name, 得意先名)
                        time.sleep(1)
                        try:
                            try:
                                # select_ankenmei= f"//div[@data-title='{物件名}']"
                                select_ankenmei = f"//div[@data-title='{物件名}' and span[@class='highlight']]"
                                time.sleep(0.5)
                                # Wait for the elements to be present in the DOM
                                WebDriverWait(driver, 10).until(
                                    EC.visibility_of_element_located((By.XPATH, select_ankenmei))
                                )
                                time.sleep(1)
                                logging.info("found ankenmei in method 1 (@data-title)")
                            except Exception:
                                logging.info(
                                    "Search ankenmei method 1 (@data-title) failed, trying method 2 with new_bukkenmei, 2nd instance"  # noqa
                                )
                                new_bukkenmei = 物件名.replace("･", "・")
                                logging.info(f"New bukkenmei is: {new_bukkenmei}")
                                time.sleep(0.5)
                                # select_ankenmei= f"//div[@data-title='{new_bukkenmei}']"
                                select_ankenmei = f"//div[@data-title='{new_bukkenmei}' and span[@class='highlight']]"
                                time.sleep(0.5)
                                # Wait for the elements to be present in the DOM
                                WebDriverWait(driver, 10).until(
                                    EC.visibility_of_element_located((By.XPATH, select_ankenmei))
                                )
                                logging.info(
                                    "found ankenmei in method 2 (@data-title) with new bukkenmei, 2nd instance"
                                )
                                logging.info("Not found with Japanese name, trying with @class now")

                        except Exception:
                            try:
                                select_ankenmei = f"//div[@data-title='{物件名}']"
                                # select_ankenmei= f"//div[@data-title='{物件名}' and span[@class='highlight']]"
                                time.sleep(0.5)
                                # Wait for the elements to be present in the DOM
                                WebDriverWait(driver, 10).until(
                                    EC.visibility_of_element_located((By.XPATH, select_ankenmei))
                                )
                                time.sleep(1)
                                logging.info("found ankenmei in method 1(@class)")
                            except Exception:
                                logging.info(
                                    "Search ankenmei method 1(@class) failed, trying method 2 with new_bukkenmei, 2nd instance"  # noqa
                                )
                                new_bukkenmei = 物件名.replace("･", "・")
                                logging.info(f"New bukkenmei is: {new_bukkenmei}")
                                time.sleep(0.5)
                                select_ankenmei = f"//div[@data-title='{new_bukkenmei}']"
                                # select_ankenmei= f"//div[@data-title='{new_bukkenmei}' and span[@class='highlight']]"
                                time.sleep(0.5)
                                # Wait for the elements to be present in the DOM
                                WebDriverWait(driver, 10).until(
                                    EC.visibility_of_element_located((By.XPATH, select_ankenmei))
                                )
                                logging.info("found ankenmei in method 2 (@class) with new bukkenmei, 2nd instance")
                                logging.info("Not found with English name as well")
                            else:
                                error_msg = "案件名見つかりません"
                                logging.info(f"{error_msg}")
                                ankenmeir.append("NG")
                                result.append(error_msg)
                                result_df = pd.DataFrame(
                                    {
                                        "支店名": shiten,
                                        "システム": system,
                                        "ビルダー": builder,
                                        "案件番号": bangou,
                                        "案件名": ankenmei,
                                        "追・不": type,
                                        "監督名": kantoku,
                                        "納期": nouki,
                                        "案件検索": ankenmeir,
                                        "不足": shiryou,
                                        "結果": result,
                                    }
                                )
                                result_df.to_excel(Write_data, index=False)
                                continue
                        time.sleep(0.5)
                    except Exception:
                        # logging.info(f"Error in select_builder, error is: {e}")
                        error_msg = "select_builder エラー"
                        logging.info(f"{error_msg}")
                        ankenmeir.append("NG")
                        result.append(error_msg)
                        result_df = pd.DataFrame(
                            {
                                "支店名": shiten,
                                "システム": system,
                                "ビルダー": builder,
                                "案件番号": bangou,
                                "案件名": ankenmei,
                                "追・不": type,
                                "監督名": kantoku,
                                "納期": nouki,
                                "案件検索": ankenmeir,
                                "不足": shiryou,
                                "結果": result,
                            }
                        )
                        result_df.to_excel(Write_data, index=False)
                        continue
                else:
                    error_msg = "案件名見つかりません"
                    logging.info(f"{error_msg}")
                    ankenmeir.append("NG")
                    result.append(error_msg)
                    result_df = pd.DataFrame(
                        {
                            "支店名": shiten,
                            "システム": system,
                            "ビルダー": builder,
                            "案件番号": bangou,
                            "案件名": ankenmei,
                            "追・不": type,
                            "監督名": kantoku,
                            "納期": nouki,
                            "案件検索": ankenmeir,
                            "不足": shiryou,
                            "結果": result,
                        }
                    )
                    result_df.to_excel(Write_data, index=False)
                    time.sleep(0.5)
                    driver.refresh()
                    time.sleep(2)
                    continue

            try:
                elements = driver.find_elements(By.XPATH, select_ankenmei)
                time.sleep(0.5)
                num_elements = len(elements)
                logging.info(f"Number of elements found: {num_elements}")

                # If there are multiple elements, skip to the next loop iteration
                if num_elements > 1:
                    logging.info("Multiple elements found, skipping this iteration.")
                    time.sleep(2)  # Optional: wait before the next iteration
                    logging.info("More than 1 案件名 found \n moving to the next anken")
                    error_msg = "案件名一つ以上"
                    logging.info(f"{error_msg}")
                    ankenmeir.append("NG")
                    result.append(error_msg)
                    result_df = pd.DataFrame(
                        {
                            "支店名": shiten,
                            "システム": system,
                            "ビルダー": builder,
                            "案件番号": bangou,
                            "案件名": ankenmei,
                            "追・不": type,
                            "監督名": kantoku,
                            "納期": nouki,
                            "案件検索": ankenmeir,
                            "不足": shiryou,
                            "結果": result,
                        }
                    )
                    result_df.to_excel(Write_data, index=False)
                    time.sleep(0.5)
                    driver.refresh()
                    time.sleep(2)
                    continue

                elif num_elements == 0:
                    logging.info("0 element found, 案件名見つかりません\n moving to the next anken")
                    error_msg = "案件名見つかりません"
                    logging.info(f"{error_msg}")
                    ankenmeir.append("NG")
                    result.append(error_msg)
                    result_df = pd.DataFrame(
                        {
                            "支店名": shiten,
                            "システム": system,
                            "ビルダー": builder,
                            "案件番号": bangou,
                            "案件名": ankenmei,
                            "追・不": type,
                            "監督名": kantoku,
                            "納期": nouki,
                            "案件検索": ankenmeir,
                            "不足": shiryou,
                            "結果": result,
                        }
                    )
                    result_df.to_excel(Write_data, index=False)
                    time.sleep(0.5)
                    driver.refresh()
                    time.sleep(2)
                    continue

                elif num_elements == 1:
                    # If only one element is found, proceed with clicking
                    select_ankenmei_click = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, select_ankenmei))
                    )
                    time.sleep(0.5)
                    select_ankenmei_click.click()
                    logging.info("案件名見つかりました")
                    ankenmeir.append("OK")
                    time.sleep(2)

            except Exception as e:
                logging.info(f"An error occurred: {e}")
                error_msg = "案件名見つかりませんでした"
                logging.info(f"{error_msg}")
                ankenmeir.append("NG")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(0.5)
                driver.refresh()
                time.sleep(2)
                continue

            time.sleep(2)

            # wait for the page to load completely
            WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div")))

            try:
                # click on 現場コメント
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div/ul/li[2]/a/span")
                    )
                ).click()
                logging.info("現場コメント found")
                time.sleep(2)

                # click on 登録
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable(
                        (
                            By.XPATH,
                            "/html/body/div[1]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div[2]/div/div/div/div/div/div/div[1]/div/button",
                        )
                    )
                ).click()
                logging.info("登録 button found")
                time.sleep(2)

            except Exception as e:
                error_msg = "現場コメントや登録ボタンエラー"
                logging.info(f"{error_msg}")
                logging.info(f"error is: {e}")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(0.5)
                driver.refresh()
                time.sleep(2)
                continue

            # 全員 only for kenshin
            # if 得意先名 == '建新':
            if "建新" in 得意先名:
                logging.info("builder is 建新, sending msg to all")
                # click on 全員
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "/html/body/div[7]/div/div/div/div/form/div[2]/table/tbody/tr[1]/td/label")
                    )
                ).click()
                time.sleep(2)

            else:
                logging.info(f"builder is {得意先名}")
                try:
                    # click on input 監督名
                    kantoku_name = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable(
                            (
                                By.XPATH,
                                "/html/body/div[7]/div/div/div/div/form/div[2]/table/tbody/tr[1]/td/span/span[1]/span/ul/li/input",
                            )
                        )
                    )
                    kantoku_name.send_keys(担当者)
                    logging.info("kantokumei sent")
                    # kantoku_name.send_keys(Keys.ENTER)
                    # time.sleep(0.5)
                    # logging.info("kantoku name sent successfully")
                    # # input('a')
                    time.sleep(2)
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable(
                                (
                                    By.XPATH,
                                    "//*[@class='select2-results__option select2-results__message' and contains(text(),'対象が見つかりません')]",  # noqa
                                )
                            )
                        )
                        # raise Exception
                        error_msg = "監督名見つかりません"
                        logging.info(f"{error_msg}")
                        result.append(error_msg)
                        result_df = pd.DataFrame(
                            {
                                "支店名": shiten,
                                "システム": system,
                                "ビルダー": builder,
                                "案件番号": bangou,
                                "案件名": ankenmei,
                                "追・不": type,
                                "監督名": kantoku,
                                "納期": nouki,
                                "案件検索": ankenmeir,
                                "不足": shiryou,
                                "結果": result,
                            }
                        )
                        result_df.to_excel(Write_data, index=False)
                        time.sleep(0.5)
                        driver.refresh()
                        time.sleep(2)
                        continue
                    except Exception:
                        logging.info("didn't find 対象が見つかりません, checking 2nd")

                    try:
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@class="select2-results__option"]'))
                        )
                        # get the number of kantoku
                        kantoku_no_path = '//*[@class="select2-results__option"]'
                        # kantoku_no_path = '//span[@class="select2-dropdown select2-dropdown--below"]'
                        kantoku_no = driver.find_element(By.XPATH, kantoku_no_path)
                        time.sleep(0.5)
                        num_kantoku = len(kantoku_no)
                        logging.info(f"Number of Kantoku found {num_kantoku}")
                        if num_kantoku > 1:
                            logging.info("Multiple kantoku found, skipping this iteration.")
                            time.sleep(2)  # Optional: wait before the next iteration
                            error_msg = "監督名一つ以上"
                            logging.info(f"{error_msg}")
                            result.append(error_msg)
                            result_df = pd.DataFrame(
                                {
                                    "支店名": shiten,
                                    "システム": system,
                                    "ビルダー": builder,
                                    "案件番号": bangou,
                                    "案件名": ankenmei,
                                    "追・不": type,
                                    "監督名": kantoku,
                                    "納期": nouki,
                                    "案件検索": ankenmeir,
                                    "不足": shiryou,
                                    "結果": result,
                                }
                            )
                            result_df.to_excel(Write_data, index=False)
                            time.sleep(0.5)
                            driver.refresh()
                            time.sleep(2)
                            continue
                        elif num_kantoku == 1:
                            logging.info("Multiple kantoku found, skipping this iteration.")
                            time.sleep(2)  # Optional: wait before the next iteration
                            error_msg = "監督名一つ以上"
                            logging.info(f"{error_msg}")
                            result.append(error_msg)
                            result_df = pd.DataFrame(
                                {
                                    "支店名": shiten,
                                    "システム": system,
                                    "ビルダー": builder,
                                    "案件番号": bangou,
                                    "案件名": ankenmei,
                                    "追・不": type,
                                    "監督名": kantoku,
                                    "納期": nouki,
                                    "案件検索": ankenmeir,
                                    "不足": shiryou,
                                    "結果": result,
                                }
                            )
                            result_df.to_excel(Write_data, index=False)
                            time.sleep(0.5)
                            driver.refresh()
                            time.sleep(2)
                            continue
                    except Exception:
                        logging.info("only 1 kantoku probably")

                    kantoku_name.send_keys(Keys.ENTER)
                    time.sleep(0.5)
                    logging.info("kantoku name sent successfully")

                except Exception as e3:
                    error_msg = "監督名エラー"
                    logging.info(f"{error_msg}")
                    logging.info(f"e3: {e3}")
                    result.append(error_msg)
                    result_df = pd.DataFrame(
                        {
                            "支店名": shiten,
                            "システム": system,
                            "ビルダー": builder,
                            "案件番号": bangou,
                            "案件名": ankenmei,
                            "追・不": type,
                            "監督名": kantoku,
                            "納期": nouki,
                            "案件検索": ankenmeir,
                            "不足": shiryou,
                            "結果": result,
                        }
                    )
                    result_df.to_excel(Write_data, index=False)
                    time.sleep(0.5)
                    driver.refresh()
                    time.sleep(2)
                    continue

            # click on input text box and send text
            msg_box = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "/html/body/div[7]/div/div/div/div/form/div[2]/table/tbody/tr[3]/td/textarea")
                )
            )
            msg_box.click()
            time.sleep(1)
            logging.info("msgbox clicked")
            msg_box.send_keys(text)
            time.sleep(2)

            try:
                # click on 編集を実行 button to send
                send = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'編集を実行')]"))
                )
                logging.info("編集を実行 button found")
                time.sleep(2)
                # //*[@class="icon-warning"]
                send.click()
                logging.info("メッセージ送信済み")
                time.sleep(4)
                try:
                    WebDriverWait(driver, 20).until(
                        EC.visibility_of_element_located((By.XPATH, "//*[contains(text(),'宛先が選択されていません')]"))
                    )
                    logging.info("No kantoku selected, checked kantoku_name")
                    # raise Exception
                    error_msg = "宛先が選択されていません。"
                    logging.info(f"{error_msg}")
                    result.append(error_msg)
                    result_df = pd.DataFrame(
                        {
                            "支店名": shiten,
                            "システム": system,
                            "ビルダー": builder,
                            "案件番号": bangou,
                            "案件名": ankenmei,
                            "追・不": type,
                            "監督名": kantoku,
                            "納期": nouki,
                            "案件検索": ankenmeir,
                            "不足": shiryou,
                            "結果": result,
                        }
                    )
                    result_df.to_excel(Write_data, index=False)
                    result_df.to_excel(Write_data, index=False)
                    time.sleep(0.5)
                    driver.refresh()
                    time.sleep(2)
                    continue
                except Exception:
                    logging.info("didn't find 宛先が選択されていません, msg_sent successfully")

            except Exception as e:
                error_msg = "メッセージ送信出来ませんでした"
                logging.info(f"{error_msg}\nError is: {e} ")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                result_df.to_excel(Write_data, index=False)
                time.sleep(0.5)
                driver.refresh()
                time.sleep(2)
                continue

        #########################################################
        # elif 'Andpad' in system:
        elif システム == "Andpad" or システム == "andpad" or システム == "ANDPAD":
            logging.info("system is Andpad")

            try:
                Andpad_login()
            except Exception as e:
                logging.info(f"Error in Andpad_login: {e}")

            # Switch to the newly opened tab
            # driver.switch_to.window(driver.window_handles[1])

            time.sleep(1)
            # pyautogui.press ('esc')
            time.sleep(1)

            enter_ankenmei_1 = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.NAME, "q[keyword_cont]"))
            )
            enter_ankenmei_1.send_keys(物件名)
            time.sleep(0.5)
            enter_ankenmei_1.submit()
            time.sleep(2)

            # enter_ankenmei_1.send_keys(name)
            # logging.info('g')
            # time.sleep(3)
            # enter_ankenmei_1.send_keys(Keys.ENTER)
            # time.sleep(2)

            # Select the ankemei after search result comes
            select_ankemei = (
                "/html/body/div[1]/div/div[1]/section/section/div[3]/div/div/div/div[2]/section/table/tbody/tr"
            )

            try:
                # Wait for the elements to be present in the DOM
                WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, select_ankemei)))
                time.sleep(1)
                logging.info("select_ankemei found")

                # Find all elements matching the XPath
                elements = driver.find_elements(By.XPATH, select_ankemei)
                time.sleep(0.5)

                # Check the length of elements found
                num_elements = len(elements)
                logging.info(f"Number of elements found: {num_elements}")
                time.sleep(0.5)

                # If there are multiple elements, skip to the next loop iteration
                if num_elements > 1:
                    logging.info("Multiple elements found, skipping this iteration.")
                    logging.info("案件名一つ以上 \n trying to get the correct ankenmei")
                    time.sleep(2)  # Optional: wait before the next iteration
                    select_ankemei = f"//tr/td/p[contains(text(),'{物件名}')]"
                    select_ankemei_click = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, select_ankemei))
                    )
                    select_ankemei_click.click()
                    logging.info("案件名見つかりました")
                    ankenmeir.append("OK")
                    time.sleep(2)

                elif num_elements == 0:
                    logging.info("0 element found, 案件名見つかりません\n moving to the next anken")
                    error_msg = "案件名見つかりません"
                    logging.info(f"{error_msg}")
                    ankenmeir.append("NG")
                    result.append(error_msg)
                    result_df = pd.DataFrame(
                        {
                            "支店名": shiten,
                            "システム": system,
                            "ビルダー": builder,
                            "案件番号": bangou,
                            "案件名": ankenmei,
                            "追・不": type,
                            "監督名": kantoku,
                            "納期": nouki,
                            "案件検索": ankenmeir,
                            "不足": shiryou,
                            "結果": result,
                        }
                    )
                    result_df.to_excel(Write_data, index=False)
                    continue

                elif num_elements == 1:
                    # If only one element is found, proceed with clicking
                    select_ankemei_click = WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located((By.XPATH, select_ankemei))
                    )
                    select_ankemei_click.click()
                    logging.info("案件名見つかりました")
                    ankenmeir.append("OK")
                    time.sleep(2)

            except Exception as e:
                logging.info(f"An error occurred: {e}")
                error_msg = "案件名見つかりませんでした"
                logging.info(f"{error_msg}")
                ankenmeir.append("NG")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(1)
                andpad_logout()
                time.sleep(1)
                continue

            time.sleep(2)

            try:
                # click chat
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "/html/body/div[3]/div/div[1]/ul//*[contains(text(), 'チャット')]")
                    )
                ).click()
                time.sleep(2)

                # Switch to the newly opened chat tab
                # WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(3))  # Wait for the new window to open
                driver.switch_to.window(driver.window_handles[-1])  # Switch to the newly opened tab
                logging.info("switched to chat tab")
                time.sleep(4)

            except TimeoutException:
                error_msg = "chat見つかりませんでした"
                logging.info(f"{error_msg}")
                logging.info("Failed to switch to the new tab")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(1)
                andpad_logout()
                time.sleep(1)
                continue

            time.sleep(3)

            # click on お知らせ
            shirase = "/html/body/div/div/div/div/main/div/div[2]/div[3]/div[3]/button[3]"
            shirase1 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, shirase)))
            shirase1.click()
            time.sleep(2)

            # click on input text box
            textbox = driver.find_element(By.CLASS_NAME, "searchbox__input")
            textbox.click()
            logging.info("textbox clicked")
            textbox.send_keys(担当者)

            time.sleep(2)

            # click on 検索 #name automatically comes as soon as pasted #not required by just to be safe
            driver.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[1]/div/div[3]/div/button").click()
            logging.info("clicked on kensaku")
            time.sleep(2)

            ####################################################################################

            # Find all elements matching the XPath
            elements_name = driver.find_elements(By.XPATH, '//*[@id="container"]/div[3]/div/div[2]/div/label')

            # Check the length of elements found
            num_elements_1 = len(elements_name)
            logging.info(f"Number of kanotku name elements found: {num_elements_1}")

            # If there are multiple elements, skip to the next loop iteration
            if num_elements_1 > 2 or num_elements_1 == 1:
                logging.info("Multiple elements found, skipping this iteration.\nKantoku not found")
                time.sleep(2)  # Optional: wait before the next iteration
                # logging.info(f'More than 1 kantokumei found\nmoving to the next anken')
                error_msg = "監督名一つ以上"
                logging.info(f"{error_msg}")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(1)
                driver.close()
                time.sleep(0.5)
                andpad_logout()
                time.sleep(1)
                continue

            elif num_elements_1 == 2:
                checkbox = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located(
                        (By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div/label[1]")
                    )
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", checkbox)
                time.sleep(0.5)
                checkbox.click()
                logging.info("kantoku selected")
                logging.info("only 1 Kantoku found")

            ####################################################################################

            # # click on checkbox
            # driver.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div/label[1]").click()
            # time.sleep(3)
            time.sleep(1)
            # click on 選択
            try:
                # //wc-tsukuri-button[@button-type='button' and @data-test='add-notify-member-button']
                sentaku = driver.find_element(
                    By.XPATH,
                    "/html/body/div/div/div/div/div[3]/div/div[3]/span/span/wc-tsukuri-buttons/wc-tsukuri-button[2]",
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", sentaku)
                time.sleep(0.5)
                sentaku.click()
                logging.info("sentaku button clicked")
                time.sleep(2)
            except Exception as e4:
                logging.info(f"error e4: {e4}")
                logging.info("選択btn not found or not clicked")
                error_msg = "選択ボタンエラー"
                logging.info(f"{error_msg}")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(1)
                andpad_logout()
                time.sleep(1)
                continue

            # select msgbox #define text above, just call it her, make sure to add date as well, before you call it.
            # msgbox = "/html/body/div/div/div/div/main/div/div[2]/div[3]/div[2]/textarea"
            msgbox = "//textarea[@placeholder='メッセージを入力']"
            msgbox1 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, msgbox)))
            msgbox1.click()
            msgbox1.send_keys(text)
            time.sleep(3)

            try:
                # click on 送信 button
                sendmsg = "/html/body/div/div/div/div/main/div/div[2]/div[3]/div[3]/div/span/wc-tsukuri-button"
                sendmsg1 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, sendmsg)))
                sendmsg1.click()
                logging.info("メッセージ送信済み")
                time.sleep(4)
                # close the chat tab
                driver.close()
                time.sleep(2)

            # except TimeoutException:
            except Exception as e:
                error_msg = "メッセージ送信出来ませんでした"
                logging.info(f"{error_msg}\nError is: {e}")
                result.append(error_msg)
                result_df = pd.DataFrame(
                    {
                        "支店名": shiten,
                        "システム": system,
                        "ビルダー": builder,
                        "案件番号": bangou,
                        "案件名": ankenmei,
                        "追・不": type,
                        "監督名": kantoku,
                        "納期": nouki,
                        "案件検索": ankenmeir,
                        "不足": shiryou,
                        "結果": result,
                    }
                )
                result_df.to_excel(Write_data, index=False)
                time.sleep(1)
                andpad_logout()
                time.sleep(1)
                continue

        #############################################
        else:
            logging.info("system is not Dandoli or Andpad")
            error_msg = "システムエラー"
            logging.info(f"{error_msg}")
            result.append(error_msg)
            result_df = pd.DataFrame(
                {
                    "支店名": shiten,
                    "システム": system,
                    "ビルダー": builder,
                    "案件番号": bangou,
                    "案件名": ankenmei,
                    "追・不": type,
                    "監督名": kantoku,
                    "納期": nouki,
                    "案件検索": ankenmeir,
                    "不足": shiryou,
                    "結果": result,
                }
            )
            result_df.to_excel(Write_data, index=False)
            continue

        time.sleep(5)
        result.append("DONE")

        if 得意先名 == "ヤマト住建㈱(アイワ)" or 得意先名 == "㈱秀光ビルド":
            time.sleep(1)
            andpad_logout()
            time.sleep(2)

        result_df = pd.DataFrame(
            {
                "支店名": shiten,
                "システム": system,
                "ビルダー": builder,
                "案件番号": bangou,
                "案件名": ankenmei,
                "追・不": type,
                "監督名": kantoku,
                "納期": nouki,
                "案件検索": ankenmeir,
                "不足": shiryou,
                "結果": result,
            }
        )
        result_df.to_excel(Write_data, index=False)

        # after all the files are downloaded, start reading the data from here and put the data in the same excel file
        logging.info("Data written to excel and saved successfully")

except Exception as e:
    logging.info(f"{e}")


finally:
    time.sleep(10)
    driver.quit()

try:
    # Load the existing workbook
    wb = load_workbook(Write_data)

    # Select the active worksheet
    ws = wb.active
    # Set the column widths
    column_widths = {
        "A": 14,  # Column A width
        "B": 9,  # Column B width
        "C": 33,  # Column C width
        "D": 9,  # Column D width
        "E": 40,  # Column E width
        "F": 8,  # Column F width
        "G": 13,  # Column G width
        "H": 10,  # Column H width
        "I": 10,  # Column I width
        "J": 13,  # Column J width
        "K": 30,  # Column K width
        # Add more columns as needed
    }
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    # Add thick outside borders to the outer columns for the header
    header_border = Border(
        left=Side(border_style="medium"),
        right=Side(border_style="medium"),
        top=Side(border_style="medium"),
        bottom=Side(border_style="medium"),
    )

    # Add thin outside borders to the outer columns for other rows
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    # Apply the border to the outer columns A, B, and C for the header (row 1)
    for cell in ws[1]:  # Iterate through cells in row 1
        cell.border = header_border

    # Apply the border to the outer columns A, B, and C for other rows
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=11):
        for cell in row:
            cell.border = thin_border

    # Set page setup for A4 size
    ws.page_setup.paperSize = ws.PAPERSIZE_A3

    # Iterate through the rows and set the fill color for "OK" entries
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11):
    #     for cell in row:
    #         if cell.column == 10:  # Skip column 10
    #             continue
    #         if cell.value == "DONE" or cell.value == "OK":
    #             # Set the fill color to green
    #             cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    #         else:
    #             cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Iterate through rows starting from the second row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11):
        # Check if column 11 (column K) has any value other than "DONE"
        if row[10].value != "DONE":  # Column 11 is at index 10 (0-based index)
            # If the condition is met, color the entire row red
            for cell in row:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            # If column 11 has "DONE", color cells in column 1 to 9 green
            for cell in row:
                if cell.column == 10:  # Skip column 10 (index 9, column J)
                    continue
                if cell.value == "DONE" or cell.value == "OK":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                # else:
                #     cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Define the range of cells containing data (excluding headers)
    data_range = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)

    # Set alignment for each cell in the data range
    for row in data_range:
        for cell in row:
            # Create a new alignment object with both horizontal and vertical center alignment
            align = Alignment(horizontal="center", vertical="center")
            # Apply the alignment to the cell
            cell.alignment = align

    # Adjust the page layout options
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Optional: Adjust margins if needed
    margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    ws.page_margins = margins

    # Define the print area
    ws.print_area = f"A1:K{ws.max_row}"

    # Save the workbook
    wb.save(Write_data)
    time.sleep(0.5)
except Exception as e:
    logging.info(f"Error in Excel allignment, error is:\n{e}")

# try:
#     export_to_pdf(Write_data, pdf_file_path)
#     time.sleep(2)
# except Exception as e:
#     logging.info(f"Error in export_to_pdf, error is:\n{e}")
