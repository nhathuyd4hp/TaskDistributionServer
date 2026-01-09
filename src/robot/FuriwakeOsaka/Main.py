# -*- coding: utf-8 -*-
# Osaka Kojo Furiwake Bot — BOM + 配車表 + USB Upload version:
# - Uses bom_downloader.py to fetch:
#       BOM  -> ./BOM/<日付>/
#       配車表 -> ./配車表/
# - Reads 配車(予定) from ./配車表
# - Matches against BOM under ./BOM/<日付>/
# - strict PDF=1 rule, Excel=floors rule (集合住宅 = floors multiple OK)
# - furiwake to ./▽USB/<号車>/(割付図, excels)
# - AFTER furiwake: uploads ▽USB under
#       https://nskkogyo.sharepoint.com/sites/yanase/Shared Documents/大阪工場　製造データ/{date}
# - results saved under ./Results, with 資料UP + OK/NG coloring
# - No SharePoint upload of BOM/results, only USB folder

import os, re, shutil, logging, unicodedata, math, base64
from datetime import datetime
from pathlib import Path, PurePosixPath

import requests
import jaconv
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
import customtkinter as ctk
from tkcalendar import Calendar

# ---------- Nasiwak / Logging ----------
from Nasiwak import *  # Bot_Update, etc.
from logging_setup import setup_logging
from bom_downloader import (
    download_factory_bom_for_date,
    download_osaka_haisha_for_date,
    download_tochigi_haisha_for_date,
    upload_usb_to_tochigi_date,
)
from config_access_token import get_access_token  # Graph access token
import argparse

setup_logging()
logger = logging.getLogger(__name__)

GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"

# ----------------- Normalisation helpers -----------------
JA_TRANS = str.maketrans({
    "（": "(", "）": ")", "－": "-", "ー": "-", "・": "･", "Ｆ": "F", "階": "F", "　": " "
})

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--工場",
        choices=["大阪工場　製造データ", "栃木工場"],
        required=True,
    )
    parser.add_argument(
        "--日付",
        required=True,
    )
    return parser.parse_args()

def ja_norm_strict(s: str) -> str:
    """For exact equality: z2h ascii/digit/kana, remove spaces/underscores, unify parens/dashes."""
    if s is None:
        return ""
    s = str(s).translate(JA_TRANS)
    s = jaconv.z2h(s, ascii=True, digit=True, kana=True)
    s = re.sub(r"[\s_＿]+", "", s).strip()
    return s

def anken_key(s: str) -> str:
    return ja_norm_strict(s)

# ★ builder＿物件名(10am訂正1027) [PDF]
# ★ builder＿物件名(10am訂正1027)＿天井割り付け1階＊ [Excel]
FILE_RE = re.compile(
    r"""^[★＊]?\s*
        (?P<builder>.+?)[＿_]\s*         # allow ＿ or _
        (?P<bukken>.+?)                  # 物件名
        (?:                              # ▼ time / 訂正 block is OPTIONAL
            \(
                (?P<ts>[^)]*)
            \)\)*                        # tolerate stray ')'
        )?
        (?:[＿_]\s*天井割り付け(?P<floor>平屋|\d+階)[＊*]?)?  # accept ＊ or *
        \s*$""",
    re.X,
)

def extract_tokens_from_name(name: str) -> dict:
    """
    Extract builder / 物件名 / floor info from BOM filename.

    Expected patterns:
      ★ビルダー＿物件名(…)(…)＿ 天井割り付け2階＊.xls
      ★ビルダー＿物件名(…)(…).pdf
      ★ビルダー＿物件名＿ 天井割り付け平屋＊.xls
      ★ビルダー＿物件名.pdf

    Strategy:
      1) Manual split:
         - strip extension
         - strip leading ★/＊
         - split at first ＿ / _
         - cut off 天井割り付け block
         - bukken keeps (3号棟) etc, but drops (4pm)/(訂正1030)
      2) If that fails, fallback to FILE_RE.
    """
    if not name:
        return {}

    # ---- strip extension, normalise parens ----
    base, _ = os.path.splitext(name)
    nm = base.replace("（", "(").replace("）", ")")

    # remove leading marks
    nm2 = nm.lstrip("★＊ ").strip()

    builder = None
    bukken  = None
    floor   = None

    # ---- 1) manual parse ----
    parts = re.split(r"[＿_]", nm2, maxsplit=1)
    if len(parts) >= 2:
        builder_candidate = parts[0].strip()
        rest              = parts[1].strip()

        # split off 天井割り付け part
        rest_main = rest
        floor_split = re.split(r"[＿_]\s*天井割り付け", rest, maxsplit=1)
        if len(floor_split) == 2:
            rest_main, floor_part = floor_split[0].strip(), floor_split[1]
            if "平屋" in floor_part:
                floor = "平屋"
            else:
                mfl = re.search(r"(\d+)階", floor_part)
                if mfl:
                    floor = f"{mfl.group(1)}階"

        # ---- bukken: keep (3号棟) etc, drop only 訂正/時間系 ----
        def looks_like_suffix_paren(content: str) -> bool:
            c = content.strip()
            if not c:
                return False
            # 訂正系
            if "訂正" in c:
                return True
            # am/pm, 4pm, 10am, etc.
            if re.search(r"(am|pm)", c.lower()):
                return True
            # pure 3–4 digit token like 1030, 1600 (時刻っぽい)
            if re.fullmatch(r"\d{3,4}", c):
                return True
            return False

        bukken_candidate = rest_main.strip()
        for m in re.finditer(r"\([^)]*\)", rest_main):
            inner = m.group(0)[1:-1]
            if looks_like_suffix_paren(inner):
                # everything BEFORE this paren is 物件名
                bukken_candidate = rest_main[:m.start()].strip()
                break

        if bukken_candidate:
            builder = builder_candidate
            bukken  = bukken_candidate

    # ---- 2) fallback to regex if bukken not resolved ----
    if not bukken:
        m = FILE_RE.match(nm2)
        if not m:
            return {}
        gd = m.groupdict()
        if not builder:
            builder = (gd.get("builder") or "").strip()
        bukken = (gd.get("bukken") or "").strip()
        floor  = gd.get("floor") or floor

    if not bukken:
        return {}

    return {
        "builder":      builder,
        "bukken":       bukken,
        "bukken_norm":  anken_key(bukken),
        "floor":        floor,
        "is_excel":     ("天井割り付け" in nm),
    }

# ----------------- Version / Setup -----------------
file_path_token = os.path.join(os.getcwd(), "Access_token", "Access_token.txt")
logging.info(f"Access token path: {file_path_token}")
try:
    with open(file_path_token, "r", encoding="utf-8") as f:
        ACCESS_TOKEN_FILE = f.read().strip()
except Exception:
    ACCESS_TOKEN_FILE = ""
logging.info("Access token loaded.")

REPO_OWNER = "Nasiwak"
REPO_NAME  = "koujou_furiwake_bot"
CURRENT_VERSION = "1.4.6-BOM-HAISHA-UPLOAD"
try:
    Bot_Update(REPO_OWNER, REPO_NAME, CURRENT_VERSION, ACCESS_TOKEN_FILE)
except Exception as e:
    logging.warning(f"Update check failed (non-blocking): {e}")

# Fresh 結果 book per run — save under ./Results
run_ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
RESULTS_DIR = os.path.join(os.getcwd(), "Results")
os.makedirs(RESULTS_DIR, exist_ok=True)
excelsheet = os.path.join(RESULTS_DIR, f"結果_{run_ts}.xlsx")

# ----------------- General helpers -----------------

def normalize_excel_id(val) -> str:
    """
    Convert Excel numeric IDs safely:
    1.0 -> '1'
    5.0 -> '5'
    'A' -> 'A'
    '01' -> '01'
    """
    if val is None:
        return ""

    # float like 1.0, 5.0
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)

    # int
    if isinstance(val, int):
        return str(val)

    # string
    s = str(val).strip()
    if s.endswith(".0") and s.replace(".", "", 1).isdigit():
        return s[:-2]

    return s

def _norm_j(text: str) -> str:
    return (
        jaconv.h2z(text or "", ascii=True, digit=True, kana=True)
        .replace(" ", "").replace("　", "")
        .replace("_", "").replace("＿", "")
    )

def jnorm(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("_", "").replace("＿", "")
    s = re.sub(r"[ \t\u3000]+", " ", s).strip()
    return s

def _clean_header(h: str) -> str:
    s = jaconv.z2h(str(h or ""), ascii=True, digit=True, kana=False)
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[ \u3000\t]", "", s)
    return s

TIME_TOKENS = re.compile(
    r'(?:(?P<h>\d{1,2}):(?P<m>\d{2}))'
    r'|(?:(?P<h2>\d{1,2})(?P<m2>\d{2}))'
    r'|(?:(?P<h3>\d{1,2})(?P<ap>am|pm))',
    re.IGNORECASE
)

def extract_time_token(name: str):
    s = (name or "").lower()
    m = TIME_TOKENS.search(s)
    if not m:
        return None
    if m.group('h') and m.group('m'):
        hh = int(m.group('h')); mm = int(m.group('m'))
    elif m.group('h2') and m.group('m2'):
        hh = int(m.group('h2')); mm = int(m.group('m2'))
    else:
        hh = int(m.group('h3')); mm = 0
        if m.group('ap') == 'pm' and hh != 12:
            hh += 12
        if m.group('ap') == 'am' and hh == 12:
            hh = 0
    return f"{hh:02d}{mm:02d}"

def safe_set(ws, cell_ref, value):
    if isinstance(value, str) and value[:1] in ('=', '+'):
        ws[cell_ref] = "'" + value
    else:
        ws[cell_ref] = value

def _parse_floor(val) -> int:
    """
    Robust floor count from '階':
      - '平屋' -> 1
      - '2F', '２階' -> 2
      - '1,2,3F' / '1・2・3F' / '1/2/3F' / '1,2,3 階' -> 3
      - numeric cell like 3 -> 3
    """
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 1
    s = str(val).strip()
    s = jaconv.z2h(s, ascii=True, digit=True, kana=False)
    if "平屋" in s:
        return 1

    if re.fullmatch(r"\d+", s):
        n = int(s)
        return n if n in (1, 2, 3) else 1

    tokens = re.split(r"[,\u3001/・\s]+", s)
    floors = set()
    for t in tokens:
        if not t:
            continue
        m = re.search(r"(\d+)", t)
        if m:
            n = int(m.group(1))
            if n in (1, 2, 3):
                floors.add(n)
    if floors:
        return len(floors)

    m = re.findall(r"(\d+)\s*(?:F|階)", s, flags=re.IGNORECASE)
    if m:
        floors = {int(x) for x in m if int(x) in (1, 2, 3)}
        if floors:
            return len(floors)

    return 1

OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def color_result_row(ws, row_idx: int, result_text: str):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        col_result = headers.index("結果") + 1
    except ValueError:
        return
    cell = ws.cell(row=row_idx, column=col_result)
    if isinstance(result_text, str) and result_text.strip().upper().startswith("OK"):
        cell.fill = OK_FILL
    elif isinstance(result_text, str) and result_text.strip().upper().startswith("NG"):
        cell.fill = NG_FILL

def create_new_result_excel(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "結果"
    headers = [
        "案件番号", "ビルダー名", "物件名", "階",
        "PDF数", "Excel数", "一致PDF", "一致Excel",
        "判定理由", "結果", "資料UP",
        "号車", "追加不足", "配送時特記事項",
    ]
    for i, h in enumerate(headers, start=1):
        safe_set(ws, f"{get_column_letter(i)}1", h)
    style_headers(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(path)
    logging.info(f"Created result workbook: {path}")


# ----------------- Styling -----------------
thin       = Side(style="thin", color="000000")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
header_font = Font(bold=True)

def style_headers(ws, header_row: int, num_cols: int):
    # Header row height a bit taller
    ws.row_dimensions[header_row].height = 32

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill      = header_fill
        cell.border    = border_all

def style_table(ws, first_data_row: int, last_row: int, last_col: int):
    if last_row < first_data_row:
        return

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Columns that should be centered (by header name)
    center_cols = set()
    for name in ("階", "PDF数", "Excel数", "結果", "資料UP", "案件番号", "号車"):
        if name in headers:
            center_cols.add(headers.index(name) + 1)

    # Columns where we expect long text and want wrap + left align
    wrap_cols = set()
    for name in ("物件名", "一致Excel", "一致PDF", "判定理由", "配送時特記事項", "追加不足"):
        if name in headers:
            wrap_cols.add(headers.index(name) + 1)

    # ----- Borders, alignment, row height -----
    for r in range(first_data_row, last_row + 1):
        # Slightly taller data rows for readability
        ws.row_dimensions[r].height = 27
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all

            if c in center_cols:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            elif c in wrap_cols:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                )

    # ----- Auto column width (respecting content length) -----
    max_width = {c: 0 for c in range(1, last_col + 1)}
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            v  = ws.cell(row=r, column=c).value
            ln = len(str(v)) if v is not None else 0
            if ln > max_width[c]:
                max_width[c] = ln

    for c in range(1, last_col + 1):
        width = min(max(10, int(max_width[c] * 1.2)), 60)
        ws.column_dimensions[get_column_letter(c)].width = width


def add_ng_sheet_and_metrics(wb_path: str, counters: dict):
    wb = load_workbook(wb_path)
    ws = wb["結果"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        idx_result = headers.index("結果") + 1
    except ValueError:
        idx_result = ws.max_column

    ng = wb.create_sheet("NG一覧")
    for i, h in enumerate(headers, 1):
        safe_set(ng, f"{get_column_letter(i)}1", h)
    style_headers(ng, 1, len(headers))
    ng.freeze_panes = "A2"
    ng.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    write_row = 2
    for r in range(2, ws.max_row + 1):
        res = str(ws.cell(row=r, column=idx_result).value or "").strip()
        if res.upper().startswith("NG"):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                safe_set(ng, f"{get_column_letter(c)}{write_row}", val)
            write_row += 1

    if write_row > 2:
        style_table(ng, 2, write_row - 1, len(headers))

    last = ws.max_row + 2
    safe_set(ws, f"A{last}",     "Summary")
    safe_set(ws, f"A{last+1}",   "合計件数");        safe_set(ws, f"B{last+1}", counters.get("total_rows", 0))
    safe_set(ws, f"A{last+2}",   "OK件数");          safe_set(ws, f"B{last+2}", counters.get("ok", 0))
    safe_set(ws, f"A{last+3}",   "NG件数");          safe_set(ws, f"B{last+3}", counters.get("ng", 0))
    safe_set(ws, f"A{last+4}",   "重複PDF件数");      safe_set(ws, f"B{last+4}", counters.get("dup", 0))
    safe_set(ws, f"A{last+5}",   "時刻不一致件数");    safe_set(ws, f"B{last+5}", counters.get("ts_mismatch", 0))
    safe_set(ws, f"A{last+6}",   "アップロード数");    safe_set(ws, f"B{last+6}", counters.get("uploaded", 0))

    style_table(ws, 2, ws.max_row, ws.max_column)
    wb.save(wb_path); wb.close()

# ----------------- SharePoint upload helpers -----------------
def _encode_share_url(url: str) -> str:
    """Encode a SharePoint URL into Graph /shares ID."""
    b = url.encode("utf-8")
    return "u!" + base64.urlsafe_b64encode(b).decode("ascii").rstrip("=")

def upload_usb_to_osaka_date(jp_date: str, usb_folder: Path) -> int:
    """
    Uploads the local ▽USB folder contents into:
      https://nskkogyo.sharepoint.com/sites/yanase/
        Shared Documents/大阪工場　製造データ/{jp_date}/▽USB/{号車}/...

    Returns number of successfully uploaded files.
    """
    usb_folder = Path(usb_folder)
    if not usb_folder.exists():
        logging.info(f"USB folder does not exist; skipping upload: {usb_folder}")
        return 0

    target_url = (
        "https://nskkogyo.sharepoint.com/sites/yanase/"
        "Shared Documents/大阪工場　製造データ/{date}"
    ).format(date=jp_date)

    logging.info(f"[Graph] USB upload target URL (date folder): {target_url}")

    share_id = _encode_share_url(target_url)
    token    = get_access_token()
    headers  = {"Authorization": f"Bearer {token}"}

    # Resolve the *date* folder driveItem
    resp = requests.get(
        f"{GRAPH_BASE_URL}/shares/{share_id}/driveItem",
        headers=headers,
    )
    if not resp.ok:
        logging.error(
            f"[Graph] Failed to resolve date folder: {resp.status_code} {resp.text}"
        )
        return 0

    info      = resp.json()
    drive_id  = info["parentReference"]["driveId"]
    parent_id = info["id"]

    usb_name = usb_folder.name  # should be "▽USB"
    uploaded = 0

    for root, _, files in os.walk(usb_folder):
        rel_root = Path(root).relative_to(usb_folder)

        for fname in files:
            local_path = Path(root) / fname

            # Build remote path so everything is under ▽USB on SharePoint:
            #   ▽USB/{truck}/... or ▽USB/file at root
            if rel_root == Path("."):
                remote_rel = str(PurePosixPath(usb_name) / fname)
            else:
                remote_rel = str(
                    PurePosixPath(usb_name)
                    / PurePosixPath(rel_root.as_posix())
                    / fname
                )

            put_url = (
                f"{GRAPH_BASE_URL}/drives/{drive_id}/items/{parent_id}:/"
                f"{remote_rel}:/content"
            )

            try:
                with open(local_path, "rb") as f:
                    put_resp = requests.put(put_url, headers=headers, data=f)
                if put_resp.status_code in (200, 201):
                    uploaded += 1
                    logging.info(f"[Graph] Uploaded {remote_rel} → {target_url}")
                else:
                    logging.error(
                        f"[Graph] Upload failed for {remote_rel}: "
                        f"{put_resp.status_code} {put_resp.text}"
                    )
            except Exception as e:
                logging.error(f"[Graph] Exception during upload for {remote_rel}: {e}")

    logging.info(f"[Graph] USB upload completed. Files uploaded: {uploaded}")
    return uploaded

# ----------------- Factory key resolver -----------------
def resolve_factory_key(factory_label: str) -> str | None:
    s = factory_label.strip()
    if "大阪" in s:
        return "大阪"
    if "九州" in s:
        return "九州"
    if "栃木" in s or "真岡" in s:
        return "栃木"
    if "千葉" in s:
        return "千葉"
    if "豊橋" in s:
        return "豊橋"
    if "滋賀" in s:
        return "滋賀"
    return None

# ----------------- Core (furiwake) -----------------
class process:
    def __init__(self, from_date, to_date) -> None:
        self.from_date = from_date        # 工場 (e.g. "大阪工場　製造データ")
        self.to_date   = to_date          # "11月15日" etc.
        logging.info(f"Run params — 工場:{self.from_date} / 日付:{self.to_date}")

        self.base_dir      = Path.cwd()
        self.usb_folder    = r'▽USB'
        self.haisha_folder = r'配車表'
        self.bom_root: Path | None = None

        # Fresh start: wipe BOM, USB, 配車表 and recreate for this run
        self.reset_run_folders()

        # 結果 book
        self.excelsheet = excelsheet
        create_new_result_excel(self.excelsheet)

        # ====== BOM + 配車 download via bom_downloader.py ======
        factory_key = resolve_factory_key(self.from_date)
        if factory_key:
            try:
                self.bom_root = download_factory_bom_for_date(
                    factory_key,
                    self.to_date,
                    self.base_dir,
                )
                if self.bom_root is None:
                    logging.warning(
                        f"No BOM files downloaded for factory '{factory_key}' "
                        f"and date '{self.to_date}'."
                    )
                else:
                    logging.info(f"BOM root for this run: {self.bom_root}")
            except Exception as e:
                logging.error(f"BOM download failed for factory '{factory_key}': {e}")
                self.bom_root = None

            # 配車表: only for 大阪
            if factory_key == "大阪":
                try:
                    haisha_root = download_osaka_haisha_for_date(
                        self.to_date,
                        self.base_dir,
                    )
                    if haisha_root is None:
                        logging.warning(
                            f"配車表 not downloaded for 大阪, date '{self.to_date}'. "
                            f"配車表 folder may be empty."
                        )
                    else:
                        logging.info(f"配車表 saved under: {haisha_root}")
                except Exception as e:
                    logging.error(
                        f"配車表 download failed for 大阪, date '{self.to_date}': {e}"
                    )
            elif factory_key == "栃木":
                try:
                    haisha_root = download_tochigi_haisha_for_date(
                        self.to_date,
                        self.base_dir,
                    )
                    if haisha_root is None:
                        logging.warning(
                            f"配車表 not downloaded for 栃木, date '{self.to_date}'. "
                            f"配車表 folder may be empty."
                        )
                    else:
                        logging.info(f"栃木 配車表 saved under: {haisha_root}")
                except Exception as e:
                    logging.error(f"配車表 download failed for 栃木, date '{self.to_date}': {e}")


        else:
            logging.info(
                f"No BOM mapping for 工場={self.from_date}; skipping BOM/配車 download."
            )

        # Metrics
        self.metrics = dict(
            total_rows=0, ok=0, ng=0, dup=0, ts_mismatch=0, uploaded=0
        )

        self.missing_bom_rows = []

        # 🔹 MAIN FURIWAKE LOGIC 🔹
        # This reads 配車表, matches against BOM, writes 結果 sheet & moves files into ▽USB
        self.compare_ankenmei()
        self.write_missing_bom_sheet() 

        factory_key = resolve_factory_key(self.from_date)
        usb_path = self.base_dir / self.usb_folder

        if factory_key == "大阪":
            try:
                uploaded_count = upload_usb_to_osaka_date(self.to_date, usb_path)
            except Exception as e:
                logging.error(f"USB upload error (大阪): {e}")
                uploaded_count = 0

        elif factory_key == "栃木":
            try:
                uploaded_count = upload_usb_to_tochigi_date(self.to_date, usb_path)
            except Exception as e:
                logging.error(f"USB upload error (栃木): {e}")
                uploaded_count = 0

        else:
            uploaded_count = 0

        # 資料UP column:
        # - If at least one file uploaded → 資料UP = OK for 結果==OK rows, NG otherwise
        # - If nothing uploaded → all 資料UP = NG
        upload_status = "OK" if uploaded_count > 0 else "NG"
        self.shiryou_upload_status(upload_status)

        # Add NG一覧 sheet + summary block at the end
        add_ng_sheet_and_metrics(self.excelsheet, self.metrics)

    def reset_run_folders(self):
        """Remove and recreate USB, 配車表, and BOM folders fully every run."""
        base = Path(self.base_dir)

        # --- 1) USB ---
        usb = base / "▽USB"
        if usb.exists():
            shutil.rmtree(usb, ignore_errors=True)
            logging.info("Removed folder: ▽USB")
        usb.mkdir(parents=True, exist_ok=True)
        logging.info("Created folder: ▽USB")

        # --- 2) 配車表 ---
        haisha = base / "配車表"
        if haisha.exists():
            shutil.rmtree(haisha, ignore_errors=True)
            logging.info("Removed folder: 配車表")
        haisha.mkdir(parents=True, exist_ok=True)
        logging.info("Created folder: 配車表")

        # --- 3) BOM ---
        bom = base / "BOM"
        if bom.exists():
            shutil.rmtree(bom, ignore_errors=True)
            logging.info("Removed folder: BOM (full)")
        bom.mkdir(parents=True, exist_ok=True)
        logging.info("Created folder: BOM (empty)")

        # Today's BOM folder
        today_dir = bom / self.to_date
        today_dir.mkdir(parents=True, exist_ok=True)
        self.bom_root = str(today_dir)
        logging.info(f"Prepared BOM folder for this date: {today_dir}")

    def create_clear_folder(self, path):
        # Kept for backward compatibility (not used now)
        if os.path.exists(path):
            shutil.rmtree(path)
            logging.info(f"Removed folder: {path}")
        os.makedirs(path, exist_ok=True)
        logging.info(f"Created folder: {path}")

    def count_uploaded_files(self) -> int:
        """Count how many files exist under ▽USB (for 'アップロード数' metric)."""
        base = Path(self.usb_folder)
        if not base.exists():
            return 0
        total = 0
        for _root, _dirs, files in os.walk(base):
            total += len(files)
        return total


    # ---- BOM artifact lookup ----
    def find_artifacts_for_anken(self, anken_key_norm: str):
        """
        Look for BOM files under ./BOM/<日付>/ (downloaded by bom_downloader).
        Fallback: ./BOM if date-specific folder doesn't exist.

        1st pass: strict bukken_norm == anken_key_norm
        2nd pass: (if strict empty) loose match where one name contains the other.
        """
        # ---- decide base BOM folder ----
        if self.bom_root is not None:
            base = Path(self.bom_root)
        else:
            candidate = self.base_dir / "BOM" / self.to_date
            base = candidate if candidate.exists() else (self.base_dir / "BOM")

        if not base.exists():
            logging.info(f"No BOM folder found at: {base}")
            return {"pdfs": [], "excels": [], "floors": set()}

        def detect_floor(name: str) -> str:
            FLOOR_PATTERNS = {
                "1F": [r"\b1F\b", r"１Ｆ", r"1階", r"１階", r"一階"],
                "2F": [r"\b2F\b", r"２Ｆ", r"2階", r"２階", r"二階"],
                "3F": [r"\b3F\b", r"３Ｆ", r"3階", r"３階", r"三階"],
                "平屋": [r"平屋"],
            }
            s = jnorm(name)
            for fl, pats in FLOOR_PATTERNS.items():
                for p in pats:
                    if re.search(p, s, flags=re.IGNORECASE):
                        return fl
            return "UNKNOWN"

        strict_pdfs, strict_excels, strict_floors = [], [], set()
        loose_pdfs,  loose_excels,  loose_floors  = [], [], set()

        for root, _, files in os.walk(base):
            for file in files:
                low = file.lower()
                if not (
                    low.endswith(".pdf")
                    or low.endswith((".xlsx", ".xls", ".csv"))
                ):
                    continue

                toks = extract_tokens_from_name(file)
                if not toks:
                    continue

                bukken_norm = toks.get("bukken_norm", "")
                if not bukken_norm:
                    continue

                full_path = os.path.join(root, file)

                # ---- strict match ----
                if bukken_norm == anken_key_norm:
                    if low.endswith(".pdf"):
                        strict_pdfs.append(full_path)
                    else:
                        strict_excels.append(full_path)
                        strict_floors.add(detect_floor(file))
                    continue

                # ---- loose match: one contains the other ----
                if anken_key_norm and (
                    anken_key_norm in bukken_norm or bukken_norm in anken_key_norm
                ):
                    if low.endswith(".pdf"):
                        loose_pdfs.append(full_path)
                    else:
                        loose_excels.append(full_path)
                        loose_floors.add(detect_floor(file))

        # Prefer strict, fallback to loose
        if strict_pdfs or strict_excels:
            return {
                "pdfs":   strict_pdfs,
                "excels": strict_excels,
                "floors": strict_floors,
            }
        else:
            return {
                "pdfs":   loose_pdfs,
                "excels": loose_excels,
                "floors": loose_floors,
            }

    def create_folder_and_move_matched_files(self, matched_files):
        try:
            go_folder  = os.path.join(self.usb_folder, self.号車)
            waritsuke  = os.path.join(go_folder, "割付図")
            os.makedirs(waritsuke, exist_ok=True)
            for file_path in matched_files:
                if not os.path.isfile(file_path):
                    continue
                name = os.path.basename(file_path)
                if name.lower().endswith(".pdf"):
                    dest = os.path.join(waritsuke, name)
                elif name.lower().endswith((".xlsx", ".xls", ".csv")):
                    dest = os.path.join(go_folder, name)
                else:
                    continue
                shutil.move(file_path, dest)
                logging.info(f"MOVED: {name} -> {dest}")
            return "OK"
        except Exception as e:
            logging.error(f"Move error: {e}")
            return "NG"

    def shiryou_upload_status(self, status):
        fp = self.excelsheet
        wb = load_workbook(fp)
        ws = wb["結果"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "資料UP" not in headers:
            wb.close()
            logging.info("資料UP column not present; skip updating.")
            return
        col_kekka   = headers.index("結果") + 1
        col_shiryou = headers.index("資料UP") + 1

        row_count = 0
        for row in ws.iter_rows(min_row=2):
            v = row[col_kekka - 1].value
            if v is not None and str(v) != "":
                row_count += 1
            else:
                break

        for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
            if i > row_count:
                break
            r    = str(row[col_kekka - 1].value or "")
            cell = row[col_shiryou - 1]
            val  = "OK" if (status == "OK" and r == "OK") else "NG"
            if isinstance(val, str) and val.startswith(('=', '+')):
                cell.value = "'" + val
            else:
                cell.value = val

        wb.save(fp); wb.close()
        logging.info("資料UP updated.")
    
    def write_missing_bom_sheet(self):
        """Create sheet '未UP案件' listing 配送先 with zero BOM files."""
        if not self.missing_bom_rows:
            logging.info("No missing BOM rows → 未UP案件 sheet not created.")
            return

        wb = load_workbook(self.excelsheet)
        ws = wb.create_sheet("未UP案件")

        headers = ["号車", "案件番号", "Builder名", "配送先", "必要階"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = h

        row = 2
        for item in self.missing_bom_rows:
            ws.cell(row=row, column=1).value = item["go"]
            ws.cell(row=row, column=2).value = item["anken"]
            ws.cell(row=row, column=3).value = item["builder"]
            ws.cell(row=row, column=4).value = item["bukken"]
            ws.cell(row=row, column=5).value = item["floors"]
            row += 1

        style_headers(ws, 1, len(headers))
        if row > 2:
            style_table(ws, 2, row - 1, len(headers))

        wb.save(self.excelsheet)
        wb.close()
        logging.info("未UP案件 sheet created.")


    # ------------ 配車読み込み & 判定 ------------"

    def compare_ankenmei(self):
        folder_path = os.path.join(os.getcwd(), "配車表")
        xls_path    = None
        xlsx_path   = None

        # ---- Find 配車 file ----
        for root, _, files in os.walk(folder_path):
            for f in files:
                if "配車" in f and f.lower().endswith(".xls"):
                    xls_path = os.path.join(root, f); break
                if "配車" in f and f.lower().endswith(".xlsx"):
                    xlsx_path = os.path.join(root, f); break
            if xls_path or xlsx_path:
                break

        if not xls_path and not xlsx_path:
            logging.error("配車ファイル(.xls/.xlsx)が見つかりません。処理を中止します。")
            return

        # ---- If .xls → convert ----
        temp_xlsx = os.path.join(folder_path, "temp_unlocked.xlsx")
        if xlsx_path:
            temp_xlsx = xlsx_path
        else:
            try:
                import win32com.client
                password = "nsk"
                excel    = win32com.client.Dispatch("Excel.Application")
                excel.DisplayAlerts = False
                wb_x = excel.Workbooks.Open(xls_path, False, True, None, password)
                wb_x.SaveAs(temp_xlsx, FileFormat=51)
                wb_x.Close(False)
                excel.Application.Quit()
            except Exception as e:
                logging.error(f".xls の変換に失敗しました: {e}")
                return

        # ---- Read 配車 ----
        try:
            raw = pd.read_excel(temp_xlsx, sheet_name="予定")
        except Exception as e:
            logging.error(f"配車ファイルの読込に失敗しました: {e}")
            return

        df = raw.copy()
        norm_map = {_clean_header(c): c for c in df.columns}

        def pick(name):
            return norm_map.get(name)

        col_bukken   = pick("配送先")
        col_kai      = pick("階")
        col_builder  = pick("ビルダー名") or pick("得意先名")
        col_anken    = pick("案件番号")
        col_go       = pick("号車") or pick("号")
        col_status   = pick("追加不足")
        col_tokki    = pick("配送時特記事項")

        if not (col_bukken and col_kai):
            logging.error("予定シートに必要列不足: 配送先/階 が見つかりません。")
            return

        df = df.replace({pd.NA: None})
        df = df[~(df[col_bukken].isna())].copy()

        wb = load_workbook(self.excelsheet)
        ws = wb["結果"]
        write_row = 2
        total_ok = total_ng = 0

        # =============================
        # 🔥 MAIN 配車 → BOM MATCH LOOP
        # =============================
        for _, r in df.iterrows():
            anken_no = normalize_excel_id(r.get(col_anken, "")) if col_anken else ""
            builder  = str(r.get(col_builder, "") or "").strip() if col_builder else ""
            bukken   = str(r.get(col_bukken, "") or "").strip()
            floors   = _parse_floor(r.get(col_kai, ""))

            # 号車
            go_raw = r.get(col_go, "") if col_go else ""
            go = normalize_excel_id(go_raw).upper()


            # 追加不足
            status = str(r.get(col_status, "") or "").strip() if col_status else ""

            # 特記事項
            tokki = str(r.get(col_tokki, "") or "").strip() if col_tokki else ""

            # ---- Missing 配送先 ----
            if not bukken:
                vals = [
                    anken_no, builder, "", floors,
                    0, 0, "", "",
                    "配送先(物件名)欠落", "NG", "NG",
                    go, status, tokki,
                ]
                for i, v in enumerate(vals, start=1):
                    safe_set(ws, f"{get_column_letter(i)}{write_row}", v)
                color_result_row(ws, write_row, "NG")
                write_row += 1
                total_ng  += 1
                continue

            # ---- Find BOM artifacts ----
            key_norm = anken_key(bukken)
            arts     = self.find_artifacts_for_anken(key_norm)
            pdf_count   = len(arts["pdfs"])
            excel_count = len(arts["excels"])

            matched_pdf_name   = os.path.basename(arts["pdfs"][0]) if pdf_count > 0 else ""
            matched_excel_name = " | ".join(sorted(os.path.basename(p) for p in arts["excels"])) if excel_count > 0 else ""

            # --------------------------
            # 🔥 NEW FEATURE:
            # If BOTH PDF=0 AND Excel=0 → add to 未UP案件 list
            # --------------------------
            if pdf_count == 0 and excel_count == 0:
                self.missing_bom_rows.append({
                    "go": go,
                    "anken": anken_no,
                    "builder": builder,
                    "bukken": bukken,
                    "floors": floors,
                })

            # ---- 判定 ----
            reasons = []
            if pdf_count != 1:
                reasons.append(f"PDF数={pdf_count} (期待:1)")
            if excel_count != floors:
                reasons.append(f"Excel数={excel_count} (期待:{floors})")

            result  = "OK" if not reasons else "NG"
            shiryou = "NG"

            # ---- FURIWAKE (OK only) ----
            if result == "OK" and (arts["pdfs"] or arts["excels"]):
                try:
                    self.号車 = go or "未指定"
                    move_status = self.create_folder_and_move_matched_files(
                        arts["pdfs"] + arts["excels"]
                    )
                except Exception as e:
                    logging.error(f"Furiwake move error for '{bukken}': {e}")

            # ---- Write row ----
            vals = [
                anken_no, builder, bukken, floors,
                pdf_count, excel_count,
                matched_pdf_name, matched_excel_name,
                " / ".join(reasons), result, shiryou,
                go, status, tokki,
            ]

            for i, v in enumerate(vals, start=1):
                if isinstance(v, float) and math.isnan(v):
                    v = ""
                safe_set(ws, f"{get_column_letter(i)}{write_row}", v)

            color_result_row(ws, write_row, result)
            write_row += 1
            total_ok  += int(result == "OK")
            total_ng  += int(result == "NG")

        # ---- Style ----
        if write_row > 2:
            style_table(ws, 2, write_row - 1, ws.max_column)

        # ---- Update metrics ----
        self.metrics["total_rows"] = write_row - 2
        self.metrics["ok"]        = total_ok
        self.metrics["ng"]        = total_ng

        wb.save(self.excelsheet)
        wb.close()


# ----------------- GUI -----------------
class DateHandler:
    def __init__(self, from_date, to_date, app=None):
        self.from_date = from_date
        self.to_date   = to_date
        self.app       = app
        self.process_dates()

    def process_dates(self):
        if self.app:
            self.app.set_busy(True)
        try:
            result = process(self.from_date, self.to_date)
            logging.info(f"From date: {self.from_date}")
            logging.info(f"To date: {self.to_date}")
            return result
        finally:
            if self.app:
                self.app.set_busy(False)

class App:
    def __init__(self, 
        root: ctk.CTk,
        工場: str,
        日付: str,
    ):
        self.root = root
        self.工場 = 工場
        self.日付 = 日付
        self.root.title("工場振り分け（大阪）")
        self.root.geometry("640x480")
        ctk.set_appearance_mode("Light")

        self.subtitle = ctk.CTkLabel(
            self.root,
            text="配車表・BOMを取得 → USB仕分け → 結果出力 → USBアップロード",
            font=("Meiryo UI", 12),
        )
        self.subtitle.pack(pady=(10, 0))

        frm = ctk.CTkFrame(self.root); frm.pack(fill="x", padx=16, pady=12)

        row1 = ctk.CTkFrame(frm); row1.pack(fill="x", pady=6)
        ctk.CTkLabel(row1, text="工場", width=80, anchor="e").pack(side="left", padx=6)
        self.from_date_entry = ctk.CTkOptionMenu(
            row1,
            values=[self.工場],
        )
        self.from_date_entry.pack(side="left", padx=6)

        row2 = ctk.CTkFrame(frm); row2.pack(fill="x", pady=6)
        ctk.CTkLabel(row2, text="日付", width=80, anchor="e").pack(side="left", padx=6)
        self.to_date_entry = ctk.CTkEntry(row2, placeholder_text="例) 11月15日")
        self.to_date_entry.pack(side="left", padx=6, fill="x", expand=True)
        self.to_date_entry.delete(0, tk.END)
        self.to_date_entry.insert(0, self.日付)

        self.progress = ctk.CTkProgressBar(self.root)
        self.progress.pack(fill="x", padx=16, pady=(6, 0))
        self.progress.set(0)
        self.status = ctk.CTkLabel(self.root, text="待機中", anchor="center")
        self.status.pack(pady=(4, 8))

        btns = ctk.CTkFrame(self.root); btns.pack(pady=8)
        self.start_button = ctk.CTkButton(btns, text="実行", width=140, command=self.on_start)
        self.start_button.pack(side="left", padx=10)
        self.close_button = ctk.CTkButton(btns, text="閉じる", width=120, command=self.root.destroy)
        self.close_button.pack(side="left", padx=10)

        self.footer = ctk.CTkLabel(
            self.root,
            text=f"Version {CURRENT_VERSION}   © {datetime.now().year} Nasiwak",
        )
        self.footer.pack(pady=(8, 12))

        self.root.after(15000,self.on_start)

    def set_busy(self, busy: bool):
        try:
            if busy:
                self.start_button.configure(state="disabled")
                self.progress.set(0.3)
                self.status.configure(text="処理中…")
            else:
                self.start_button.configure(state="normal")
                self.progress.set(1.0)
                self.status.configure(text="完了")
            self.root.update_idletasks()
        except Exception:
            pass

    def on_start(self):
        factory = self.from_date_entry.get()
        date_txt = (self.to_date_entry.get() or "").strip()
        if not date_txt:
            self.status.configure(text="日付を入力してください")
            return
        DateHandler(factory, date_txt, app=self)
        self.root.quit()
        self.root.destroy()

if __name__ == "__main__":
    args = parse_args()
    
    root = ctk.CTk()
    app = App(
        root,
        工場 = args.工場,
        日付 = args.日付,
    )
    root.mainloop()
