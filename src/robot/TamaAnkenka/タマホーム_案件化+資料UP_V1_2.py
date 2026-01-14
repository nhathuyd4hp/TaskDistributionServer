import os
from pathlib import Path
import re
import shutil
import subprocess
import pandas as pd
import pyautogui
import pyperclip
from selenium.webdriver import ActionChains
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException,NoSuchElementException,NoSuchFrameException
import time
import logging
from lxml import html
from openpyxl import Workbook
from openpyxl import load_workbook
import xlwings as xw
from openpyxl.styles import PatternFill, Font
from selenium.webdriver.common.action_chains import ActionChains 
from openpyxl.utils import get_column_letter
import win32com.client as win32
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment


#logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', handlers=[logging.FileHandler('タマホーム_案件化+資料UP_bot.log',encoding='utf-8'), logging.StreamHandler()])

from sharePoint_config import builder_sharepoint


Maildealer="https://mds3310.maildealer.jp/index.php"
sharepoint="https://nskkogyo.sharepoint.com"
web_access ="https://webaccess.nsk-cad.com/index.php"
genba_site="https://apg.kensetsu-cloud.jp/main/dyapp/a/AX4DD27C/assoclogin"
# tamahome_share = " https://nskkogyo.sharepoint.com/sites/2019/DocLib/Forms/AllItems.aspx?viewid=2993e1d4%2Dd235%2D4556%2D9fa7%2D5547911ca81f"
tamahome_share= "https://nskkogyo.sharepoint.com/:f:/s/2019/EjNHdPtC49BHrK_l3zsUTLIBCbNlqp1Vwlbjjmb8oEX7iA?e=6daKAh"


chrome_options = Options()

prefs = {
    "credentials_enable_service": False,
    "profile.password_manager_enabled": False,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": False
}
chrome_options.add_experimental_option("prefs", prefs)

chrome_options.add_argument("--guest")  # prevents profile sync
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-notifications")

driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
window_handles = driver.window_handles
current_w = driver.current_window_handle

# download_path=os.path.join(os.getcwd())
# logging.info(download_path)
def clean_excel(result):
        wb = load_workbook(result)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
                    cell.fill=PatternFill(fill_type=None)
                    cell.border = Border()
        wb.save(result)
        logging.info("clear access_touroku file")
        
def clear_folder(path):
            if os.path.exists(path):
                logging.info("Folder exists, removing...")
                shutil.rmtree(path)
                logging.info("Folder removed.")
            os.makedirs(path)
            logging.info("Folder created.")

download_directory = os.path.join(os.getcwd(), '新規案件')
clear_folder(download_directory)

# Define the folder path where the ZIP files are located

winrar_path = r"C:\Program Files\WinRAR\WinRAR.exe"  # Update this path if needed

def zip_extract(folder_path):
    # List all ZIP files in the specified folder
    zip_files = [f for f in os.listdir(folder_path) if f.endswith('.zip')]

    if not zip_files:
        raise FileNotFoundError("No ZIP files found in the specified folder.")

    # Process each ZIP file
    for zip_filename in zip_files:
        zip_file_path = os.path.join(folder_path, zip_filename)
        
        # Define the extraction command
        extraction_command = [
            winrar_path,  # Path to WinRAR executable
            'x',          # Extract command
            zip_file_path,  # The ZIP file to extract
            folder_path    # Destination folder
        ]
        
        # Run the command
        try:
            subprocess.run(extraction_command, check=True)
            logging.info(f"Extracted contents of {zip_filename} to: {folder_path}")
            
            # # Remove the original ZIP file after extraction
            os.remove(zip_file_path)
            logging.info(f"Removed ZIP file: {zip_filename}")
            
        except subprocess.CalledProcessError as e:
            logging.info(f"An error occurred while extracting {zip_filename}: {e}")

def move_files_to_root(root_folder):
    """
    Move all files from subfolders to the root folder and remove empty subfolders.
    """
    for dirpath, dirnames, filenames in os.walk(root_folder, topdown=False):
        # Move all files in the current directory to the root folder
        for file in filenames:
            source_file = os.path.join(dirpath, file)
            destination_file = os.path.join(root_folder, file)
            
            shutil.move(source_file, destination_file)
            logging.info(f"Moved file: {source_file} to {destination_file}")
        
        # Remove empty directories
        if not os.listdir(dirpath):
            os.rmdir(dirpath)
            logging.info(f"Removed empty folder: {dirpath}")

def mail_dealer():

    # Open a sharepoint in the new tab
    driver.get(Maildealer)
    logging.info("opening mail dealer")
    time.sleep(2)

    login = driver.find_element(By.XPATH,'//*[@id="fUName"]')
    logpassword = driver.find_element("name","fPassword")

    login.clear()
    logpassword.clear()

    login.send_keys("Nasiwakロボ")
    logpassword.send_keys("ouocf68l")

    driver.find_element(By.XPATH,"//input[@value='ログイン']").click()
    time.sleep(4)

    sideFrame = WebDriverWait(driver, 15).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifmSide")))
    logging.info('Switched to side Frame')
    time.sleep(1)

    # Wait for and click the "専用アドレス・タマホーム" element
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//span[contains(text(), "専用アドレス・タマホーム")]'))).click()
    logging.info("専用アドレス・タマホーム clicked")
    time.sleep(3)
    # Switch back to the default content and then to the main frame
    driver.switch_to.default_content()
    mainFrame = WebDriverWait(driver, 15).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifmMain")))
    logging.info('Back to main Frame')
    time.sleep(1)
    

def genba_login():

    driver.get(genba_site)
    time.sleep(1)
    genba_code = WebDriverWait(driver, 15).until( EC.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/div/div[1]/div[3]/form/table/tbody/tr[1]/td/div/input")))
    time.sleep(1)
    genba_code.clear()
    time.sleep(1)
    genba_code.send_keys("5101603")

    genba_ID = WebDriverWait(driver, 15).until( EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[1]/div[3]/form/table/tbody/tr[2]/td/div/input")))
    genba_ID.clear()
    genba_ID.send_keys("tama@nsk-cad.com")
    time.sleep(1)

    genba_pass = WebDriverWait(driver, 15).until( EC.element_to_be_clickable((By.XPATH,"//*[@id='password']")))
    genba_pass.clear()
    genba_pass.send_keys("tamahome2021!")

    genba_login = WebDriverWait(driver, 15).until( EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[1]/div[3]/form/table/tbody/tr[5]/td/button"))).click()
    logging.info("login to genbaplus")
    time.sleep(2)

def back_mail():   
    try:
        # Attempt to switch to the iframe
        driver.switch_to.frame("ifmMain")
        logging.info("Switched to ifmMain iframe")
    except NoSuchFrameException:
        # If NoSuchFrameException is raised, we are already in the iframe
        logging.info("Already in ifmMain iframe")
    Back_xpath = '/html/body/div/div/form/div[1]/div[1]/div[1]/button'    
    Back=driver.find_element(By.XPATH,Back_xpath)
    Back.click()
    logging.info('clicked on back button ')
    time.sleep(8)

def move_to_another_folder():
    
    try:
        # Attempt to switch to the iframe
        driver.switch_to.frame("ifmMain")
        logging.info("Switched to ifmMain iframe")
    except NoSuchFrameException:
        # If NoSuchFrameException is raised, we are already in the iframe
        logging.info("Already in ifmMain iframe")
    shinchaku = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '新着')]")))
    logging.info("shinchaku")
    driver.execute_script("arguments[0].click();", shinchaku)  # Click the element via JavaScript

    logging.info("新着 dropdown clicked")

    WebDriverWait(driver,10).until(EC.visibility_of_all_elements_located((By.XPATH,'/html/body/div/div/form/div[1]/div[2]/div[1]/div/div[1]/div[2]/ul')))
    
    # select the 対応中(作図まち)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"//li[text()='新着放置']"))).click()
    logging.info("successfully moved to shinchaku hochu folder")
    time.sleep(4)
    
def run_macro(wb):
    try:
        # Run the 'ClearSheet1Data' macro to clear existing data
        wb.macro('ClearSheet1Data')()
        logging.info("ClearSheet1Data macro executed")
        time.sleep(2)
    except Exception as e:
        logging.error(f"Error running ClearSheet1Data macro: {e}")
        
        
def write_data_access_excel(result):
    macro_file = "案件化.xlsm"
    cwd = os.getcwd()
    kekka_file = os.path.join(cwd, result)
    ankenka_file = os.path.join(cwd, macro_file)

    try:
        # Open the macro workbook in background
        app = xw.App(visible=False)
        wb = app.books.open(ankenka_file)
        logging.info("Excel macro file opened")

        # Run ClearSheet1Data macro
        run_macro(wb)

        # Read data from Excel files
        kekka_df = pd.read_excel(kekka_file, sheet_name="Sheet1")
        logging.info("Result file read")

        # Filter rows where '結果' is 'OK'
        filtered_kekka = kekka_df[kekka_df['結果'] == "OK"]

        # Create mapped DataFrame
        mapped_data = {
            "ビルダーコード": "014400",
            "案件番号": filtered_kekka["案件番号"],
            "ビルダー名": "タマホーム（株）",
            "案件名": filtered_kekka["案件名"],
            "納期": "",
            "希望納期": "",
            "納期確未": "",
            "監督1": "",
            "監督2": "",
            "発注番号": "",
            "住所": filtered_kekka["住所"],
            "車両": "",
            "階数": "",
            "目地有無": "",
            "入隅有無": "",
            "目地裏": "",
            "入隅金具": "",
            "延床面積": "",
            "配送時特記": "",
            "適用1": "",
            "メモ": "",
            "工事番号": filtered_kekka["現場番号"]
        }
        new_data = pd.DataFrame(mapped_data)

        # Write to 'シート' from cell A2
        sht = wb.sheets["シート"]
        sht.range("A2").value = new_data.values.tolist()
        logging.info("Data written to シート")

        # Run UpdateCells macro
        # wb.macro('UpdateCells')()
        # logging.info(" UpdateCells macro executed")

        # Save and close
        wb.save()
        wb.close()
        app.quit()
        logging.info("Excel macro file saved and closed")

    except Exception as e:
        logging.error(f" Error in write_data_access_excel: {e}")

genba_login()

mail_dealer()

# sharepointopen()

# Create a new Workbook
wb = Workbook()

# Create the first sheet and set its title
sheet1 = wb.active
sheet1.title = "Sheet1"

# Define the headings for the first sheet
headings = {
    "A": "案件番号",  
    "B": "案件名",  
    "C": "住所",
    "D": "現場番号",      
    "E": "結果" 
}

# Write the headings to the first row of the first sheet
for col_index, (col_letter, heading) in enumerate(headings.items(), start=1):  # start=1 for 1-based indexing
    sheet1.cell(row=1, column=col_index, value=heading)

# Save the workbook
Write_data = '結果.xlsx'
wb.save(Write_data)
time.sleep(0.5)

logging.info(f"Excel file '{Write_data}' created successfully with sheet '5ishome'.")

wb = load_workbook(Write_data)
sheet = wb['Sheet1']
sheet = wb.active

excellinenumber = 2

total_mails = 1

while total_mails>0:
        
        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.default_content()
        mainmenu = driver.find_element("id", "ifmMain")
        driver.switch_to.frame(mainmenu)
        time.sleep(3)
        shinki_level = False
        sakuzu_level = False
        try:
            # Wait until at least one matching element is present within the iframe
            try:
                shinki_level=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//td[contains(@class, 'olv-c-table__tbody-td') and count(.//span)=1 and .//span[text() = '新規案件']]")))
                shinki_level.click()
                logging.info("shinki_level clicked")
                shinki_level = True
            except:
                try:
                    sakuzumachi_lavel=WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//td[contains(@class, 'olv-c-table__tbody-td') and count(.//span)=1 and .//span[text() = '作図まち']]")))
                    sakuzumachi_lavel.click()
                    logging.info("sakuzumachi_lavel clicked")
                    sakuzu_level = True
                    
                    driver.switch_to.window(driver.window_handles[0])
                    #Switch it back to the main frame
                    driver.switch_to.default_content()
                    mainmenu = driver.find_element("id", "ifmMain")
                    driver.switch_to.frame(mainmenu)
                    logging.info("Move to mail dealer")
                    time.sleep(4)
                    try:
                        try:
                            # Attempt to switch to the iframe
                            driver.switch_to.frame("ifmMain")
                            logging.info("Switched to ifmMain iframe")
                        except NoSuchFrameException:
                            # If NoSuchFrameException is raised, we are already in the iframe
                            logging.info("Already in ifmMain iframe")
                        # click on blue tick
                        WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//div/span/button[@class='olv-p-mail-view-header__ops-bulk']"))).click()
                        logging.info("blue tick clicked")
                        time.sleep(1)
                        
                        side_frame = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/form/div[2]/section/div[1]/div[2]/div[1]/div[2]/div")))
                        logging.info("side_frame located")
                        
                        bangou = WebDriverWait(driver, 20).until(
                                EC.visibility_of_element_located(
                                    (By.XPATH, "/html/body/div/div/form/div[2]/section/div[1]/div[2]/div[1]/div[2]/div/div[2]/ul/li[5]/div/div[2]/div[1]/div[2]/span")
                                )
                            )
                        案件番号 = bangou.text
                        logging.info(f"案件番号: {案件番号}")
                        sheet[f'A{excellinenumber}'].value = 案件番号
                        time.sleep(1)

                        # wait for 案件名 element
                        ankken_mei = WebDriverWait(driver, 20).until(
                            EC.visibility_of_element_located(
                                (By.XPATH, "/html/body/div/div/form/div[2]/section/div[1]/div[2]/div[1]/div[2]/div/div[2]/ul/li[5]/div/div[2]/div[3]/div[2]/span")
                            )
                        )
                        bukken = ankken_mei.text
                        logging.info(f"案件名: {bukken}")
                        sheet[f'B{excellinenumber}'].value = bukken
                        time.sleep(1)

                       

                    except Exception as e:
                        logging.error(f"Error when doing 案件紐付け: {e}")

                        driver.switch_to.window(driver.window_handles[0])
                        #Switch it back to the main frame
                        driver.switch_to.default_content()
                        mainmenu = driver.find_element("id", "ifmMain")
                        driver.switch_to.frame(mainmenu)
                        logging.info("Move to mail dealer")
                        time.sleep(4)
                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(0.5)
                        
                        move_to_another_folder()
                        back_mail()
                        sheet[f'A{excellinenumber}'].value = "Already register"
                        sheet[f'E{excellinenumber}'].value = "NG"
                        wb.save(Write_data) 
                        excellinenumber += 1                   
                        time.sleep(1)
                        continue

                except:
                    logging.info(f"No mail for process")
                    break

        except TimeoutException:
            logging.info(f"No mail for process")
            break

        time.sleep(4)
        mail_body=WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,"//div[@class='olv-p-mail-view-body']/pre"))).text
        time.sleep(1)
        if shinki_level:
            # Extract text after [現場名]
            start_text = "[現場名]"
            start_index = mail_body.find(start_text) + len(start_text)

            # Get text after [現場名] until the end of the line
            if start_index != -1:
                bukken_text= mail_body[start_index:].splitlines()[0].strip()
                logging.info(f"bukken_text:{bukken_text}")
            else:
                logging.info("Text '[現場名]' not found.")

            def clean_name(name):
            # Step 1: Remove everything before and including the colon (both English and Japanese colons)
                if "：" in name:  # Handle Japanese colon
                    name = name.split("：", 1)[1].strip()
                elif ":" in name:  # Handle English colon
                    name = name.split(":", 1)[1].strip()
                
                # Step 2: If the name ends with " (建物)", remove it
                if " (建物)" in name:
                    name = name.replace(" (建物)", "")
                
                # Step 3: If the name contains "様邸", keep only the part up to "様邸"
                if "様邸" in name:
                    match = re.search(r".*?様邸", name)  # Match up to "様邸"
                    if match:
                        name = match.group(0)
                return name
            bukken=clean_name(bukken_text)
            logging.info(f"bukken:{bukken}")
            sheet[f'B{excellinenumber}'].value = bukken

         # Store the current window handle before switching
        current_window = driver.current_window_handle
        windows_before = driver.window_handles
        logging.info(f"Windows before clicking: {windows_before}")
        time.sleep(1)

        open_url=WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="form-olv-p-viewmail"]/div[2]/section/div/pre/a[1]')))
        open_url.click()
        logging.info("open_url")
        time.sleep(5)

        # Get all windows after the click
        windows_after = driver.window_handles
        logging.info(f"Windows after clicking: {windows_after}")
        time.sleep(1)


        windows_after = driver.window_handles
        if len(windows_after) > len(windows_before):
            driver.switch_to.window(windows_after[-1])
            logging.info(f"windows_after:,{windows_after}")
            logging.info(f"windows_after:,{windows_before}")
            logging.info("Switched to the new tab.")
            time.sleep(2)
        else:
            logging.info("no new tab found")
            # driver.close()
            driver.switch_to.window(driver.window_handles[0])
            logging.info("Closed the new tab and switched back to the original tab.")
            time.sleep(2)
            #Switch it back to the main frame
            driver.switch_to.default_content()
            mainmenu = driver.find_element("id", "ifmMain")
            driver.switch_to.frame(mainmenu)
            logging.info("Move to mail dealer")
            time.sleep(2)
        
            move_to_another_folder()
            back_mail()
            sheet[f'A{excellinenumber}'].value = ""
            sheet[f'C{excellinenumber}'].value = ""
            sheet[f'D{excellinenumber}'].value = ""
            sheet[f'E{excellinenumber}'].value = "現場リンクがない"
            wb.save(Write_data) 
            excellinenumber += 1                   
            time.sleep(1)
            continue

     

        driver.execute_cdp_cmd('Page.setDownloadBehavior', {'behavior': 'allow', 'downloadPath': download_directory})

        try:
            #if there is any shiryou available then download it
            elements = driver.find_elements(By.XPATH, "//span[@class='attachmentList']/span[@class='imageFrame attachmentFile']/a")

            # Click each element one by one
            for element in elements:
                element.click()
                # Wait for the popup to appear
                popup = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@id='dynagonImagePopup-popup']")))
                logging.info("Popup opened.")
                time.sleep(1)
                
                download_Btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dynagonImagePopup"]/div[2]/div[1]/div[4]/div[1]/a')))
                download_link = download_Btn.get_attribute('href')
                logging.info(download_link)
                
                driver.execute_script("window.open('');")  
                driver.switch_to.window(driver.window_handles[-1]) 
                time.sleep(1)
                driver.get(download_link)
                time.sleep(3)
                
                # Press Ctrl + S to save the PDF
                pyautogui.hotkey('ctrl', 's')
                logging.info("Pressed Ctrl + S")
                time.sleep(2)

                # Press Enter to confirm the save
                pyautogui.press('enter')
                logging.info("Pressed Enter to confirm save")
                time.sleep(6)
                logging.info("closing tab")
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                logging.info("Closed the new tab and switched back to the original tab.")
                time.sleep(1)
                popup_frame = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@id='dynagonImagePopup-popup']")))
                time.sleep(1)
                close_Btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dynagonImagePopup"]/a'))).click()
                time.sleep(1)  # Adjust the delay as needed
                chat_close_btn_pupup = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//span[@class="closeButton"]'))).click() 
                time.sleep(1)
        except:
            logging.info("No file found in toku") 
                  

        element_genba = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//a/div[contains(text(),"現場情報")]')))
        element_genba.click()
        logging.info("現場情報 ckicked")

        try:
            現場住所=WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH,"//span[@class='jobItemAddress']")))
            現場住所_tex=現場住所.text
            logging.info(f"現場住所:{現場住所_tex}")
            def extract_area(現場住所_text):
                # Match everything up to and including the first occurrence of "市" or "区"
                match = re.search(r".*?[市区郡]", 現場住所_text)
                if match:
                    return match.group(0)
                return 現場住所_text 
            現場住所_text=extract_area(現場住所_tex)
            logging.info(f"現場住所: {現場住所_text}")

            sheet[f'C{excellinenumber}'].value = 現場住所_tex
        except:
            logging.info("address not found")
            現場住所_text=""
            sheet[f'C{excellinenumber}'].value = ""

        try:
            genba_bangou=WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH,"//span[@class='cell content jobItemNumber']"))).text
            logging.info(f"genba_bangou:{genba_bangou}")
            match = re.search(r"-(.*)", genba_bangou)
            if match:
                現場番号 = match.group(1)
                logging.info(f"現場番号:{現場番号}")
                sheet[f'D{excellinenumber}'].value = 現場番号
        except:
            logging.info("現場番号 not found")
            sheet[f'D{excellinenumber}'].value = ""
        

        time.sleep(1)
        
        if shinki_level:

            driver.switch_to.window(driver.window_handles[0])

            try:
                # Wait for the pop-up to appear and close it
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="notice-balloon-balloon_display_hdd"]/div[2]/a'))).click()
                logging.info("Pop-up closed")
            except:
                logging.info("No pop-up found or pop-up close button not found within the timeout period")

            three_line=driver.find_element(By.XPATH,"//button[@class='icon-button olv-p-gnav__icon' and @title='その他']")
            driver.execute_script("arguments[0].click();", three_line)
            time.sleep(1)
            logging.info("clicked on 3 dots")
        

            WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[text()='案件管理']"))).click()
            logging.info("clicked on 案件管理")

            # check anken register or not
            current_w = driver.current_window_handle
            cwhs=driver.window_handles # all the windows that are opened
            driver.switch_to.window(cwhs[-1])
            WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="wrap"]/table/tbody/tr/td[2]/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/a'))).click()
            WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="itemType"]'))).click()
            WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="itemType"]/option[3]'))).click()
            anken_search = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div/table/tbody/tr/td[2]/table[1]/tbody/tr/td/table/tbody/tr/td/form/nobr/span[1]/input')))
            time.sleep(3)
    #################################
            anken_search.send_keys(bukken)
            WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="wrap"]/table/tbody/tr/td[2]/table[1]/tbody/tr/td/table/tbody/tr/td/form/nobr/button'))).click()
            try:
                time.sleep(3)
                driver.find_element(By.XPATH,'//*[@id="wrap"]/table/tbody/tr/td[2]/table[2]/tbody/tr[2]/td/div/table/tbody/tr/td/table/tbody/tr/td/div')
                logging.info('Not Registered yet') 

            except:
                reg = False
                data = driver.find_elements(By.CLASS_NAME,'bgclear_a_l')
                for value in data:
                    if 'タマホーム' == value.text:
                        reg = True
                        logging.info('Already registered')
                        time.sleep(1)
                        break

                    else:
                        logging.info('registered for other builder')
                        reg = False
                if reg:
                    driver.close()
                    time.sleep(1)

                    driver.switch_to.window(driver.window_handles[0])
                    #Switch it back to the main frame
                    driver.switch_to.default_content()
                    mainmenu = driver.find_element("id", "ifmMain")
                    driver.switch_to.frame(mainmenu)
                    logging.info("Move to mail dealer")
                    time.sleep(4)
                    driver.switch_to.window(driver.window_handles[-1])
                    time.sleep(0.5)
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(0.5)
                    move_to_another_folder()
                    back_mail()
                    sheet[f'A{excellinenumber}'].value = "Already register"
                    sheet[f'E{excellinenumber}'].value = "NG"
                    wb.save(Write_data) 
                    excellinenumber += 1                   
                    time.sleep(1)
                    continue

            # register anken 
            driver.find_element(By.XPATH,"//*[contains(text(),'案件を登録する')]").click()
            time.sleep(2)
            logging.info("案件を登録する")
            WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="matterForm"]/div[1]/table/tbody/tr[1]/td/input'))).send_keys("タマホーム")
            # SEND ANKEN MEI
            driver.find_element(By.XPATH,'//*[@id="matterForm"]/div[1]/table/tbody/tr[2]/td/input').send_keys(bukken)
            time.sleep(1)
            #send address
            driver.find_element(By.XPATH,"//input[@name='fDQuery[t12]']").send_keys(現場住所_text)
            time.sleep(1)
            # select ステータス	
            driver.find_element(By.XPATH,"//select[@name='fDQuery[statusid]']").click()
            time.sleep(1)
            # choose option open
            driver.find_element(By.XPATH,"//select[@name='fDQuery[statusid]']/option[@value='1']").click()
            time.sleep(2)
            # select current date
            date_select=WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"//nobr[text()='受付日時(入力不要)']/following::a[contains(@href, '__setCurrentDate')][1]")))
            driver.execute_script("arguments[0].click();", date_select)
            time.sleep(1)

        ################       #################

            # register
            driver.find_element(By.XPATH,'//*[@id="matterForm"]/div[2]/nobr/input').click()
            time.sleep(2)

            # copy anken bangou
            案件番号 = driver.find_element(By.XPATH,'//*[@id="matterForm"]/div[1]/table/tbody/tr[3]/td').text
            time.sleep(1)
            logging.info(f"案件番号:,{案件番号}")

            # 案件番号="457623"
            # logging.info(f"案件番号:,{案件番号}")
            driver.close()
            time.sleep(1)
            sheet[f'A{excellinenumber}'].value = 案件番号
            

        driver.switch_to.window(driver.window_handles[0])
        #Switch it back to the main frame
        driver.switch_to.default_content()
        mainmenu = driver.find_element("id", "ifmMain")
        driver.switch_to.frame(mainmenu)
        logging.info("Move to mail dealer")
        time.sleep(4)

        shiryou_path = os.path.join(download_directory, fr"{案件番号} {bukken}", "資料")
        os.makedirs(shiryou_path, exist_ok=True)
        logging.info(f"shiryou_path:{shiryou_path}")
        folder_path = os.path.dirname(shiryou_path)
        time.sleep(1)

        zip_folder_path = os.path.join(download_directory, "zip folder")

        # Check if the folder exists, if not, create it
        if not os.path.exists(zip_folder_path):
            os.makedirs(zip_folder_path)

        # List all files in the download directory which is download by chat 
        files = os.listdir(download_directory)
        # Loop through the files and move them to the shiyou path
        for file in files:
            source_file = os.path.join(download_directory, file)
            destination_file = os.path.join(shiryou_path, file)
            
            # Check if it's a file and move it
            if os.path.isfile(source_file):
                shutil.move(source_file, destination_file)
                print(f"Moved: {file}")
            else:
                print(f"Skipping non-file: {file}")
            logging.info(f"The file in the chat has been moved to Shiryou")

         #create empty text file in shiryou
        def create_text(shiryou_path, file_name="default.txt", file_content=""):
            # Define the full path for the text file
            text_path = os.path.join(shiryou_path, file_name)
            
            # Create and write to the text file
            with open(text_path, 'w') as file:
                file.write(file_content)
                logging.info(f"File '{file_name}' created in '{shiryou_path}' with content: '{file_content}'")

        # create text file
        create_text(shiryou_path, 'empty.txt', 'empty')

        driver.switch_to.window(driver.window_handles[-1])

        element_shryou = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//a/div[contains(text(),"図面")]')))
        element_shryou.click()
        logging.info("図面 ckicked")
        time.sleep(5)

        driver.execute_cdp_cmd('Page.setDownloadBehavior', {'behavior': 'allow', 'downloadPath': zip_folder_path})

        select_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[1]/div[1]/span'))
            )
        select_btn.click()
        logging.info("select_btn clicked")
        time.sleep(1)

        # Click the second button (all_selectbtn)
        all_selectbtn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[2]/div[2]/span[1]'))
        )
        all_selectbtn.click()
        logging.info("all_selectbtn clicked")
        time.sleep(1)

        # Try to click the download button
        try:
            download_file = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[2]/div[2]/span[6]'))
            )
            download_file.click()
            logging.info("download clicked")
            time.sleep(2)

            try:
                # Wait for the element to indicate no file available for download
                file_not_found = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@class='message' and text()='ダウンロードできるファイルがありません']"))
                )
                logging.info("ダウンロードできるファイルがありません")
                
                # If the element is found, raise the FileNotFoundError
                raise FileExistsError("ダウンロードできるファイルがありません")
            except TimeoutException:
                # If the element is not found within 15 seconds, proceed with normal execution
                logging.info("File found, proceeding with download.")


            waiting = WebDriverWait(driver, 600).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='barPercent' and contains(text(), '100%')]")))
            logging.info("ZIPファイルをダウンロードしました")

            pyautogui.press('esc')
            time.sleep(1)
            pyautogui.press('esc')
            time.sleep(2)
            logging.info("esc clicked")
            
            # Define the path to the ZIP folder
            abc = Path(zip_folder_path)
            zip_extract(abc)
            time.sleep(2)
            logging.info("zip file extracted")
            time.sleep(2)

            move_files_to_root(abc)
            logging.info("Files moved and extra folders cleaned up.")
            
            # Wait for the download process
            cancel_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[@class='fileMultiSelectModeOffButton button' and .//span[@class='value' and text()='キャンセル']]"))
            )
            driver.execute_script("arguments[0].click();", cancel_btn)
            logging.info("Alternative button clicked")
            time.sleep(2)
            logging.info(f"資料:,{zip_folder_path}")

            #### Check another folder
            try:
                WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='住宅仕様確認書']")))
                driver.refresh()
                
                another_folder=WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='住宅仕様確認書']")))
                driver.execute_script("arguments[0].click();", another_folder)
                logging.info(f"file download form another folder")

                select_btn = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[1]/div[1]/span'))
                    )
                select_btn.click()
                logging.info("select_btn clicked")
                time.sleep(1)

                # Click the second button (all_selectbtn)
                all_selectbtn = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[2]/div[2]/span[1]'))
                )
                all_selectbtn.click()
                logging.info("all_selectbtn clicked")
                time.sleep(1)

                # Try to click the download button
                try:
                    download_file = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[2]/div[2]/span[6]'))
                    )
                    download_file.click()
                    logging.info("download clicked")
                    time.sleep(2)

                    try:
                        # Wait for the element to indicate no file available for download
                        file_not_found = WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.XPATH, "//div[@class='message' and text()='ダウンロードできるファイルがありません']"))
                        )
                        logging.info("ダウンロードできるファイルがありません")
                        
                        # If the element is found, raise the FileNotFoundError
                        raise FileNotFoundError("ダウンロードできるファイルがありません")
                    except TimeoutException:
                        # If the element is not found within 15 seconds, proceed with normal execution
                        logging.info("File found, proceeding with download.")


                    waiting = WebDriverWait(driver, 180).until(
                        EC.presence_of_element_located((By.XPATH, "//div[@class='barPercent' and contains(text(), '100%')]")))
                    logging.info("ZIPファイルをダウンロードしました")

                    time.sleep(2)
                    logging.info(f"資料:,{zip_folder_path}")

                    # Define the path to the ZIP folder
                    abc = Path(zip_folder_path)
                    zip_extract(abc)
                    time.sleep(2)
                    logging.info("zip file extracted")
                    time.sleep(2)

                    move_files_to_root(abc)
                    logging.info("Files moved and extra folders cleaned up.")

                    pyautogui.press('esc')
                    time.sleep(1)
                    pyautogui.press('esc')
                    time.sleep(2)
                    logging.info("esc clicked")

                    # Wait for the download process
                    cancel_btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@class='fileMultiSelectModeOffButton button' and .//span[@class='value' and text()='キャンセル']]"))
                    )
                    driver.execute_script("arguments[0].click();", cancel_btn)
                    logging.info("cancel button clicked")
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='1つ上に戻る']"))).click()
                    time.sleep(2)

                except FileNotFoundError:
                    logging.info("No file to download")
                    time.sleep(1)

                    pyautogui.press('esc')
                    time.sleep(1)
                    pyautogui.press('esc')
                    # Wait for the download process
                    cancel_btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@class='fileMultiSelectModeOffButton button' and .//span[@class='value' and text()='キャンセル']]"))
                    )
                    driver.execute_script("arguments[0].click();", cancel_btn)
                    logging.info("cancel button clicked")
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='1つ上に戻る']"))).click()
                    time.sleep(2)


            except:
                logging.info(f"no other folder for download")
                # Wait for the download process
                cancel_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//span[@class='fileMultiSelectModeOffButton button' and .//span[@class='value' and text()='キャンセル']]"))
                )
                driver.execute_script("arguments[0].click();", cancel_btn)
                logging.info("cancel button clicked")
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='1つ上に戻る']"))).click()
                time.sleep(2)


            #check for one more another folder

            try:
                WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='確定設備図（エプコ）']")))
                driver.refresh()
                another_folder=WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@class='subDirectoryName overflow' and text()='確定設備図（エプコ）']")))
                driver.execute_script("arguments[0].click();", another_folder)
                logging.info(f"file download form another folder")

                select_btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[1]/div[1]/span')))
                select_btn.click()
                logging.info("select_btn clicked")
                time.sleep(1)

                # Click the second button (all_selectbtn)
                all_selectbtn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[2]/div[2]/span[1]'))
                )
                all_selectbtn.click()
                logging.info("all_selectbtn clicked")
                time.sleep(1)

                # Try to click the download button
                try:
                    download_file = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="directoryList"]/li[2]/div[2]/div[2]/span[6]'))
                    )
                    download_file.click()
                    logging.info("download clicked")
                    time.sleep(2)

                    try:
                        # Wait for the element to indicate no file available for download
                        file_not_found = WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.XPATH, "//div[@class='message' and text()='ダウンロードできるファイルがありません']"))
                        )
                        logging.info("ダウンロードできるファイルがありません")
                        
                        # If the element is found, raise the FileNotFoundError
                        raise FileNotFoundError("ダウンロードできるファイルがありません")
                    except TimeoutException:
                        # If the element is not found within 15 seconds, proceed with normal execution
                        logging.info("File found, proceeding with download.")


                    waiting = WebDriverWait(driver, 180).until(
                        EC.presence_of_element_located((By.XPATH, "//div[@class='barPercent' and contains(text(), '100%')]")))
                    logging.info("ZIPファイルをダウンロードしました")

                    time.sleep(2)
                    logging.info(f"資料:,{zip_folder_path}")

                    # Define the path to the ZIP folder
                    abc = Path(zip_folder_path)
                    zip_extract(abc)
                    time.sleep(2)
                    logging.info("zip file extracted")
                    time.sleep(2)

                    move_files_to_root(abc)
                    logging.info("Files moved and extra folders cleaned up.")

                    pyautogui.press('esc')
                    time.sleep(1)
                    pyautogui.press('esc')
                    time.sleep(2)
                    logging.info("esc clicked")
                except FileNotFoundError:
                    logging.info("No file to download")
                    time.sleep(1)
                    pyautogui.press('esc')
                    time.sleep(1)
                    pyautogui.press('esc')

            except:
                logging.info(f"no any other folder for download")


            for root, dirs, files in os.walk(abc):
                for file in files:
                    # Define the source and destination paths
                    src_path = os.path.join(root, file)
                    dest_path = os.path.join(zip_folder_path, file)
                    
                    # Move the file
                    shutil.move(src_path, dest_path)
                    logging.info(f'Moved file {src_path} to {dest_path}')

            # Optionally, remove empty directories in extracted_folder
            for root, dirs, files in os.walk(abc, topdown=False):
                for name in dirs:
                    dir_path = os.path.join(root, name)
                    try:
                        os.rmdir(dir_path)
                        logging.info(f'Removed empty directory: {dir_path}')
                    except OSError as e:
                        logging.info(f'Error removing directory {dir_path}: {e}')

            # Define keywords
            keywords = ['鋼製野縁割付依頼書', '鋼製野縁依頼', '平面図', '立面図', '配置図', 'プレカット図',
                        'PC図', '現場用伏図一式', '土台伏図', '電気図面', '24h換気図', '４ＬＤＫプラン', 
                        '２４Ｈ', '矩計図', '手書き配線図', '造作', '決定図面', '打合せ図面', '図面', '照明プラン',
                        '鋼製野縁','ﾌﾟﾚｶｯﾄ図','換気','24換気','平面・求積','電気打合せ図','手加工指示図',
                        '配線図','電気',]
            
            # Check if zip_folder_path has files
            if os.listdir(zip_folder_path):  # Only proceed if there are files in the folder
                # Loop through files in zip_folder_path
                for filename in os.listdir(zip_folder_path):
                    file_path = os.path.join(zip_folder_path, filename)
                    
                    # Check if the file name matches any keyword
                    if any([keyword in filename for keyword in keywords]):
                        # Move the matching file to shiryou_path
                        shutil.move(file_path, os.path.join(shiryou_path, filename))
                        time.sleep(2)

                # Clear all files in zip_folder_path
                for filename in os.listdir(zip_folder_path):
                    file_path = os.path.join(zip_folder_path, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)  # Delete the file
                        time.sleep(1)
            else:
                print("No files in zip_folder_path. Skipping file processing.")

            driver.close()
            time.sleep(1)

        except FileExistsError:
                logging.info("No folder for download")
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                logging.info("Closed the new tab and switched back to the original tab.")
                time.sleep(2)

        except Exception as e:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                logging.info("Closed the new tab and switched back to the original tab.")
                time.sleep(2)
                #Switch it back to the main frame
                driver.switch_to.default_content()
                mainmenu = driver.find_element("id", "ifmMain")
                driver.switch_to.frame(mainmenu)
                logging.info("Move to mail dealer")
                time.sleep(2)
                move_to_another_folder()
                back_mail()
                clear_folder(zip_folder_path)
                sheet[f'E{excellinenumber}'].value = "資料未ダウンロード"
                wb.save(Write_data) 
                excellinenumber += 1                   
                time.sleep(1)
                continue
            
        
        builder = "□案件番号500000～□"
       
        upload=builder_sharepoint(builder,案件番号, bukken)
        if upload:
            logging.info(f"Upload successful: {upload}")

        else:
            logging.info(f"Upload failed: {upload}")
            driver.switch_to.window(driver.window_handles[0])
            logging.info("Switched to Mail dealer")
            move_to_another_folder()
            back_mail()
            sheet[f'E{excellinenumber}'].value = "資料未アップロード"
            wb.save(Write_data) 
            excellinenumber += 1                   
            time.sleep(1)
            continue

        driver.switch_to.window(driver.window_handles[0])
        logging.info("Switched to Mail dealer")
        
#####################
        # 案件紐付け using ankenbango
        if shinki_level:
            try:
                try:
                    # Attempt to switch to the iframe
                    driver.switch_to.frame("ifmMain")
                    logging.info("Switched to ifmMain iframe")
                except NoSuchFrameException:
                    # If NoSuchFrameException is raised, we are already in the iframe
                    logging.info("Already in ifmMain iframe")
                # click on blue tick
                WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//div/span/button[@class='olv-p-mail-view-header__ops-bulk']"))).click()
                logging.info("blue tick clicked")
                time.sleep(1)

                # select the input box and send ankenbango
                input_box = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.NAME, 'fMatterID')))
                logging.info("input box selected")
                input_box.click()
                input_box.send_keys(案件番号)
                logging.info(f"案件番号:{案件番号} sent")
                time.sleep(0.5)

                # click on kanrenzuke
                kanrenzuke = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'olv-p-matter-search__ops')]//button[normalize-space(text())='関連付ける']")))
                kanrenzuke.click()
                # driver.execute_script("arguments[1].click();", kanrenzuke)
                time.sleep(2)
                logging.info("案件紐付け Successful")

            except Exception as e:
                logging.error(f"Error when doing 案件紐付け..... {e}")

            # Select level
        try:
            try:
                # Attempt to switch to the iframe
                driver.switch_to.frame("ifmMain")
                logging.info("Switched to ifmMain iframe")
            except NoSuchFrameException:
                # If NoSuchFrameException is raised, we are already in the iframe
                logging.info("Already in ifmMain iframe")
            
            level_btn = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='dropdown__trigger']"))).click()
            logging.info("Clicked on level selection btn")
            time.sleep(0.5)

            select_frame = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH, "//ul[@class='list has-scroll']")))
            logging.info("Frame selected")
            time.sleep(0.5)
            if shinki_level:
                作図まち_select = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH, "//li[text()='作図まち']"))).click()
                logging.info("作図まち selected")
                time.sleep(5)
            else:
                大至急_select = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH, "//li[text()='大至急']"))).click()
                logging.info("大至急 selected")
                time.sleep(5)
        except:
            logging.info("error when selecting level")
##############################################
        move_to_another_folder()
        back_mail()
        logging.info("Mail processed")

        sheet[f'E{excellinenumber}'].value = "OK"
        wb.save(Write_data) 
        excellinenumber += 1                   
        time.sleep(1)
time.sleep(1)

wb = load_workbook(Write_data)

# Select the active worksheet
ws = wb.active
# Set the column widths
column_widths = {
    "A": 15,  # Column A width
    "B": 40,  # Column B width
    "C": 40,   # Column C width
    "D": 15,  
    "E": 15,  
    # Add more columns as needed
}
for column, width in column_widths.items():
    ws.column_dimensions[column].width = width

# Add thick outside borders to the outer columns for the header
header_border = Border(
    left=Side(border_style='medium'),
    right=Side(border_style='medium'),
    top=Side(border_style='medium'),
    bottom=Side(border_style='medium')
)

# Add thin outside borders to the outer columns for other rows
thin_border = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)

header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange color
header_alignment = Alignment(horizontal='center', vertical='center')

# Apply the border to the outer columns A, B, and C for the header (row 1)
for cell in ws[1]:  # Iterate through cells in row 1
    cell.border = header_border
    cell.fill = header_fill
    cell.alignment = header_alignment
# Apply the border to the outer columns A, B, and C for other rows
for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# Iterate through the rows and set the fill color for "OK" entries
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value == "OK":
            # Set the fill color to green
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif cell.value == "NG" or cell.value == "資料未アップロード" or cell.value=="現場リンクがない":
            # Set the fill color to red
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            # Set the fill color to light blue
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        

# Define the range of cells containing data (excluding headers)
data_range = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)

# Set alignment for each cell in the data range
for row in data_range:
    for cell in row:
        # Create a new alignment object with both horizontal and vertical center alignment
        align = Alignment(horizontal='center', vertical='center')
        # Apply the alignment to the cell
        cell.alignment = align

# Save the workbook
wb.save(Write_data)
logging.info("changes in excel file")
time.sleep(1)

# access_result="アクセス登録.xlsx"
# clean_excel(access_result)
write_data_access_excel(Write_data)
time.sleep(1)
driver.quit()