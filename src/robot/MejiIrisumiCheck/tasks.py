import math
import os
import re
import tempfile
from datetime import datetime

import pandas as pd
import redis
from celery import shared_task
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from src.core.config import settings
from src.core.logger import Log
from src.core.redis import REDIS_POOL
from src.robot.MejiIrisumiCheck.automation import SharePoint, WebAccess
from src.service import ResultService as minio


def search_目地(excelPath: str):
    def get_visible_sheets(excelPath: str) -> list[str]:
        try:
            wb = load_workbook(excelPath, read_only=True)
            visible_sheets = [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
            wb.close()
            return visible_sheets
        except Exception:
            with pd.ExcelFile(excelPath) as excel:
                sheets: list[str] = excel.sheet_names
            return sheets

    for sheet in get_visible_sheets(excelPath):
        data = pd.read_excel(
            io=excelPath,
            sheet_name=sheet,
            header=None,
        )
        for _, row in data.iterrows():
            cells = [e for e in row.values.tolist() if pd.notna(e)]
            row = " ".join(map(str, cells)).strip()
            if re.match(r"^\d+\s+目地", row):
                if match := re.search(r"枚\s*(\d+)", row):
                    return int(match.group(1))
    return None


def search_入隅(excelPath: str):
    def get_visible_sheets(excelPath: str) -> list[str]:
        try:
            wb = load_workbook(excelPath, read_only=True)
            visible_sheets = [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
            wb.close()
            return visible_sheets
        except Exception:
            with pd.ExcelFile(excelPath) as excel:
                sheets: list[str] = excel.sheet_names
            return sheets

    for sheet in get_visible_sheets(excelPath):
        data = pd.read_excel(
            io=excelPath,
            sheet_name=sheet,
            header=None,
        )
        for _, row in data.iterrows():
            cells = [e for e in row.values.tolist() if pd.notna(e)]
            row = " ".join(map(str, cells)).strip()
            if re.match(r"^\d+\s+入隅", row):
                if match := re.search(r"枚\s*(\d+)", row):
                    return int(match.group(1))
    return None


@shared_task(bind=True, name="Meji-Irisumi Check")
def MejiIrisumiCheck(
    self,
    from_date: datetime | str,
    to_date: datetime | str,
):
    try:
        if isinstance(from_date, str):
            from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S.%f")
        if isinstance(to_date, str):
            to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S.%f")
        logger = Log.get_logger(channel=self.request.id, redis_client=redis.Redis(connection_pool=REDIS_POOL))
        logger.info(f"{from_date} ~ {to_date}")
        with sync_playwright() as p:
            browser = p.chromium.launch(
                channel="chrome",
                headless=False,
                args=[
                    "--start-maximized",
                ],
                timeout=10000,
            )
            context = browser.new_context(no_viewport=True)
            context.tracing.start(screenshots=True, snapshots=True, sources=True)
            with WebAccess(
                domain="https://webaccess.nsk-cad.com/",
                username="hanh0704",
                password="159753",
                playwright=p,
                logger=logger,
                browser=browser,
                context=context,
            ) as wa:
                logger.info("Downloading data from WebAccess...")
                orders = wa.download_orders(
                    from_date=from_date.strftime("%Y/%m/%d"),
                    to_date=to_date.strftime("%Y/%m/%d"),
                )
            # Clean Data
            orders = orders[orders["事業所"].isin(["横浜事務所", "ベトナム事務所　横浜事務所担当"])]
            orders = orders[~orders["得意先名"].isin(["タマホーム㈱, 一建設"])]
            if orders.empty:
                raise ValueError("No data to process.")
            orders = orders[["案件番号", "得意先名", "物件名", "目地", "目地数量", "入隅", "入隅数量", "資料リンク"]]
            # Append Column
            # 1️⃣ chèn "目地数量 Found" và "目地数量 Note" sau "目地数量" và trước "入隅"
            idx_meji = orders.columns.get_loc("目地数量") + 1
            orders.insert(idx_meji, "目地数量 Found", None)
            orders.insert(idx_meji + 1, "目地数量 Note", "")
            # 2️⃣ chèn "入隅数量 Found" và "入隅数量 Note" sau "入隅数量" và trước "資料リンク"
            idx_irisumi = orders.columns.get_loc("入隅数量") + 1
            orders.insert(idx_irisumi, "入隅数量 Found", None)
            orders.insert(idx_irisumi + 1, "入隅数量 Note", "")
            # Save
            save_path = os.path.abspath(
                f"{from_date.strftime("%Y/%m/%d")} {to_date.strftime("%Y/%m/%d")}.xlsx".replace("/", "-")
            )
            orders.to_excel(save_path, index=False)
            # Load
            wb = load_workbook(save_path)
            ws = wb.active
            with (
                SharePoint(
                    domain="https://nskkogyo.sharepoint.com/",
                    email="hanh3@nskkogyo.onmicrosoft.com",
                    password="Got21095",
                    playwright=p,
                    browser=browser,
                    logger=logger,
                    context=context,
                ) as sp,
                tempfile.TemporaryDirectory() as temp_dir,
            ):
                download_path = temp_dir
                orders = pd.read_excel(save_path)
                for index, row in orders.iterrows():
                    案件番号, _, _, _, 目地数量, _, _, _, 入隅数量, _, _, 資料リンク = row
                    # ---- #
                    # 目地数量 = 0 if math.isnan(目地数量) else 目地数量
                    # 入隅数量 = 0 if math.isnan(入隅数量) else 入隅数量
                    # ---- #
                    logger.info(f"{int(案件番号)} - {資料リンク} (Remaining: {orders.shape[0] - index - 1})")
                    if pd.isna(資料リンク):
                        continue
                    downloads = sp.download_files(
                        url=資料リンク,
                        steps=[re.compile("^見積書$")],
                        file=re.compile(r".*\.(xlsx|xls|xlsm|xlsb|xltx|xltm|xlam|xla|csv|ods)$", re.IGNORECASE),
                        save_to=download_path,
                    )
                    if not downloads:
                        ws[f"G{index+2}"].value = "情報足りない"
                        ws[f"K{index+2}"].value = "情報足りない"
                        continue
                    for download in downloads:
                        目地 = search_目地(download)
                        ws[f"F{index+2}"] = 目地
                        if 目地 is None or math.isnan(目地数量):
                            ws[f"G{index+2}"].value = "情報足りない"
                        else:
                            ws[f"G{index+2}"].value = 目地 == 目地数量
                        break
                    for download in downloads:
                        入隅 = search_入隅(download)
                        ws[f"J{index+2}"] = 入隅
                        if 入隅 is None or math.isnan(目地数量):
                            ws[f"K{index+2}"].value = "情報足りない"
                        else:
                            ws[f"K{index+2}"].value = 入隅 == 入隅数量
                        break
                wb.save(save_path)
                wb.close()
            result = minio.fput_object(
                bucket_name=settings.RESULT_BUCKET,
                object_name=f"MejiIrisumiCheck/{self.request.id}/{os.path.basename(save_path)}",
                file_path=save_path,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            os.remove(save_path)
            return f"{settings.RESULT_BUCKET}/{result.object_name}"
    except TimeoutError:
        pass
