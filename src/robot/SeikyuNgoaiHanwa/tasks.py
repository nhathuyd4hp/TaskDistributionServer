import io
import os
import re
import shutil
import tempfile
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal

import pandas as pd
import redis
from celery import shared_task
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from src.core.config import settings
from src.core.logger import Log
from src.core.redis import REDIS_POOL
from src.robot.SeikyuNgoaiHanwa.api import APISharePoint
from src.robot.SeikyuNgoaiHanwa.automation import Excel, SharePoint, WebAccess
from src.service.result import ResultService


@shared_task(bind=True, name="Seikyu Ngoài Hanwa")
def seikyu(
    self,
    from_date: datetime | str,
    to_date: datetime | str,
):
    # ----- Type Check
    if isinstance(from_date, str):
        from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S.%f").date()
    if isinstance(to_date, str):
        to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S.%f").date()
    # ----- Logger
    logger = Log.get_logger(channel=self.request.id, redis_client=redis.Redis(connection_pool=REDIS_POOL))
    logger.info(f"{from_date} ~ {to_date}")
    with (
        tempfile.TemporaryDirectory() as temp_dir,
        sync_playwright() as p,
    ):
        browser = p.chromium.launch(headless=False, args=["--start-maximized"])
        context = browser.new_context(no_viewport=True)
        context.tracing.start(screenshots=True, snapshots=True, sources=True)
        # --- #
        SeikyuFile: str = (
            "≪ベトナム≫請求書　阪和以外　(9日AM(10日分)、14日AM(15日分)、19日AM(20日分)、29日AM(末日分)に完成).xlsm"
        )
        if (
            APISharePoint(
                TENANT_ID=settings.API_SHAREPOINT_TENANT_ID,
                CLIENT_ID=settings.API_SHAREPOINT_CLIENT_ID,
                CLIENT_SECRET=settings.API_SHAREPOINT_CLIENT_SECRET,
            ).download_item(
                site_id="nskkogyo.sharepoint.com,fcec6ca2-58f4-4488-abf8-34e8ffbb741d,3136a8b2-a506-44d2-ad49-324bd156147c",
                breadcrumb=f"◆請求書　ベトナム専用◆/{SeikyuFile}",
                save_to=temp_dir,
            )
            is None
        ):
            logger.warning(f"Không tìm thấy file `{SeikyuFile}`")
            raise FileNotFoundError(f"Không tìm thấy file `{SeikyuFile}`")
        SeikyuFile = os.path.join(temp_dir, SeikyuFile)

        data: pd.DataFrame = WebAccess(
            username=settings.WEBACCESS_USERNAME,
            password=settings.WEBACCESS_PASSWORD,
            playwright=p,
            browser=browser,
            context=context,
        ).download_data(from_date, to_date)
        logger.info(f"Raw Data Shape: {data.shape}")
        if data.empty:
            return
        data = data[~data["商社名"].str.contains("阪和", na=False)]
        # Tìm dòng có công thức
        wb = load_workbook(SeikyuFile, data_only=False)
        ws = wb["請求一覧"]
        first_row_with_formula = None
        for cell in ws["A"]:
            row = cell.row
            if row < 11:
                continue
            if (
                all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 1])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 2])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 3])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 4])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 5])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 6])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 7])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 8])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 9])
            ):
                first_row_with_formula = row
                break
        # -- Copy Data -- #
        for sheet_name, _ in Excel.read(
            file_path=SeikyuFile,
            visible=False,
        ):
            if sheet_name == "請求一覧":
                Excel.write(
                    file_path=SeikyuFile,
                    data=[[item] for item in data["受注NO"].to_list()],
                    cell_range=f"I{first_row_with_formula}",
                    sheet_name="請求一覧",
                    visible=False,
                )
        # -- Extract Data--#
        quote_url = []
        prices = []
        with SharePoint(
            domain=settings.SHAREPOINT_DOMAIN,
            username=settings.SHAREPOINT_EMAIL,
            password=settings.SHAREPOINT_PASSWORD,
            playwright=p,
            browser=browser,
            context=context,
        ) as sp:
            data = pd.read_excel(
                io=SeikyuFile,
                sheet_name="請求一覧",
            )
            data = data.iloc[8:]
            data.columns = data.iloc[0]
            data = data.iloc[1:].reset_index(drop=True)
            for current, row in data.iterrows():
                logger.info(f"{current}/{data['受注NO'].last_valid_index()}")
                if (
                    pd.isna(row["365URL"])
                    # Không xử lí cột NaN
                    or pd.isna(row["受注NO"])
                    # Không xử lí giá trị màu xanh ở cột A
                    or pd.isna(row["締日"])
                    # Không xử lí dòng có giá trị 不足 ở cột B
                    or row["追加/先行"].find("不足") != -1
                    # Không xử lí ngày ngoài phạm vi
                    or row["納期"].date() < from_date
                    or row["納期"].date() > to_date
                ):
                    quote_url.append(row["見積URL"])
                    prices.append(row["税抜金額"])
                    continue
                logger.info(row["365URL"])
                downloads = sp.download(
                    url=row["365URL"],
                    steps=[
                        re.compile("^(見積|見積書)$"),
                    ],
                    file=re.compile(".*.(xlsx|xlsm|xls)"),
                )
                if not downloads or len(downloads) > 1:
                    quote_url.append(row["見積URL"])
                    prices.append(row["税抜金額"])
                    continue
                link, file_path = downloads[0]
                prices_found = []
                try:
                    for _, sheet_data in Excel.read(file_path, cell_range="A1:CC1000"):
                        for _, row in sheet_data.iterrows():
                            row = " ".join([str(cell) for cell in row])
                            if match := re.search(r"小\s*.*\s*計.*?(\d+(?:,\d{3})*(?:\.\d+)?)", row):
                                cleaned_row = " ".join([word for word in row.split() if word not in ["nan", "None"]])
                                print(f"Row: {cleaned_row}")
                                prices_found.append(match.group(1))
                                break
                except Exception:
                    prices_found = [0]
                quote_url.append(link)
                price: Decimal = Decimal(0)
                if prices_found:
                    price = Decimal(prices_found[0])
                price = price.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                prices.append(price)
        Excel.write(file_path=SeikyuFile, data=[[item] for item in quote_url], sheet_name="請求一覧", cell_range="H11")
        Excel.write(file_path=SeikyuFile, data=[[item] for item in prices], sheet_name="請求一覧", cell_range="F11")
        result_file = f"≪ボット≫請求書 阪和以外 {from_date}-{to_date}.xlsm"
        result_path = os.path.join(temp_dir, result_file)

        # Copy file
        shutil.copy2(SeikyuFile, result_path)

        # Edit Excel (giữ macro)
        wb = load_workbook(result_path, keep_vba=True)
        ws = wb["請求一覧"]

        for row in range(2, ws.max_row + 1):
            ws[f"G{row}"].number_format = "yyyy/m/d"

        wb.save(result_path)

        # Upload to MinIO (binary)
        with open(result_path, "rb") as f:
            data = f.read()

        result = ResultService.put_object(
            bucket_name=settings.RESULT_BUCKET,
            object_name=f"SeikyuNgoaiHanwa/{self.request.id}/seikuy_ngoai_hanwa.xlsm",
            data=io.BytesIO(data),
            length=len(data),
        )

        return f"{settings.RESULT_BUCKET}/{result.object_name}"
