import datetime
import io
import re
import tempfile

import pandas as pd
import redis
import validators
from celery import shared_task
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from src.core.config import settings
from src.core.logger import Log
from src.core.redis import REDIS_POOL
from src.core.type import API
from src.robot.SeikyuOnline.api import APISharePoint
from src.robot.SeikyuOnline.automation import SharePoint
from src.service.result import ResultService


def get_chubo(excelPath: str) -> float:
    def get_visible_sheets(excelPath: str) -> list[str]:
        try:
            wb = load_workbook(excelPath, read_only=True)
            visible_sheets = [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
            active_sheet = wb.active.title
            if active_sheet in visible_sheets:
                visible_sheets.remove(active_sheet)
                visible_sheets.insert(0, active_sheet)
            wb.close()
            return visible_sheets
        except Exception:
            with pd.ExcelFile(excelPath) as excel:
                sheets: list[str] = excel.sheet_names
            return sheets

    for sheet in get_visible_sheets(excelPath):
        data = pd.read_excel(
            io=excelPath,
            sheet_name=sheet,
            header=None,
        )
        for _, row in data.iterrows():
            row: str = " ".join(map(str, row.values))
            row = " ".join(row.replace("nan", "").split())
            if match := re.search(r"延床面積\s+(\d+\.\d+)\s*㎡\s*=\s*(\d+\.\d+)", row):
                return float(match.group(2))
    return 0


def get_price(excelPath: str) -> float:
    def get_visible_sheets(excelPath: str) -> list[str]:
        try:
            wb = load_workbook(excelPath, read_only=True)
            visible_sheets = [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
            wb.close()
            return visible_sheets
        except Exception:
            with pd.ExcelFile(excelPath) as excel:
                sheets: list[str] = excel.sheet_names
            return sheets

    for sheet in get_visible_sheets(excelPath):
        data = pd.read_excel(
            io=excelPath,
            sheet_name=sheet,
            header=None,
        )
        for index, row in data.iterrows():
            row: str = " ".join(map(str, row.values))
            row = " ".join(row.replace("nan", "").split())
            if match := re.search(r"小\s*.*\s*計.*?(\d+(?:,\d{3})*(?:\.\d+)?)", row):
                return float(match.group(1)), match.group(), index + 1, sheet
    return 0, None, None, None


@shared_task(
    bind=True,
    name="Seikyu Hanwa",
)
def seikyu_online(self, sheet_name: API | str = "/api/type/seikyu-online"):
    TaskID = self.request.id
    logger = Log.get_logger(channel=TaskID, redis_client=redis.Redis(connection_pool=REDIS_POOL))
    logger.info(f"Sheet name: {sheet_name}")
    if isinstance(sheet_name, API):
        sheet_name = sheet_name.url
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--start-maximized",
            ],
        )
        context = browser.new_context(no_viewport=True)
        with (
            SharePoint(
                domain=settings.SHAREPOINT_DOMAIN,
                username=settings.SHAREPOINT_EMAIL,
                password=settings.SHAREPOINT_PASSWORD,
                playwright=p,
                browser=browser,
                context=context,
            ) as sp,
            tempfile.TemporaryDirectory() as temp_dir,
        ):
            api = APISharePoint(
                TENANT_ID=settings.API_SHAREPOINT_TENANT_ID,
                CLIENT_ID=settings.API_SHAREPOINT_CLIENT_ID,
                CLIENT_SECRET=settings.API_SHAREPOINT_CLIENT_SECRET,
            )
            excel_file: str = api.download_item(
                site_id="nskkogyo.sharepoint.com,fcec6ca2-58f4-4488-abf8-34e8ffbb741d,3136a8b2-a506-44d2-ad49-324bd156147c",
                breadcrumb="◆請求書　ベトナム専用◆/≪ベトナム≫阪和興業　新(9日(10日分)、14日(15日分)、19日(20日分)、29日(末日分)に完成).xlsm",  # noqa
                save_to=temp_dir,
            )
            wb = load_workbook(excel_file, keep_vba=True)
            ws = wb[sheet_name]
            data = pd.read_excel(
                io=excel_file,
                sheet_name=sheet_name,
                header=None,
            )
            data.columns = [get_column_letter(i + 1) for i in range(len(data.columns))]  # Chuyển tên cột về chữ
            data.index = data.index + 1  # Index bắt đầu từ 1
            # ---- #
            for index, row in data.iterrows():
                LinkData: str = row["Y"]
                if not validators.url(LinkData):
                    continue
                # Trường hợp đặc biệt
                if pd.notna(row["J"]) and pd.notna(row["T"]):
                    continue
                if pd.notna(row["Z"]):
                    continue
                logger.info(f"{index} - {row["B"]} - {row["C"]} - {LinkData}")
                downloads = sp.download(
                    url=LinkData,
                    file=re.compile(".*.(xlsx|xlsm|xls)"),
                    steps=[
                        re.compile("^(見積|見積書)$"),
                    ],
                    save_to=temp_dir,
                )
                if len(downloads) != 1:
                    logger.warning("No downloaded files detected.")
                    continue
                for link, download in downloads:
                    price, raw, no_row, sheet_name = get_price(download)
                    if price != 0:
                        logger.info(f"Link: {link}")
                        logger.info(f"Price: {price} (Row {no_row} - Sheet {sheet_name} - Raw {raw})")
                        ws[f"J{index}"] = price
                        ws[f"T{index}"] = link
                        break
                if row["B"] in [
                    "阪和NO48・野原グループ㈱名古屋",
                    "阪和木材第2課NO2・金沢木材協同組合",
                    "住協建設㈱",
                ]:
                    for _, download in downloads:
                        chubo = get_chubo(download)
                        if chubo != 0:
                            logger.info(f"Chubo: {chubo}")
                            ws[f"Z{index}"] = chubo
                            break
            wb.save(excel_file)
            wb.close()

            with open(excel_file, "rb") as f:
                data = f.read()
            result = ResultService.put_object(
                bucket_name=settings.MINIO_BUCKET,
                object_name=f"SeikyuNgoaiHanwa/{datetime.datetime.now().strftime("%Y%m%d")}/{sheet_name}/SeikyuHanwa.xlsm",  # noqa
                data=io.BytesIO(data),
                length=len(data),
            )

            return result.object_name
