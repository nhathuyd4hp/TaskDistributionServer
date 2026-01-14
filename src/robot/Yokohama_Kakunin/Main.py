import argparse
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException, ElementNotInteractableException
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import PatternFill, Border, Side, Alignment
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from Nasiwak import Bot_Update,create_json_config
from openpyxl.worksheet.page import PageMargins
from selenium.webdriver.common.by import By
from tkinter import Label, Button, CENTER
from openpyxl import load_workbook
from tkcalendar import DateEntry
from selenium import webdriver
from datetime import timedelta
from datetime import datetime
import customtkinter as ctk 
import xlwings as xw
import babel.numbers
import tkinter as tk
import pandas as pd
import threading
import logging
import shutil
import time
import os
import requests
from config import BASE_URL
from token_manager import get_access_token
import locale
from logging_setup import setup_logging

# 🚀 Setup Log File
setup_logging()

# Set the locale to Japanese
locale.setlocale(locale.LC_ALL, 'ja_JP.UTF-8')

from config_access_token import token_file # noqa
# Replace with your actual file path
file_path = os.path.join(os.getcwd(), "Access_token", "Access_token.txt")
# logging.info(f"file path for text file is: {file_path}")
# Open and read the file
with open(file_path, "r", encoding="utf-8") as file:
    content = file.read()
logging.info(f"Extracted text from .txt file is: {content}")

class Yokohama_Kakunin:
    version = "2.0"

    def __init__(self, from_date, to_date):
        self.from_date = str(from_date)
        
        self.to_date = str(to_date)
        self.options = Options()
        self.options.add_argument("--headless")
    
    def format_date(self, date):
        date2 = datetime.strptime(str(date), "%Y-%m-%d")
        formatted_date = date2.strftime("%m月%d日")
        # Remove leading zero from the month
        if formatted_date[0] == '0':
            formatted_date = formatted_date[1:]
        return formatted_date
    
    def process_data(self):

        #Json
        WebAccess_config_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/webaccess.json"
        ACCESS_TOKEN = content
        Webaccess_Data= create_json_config(WebAccess_config_url,ACCESS_TOKEN)

        #Version Control
        REPO_OWNER = "Nasiwak"  # Your GitHub username
        REPO_NAME = "Yokohama_Kakunin"   # Your repo name
        CURRENT_VERSION = "v2.0" # This bot version
        Bot_Update(REPO_OWNER,REPO_NAME,CURRENT_VERSION,ACCESS_TOKEN)

        # Definitions
        excelFile = r"Data.xlsm"
        Over = r"結果.xlsx"
        CSVfolder = "CSV"
        current_date = datetime.now()
        logging.info(f"Current date: {current_date}")
        twodays_before = current_date - timedelta(days=2)

        def element_send_keys(locator, text, locator_type="xpath", timeout=20):
            try:
                # Choose the appropriate By strategy based on locator_type
                locator_by = {
                    "id": By.ID,
                    "name": By.NAME,
                    "xpath": By.XPATH,
                    "css_selector": By.CSS_SELECTOR,
                    "class_name": By.CLASS_NAME,
                    "tag_name": By.TAG_NAME,
                    "link_text": By.LINK_TEXT,
                    "partial_link_text": By.PARTIAL_LINK_TEXT
                }.get(locator_type.lower(), By.XPATH)  # Default to XPath if type is not recognized

                # Wait until the element is interactable and then send keys
                browserelement = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((locator_by, locator)))
                browserelement.clear()  # Clear any pre-existing text if needed
                browserelement.send_keys(text)
                logging.info(f"Successfully sent keys to element with {locator_type} '{locator}'")
                return True  # Indicate success
            except (NoSuchElementException, TimeoutException, ElementNotInteractableException) as e:
                logging.error(f"Failed to send keys to element with {locator_type} '{locator}': {e}")
                return False  # Indicate failure

        def element_click(locator, locator_type="xpath", timeout=20):
            try:
                # Choose the appropriate By strategy based on locator_type
                locator_by = {
                    "id": By.ID,
                    "name": By.NAME,
                    "xpath": By.XPATH,
                    "css_selector": By.CSS_SELECTOR,
                    "class_name": By.CLASS_NAME,
                    "tag_name": By.TAG_NAME,
                    "link_text": By.LINK_TEXT,
                    "partial_link_text": By.PARTIAL_LINK_TEXT
                }.get(locator_type.lower(), By.XPATH)  # Default to XPath if type is not recognized

                # Wait until the browserelement is interactable and then send keys
                browserelement = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((locator_by, locator)))
                browserelement.click()  # Click the browserelement
                logging.info(f"Successfully clicked on the element with {locator_type} '{locator}'")
                return True  # Indicate success
            except (NoSuchElementException, TimeoutException, ElementNotInteractableException) as e:
                logging.error(f"Failed to clicked on the element with {locator_type} '{locator}': {e}")
                return False  # Indicate failure

        def element_get_text(locator, locator_type="xpath", timeout=20):
            try:
                # Choose the appropriate By strategy based on locator_type
                locator_by = {
                    "id": By.ID,
                    "name": By.NAME,
                    "xpath": By.XPATH,
                    "css_selector": By.CSS_SELECTOR,
                    "class_name": By.CLASS_NAME,
                    "tag_name": By.TAG_NAME,
                    "link_text": By.LINK_TEXT,
                    "partial_link_text": By.PARTIAL_LINK_TEXT
                }.get(locator_type.lower(), By.XPATH)  # Default to XPath if type is not recognized

                # Wait until the element is visible and then retrieve its text
                browserelement = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((locator_by, locator)))
                browserelement_text = browserelement.text
                logging.info(f"Successfully retrieved text from element with {locator_type} '{locator}': {browserelement_text}")
                return browserelement_text  # Return the text of the element
            except (NoSuchElementException, TimeoutException, ElementNotInteractableException) as e:
                logging.error(f"Failed to retrieve text from element with {locator_type} '{locator}': {e}")
                return None  # Indicate failure by returning None

        def Accesslogin(driver):
            try:
                driver.get(Webaccess_Data['webaccess_url'])

                driver.switch_to.window(driver.window_handles[0])
                time.sleep(2)

                element_send_keys(Webaccess_Data['xpaths']['ログイン_xpaths']['ログインID'], "NasiwakRobot", locator_type="xpath")
                logging.info("ログインID input.")

                element_send_keys(Webaccess_Data['xpaths']['ログイン_xpaths']['パスワード'], "159753", locator_type="xpath")
                logging.info("パスワード input.")

                element_click(Webaccess_Data['xpaths']['ログイン_xpaths']['ログイン'], locator_type="xpath")
                logging.info("ログイン button clicked.")

                logging.info("Successfully logged in to Webaccess")
                time.sleep(2)
            except (NoSuchElementException, TimeoutException, WebDriverException) as e:
                logging.error(f"Failed during Access: {e}")
                return False

        def Access(driver,from_date, to_date):
            try:
                element_click(Webaccess_Data['xpaths']['受注一覧'], locator_type="xpath")
                logging.info("受注一覧 button clicked.")

                element_click(Webaccess_Data['xpaths']['受注一覧_xpaths']['リセット'], locator_type="xpath")
                logging.info("リセット button clicked.")

                fromD = str(from_date)
                fromDF = datetime.strptime(fromD, "%Y-%m-%d")
                f_fromD = fromDF.strftime("%Y/%m/%d")
                logging.info(f"From date: {f_fromD}")
                time.sleep(2)
                
                toD = str(to_date)
                toDF = datetime.strptime(toD,"%Y-%m-%d")
                f_toDFD = toDF.strftime("%Y/%m/%d")
                logging.info(f"To date :{f_toDFD}")
                time.sleep(2)

                # Clear the date delivery fields and enter the date range
                fromDateField = driver.find_element(By.NAME, "search_fix_deliver_date_from")
                fromDateField.clear()
                fromDateField.send_keys(f_fromD)
                logging.info ("From date sent")
                time.sleep(2)
                
                toDateField = driver.find_element(By.NAME, "search_fix_deliver_date_to")
                toDateField.send_keys(f_toDFD)
                logging.info("To date sent")
                time.sleep(2)

                element_click(Webaccess_Data['xpaths']['受注一覧_xpaths']['検索'], locator_type = "xpath")
                logging.info("検索 button clicked.")
                time.sleep(2)

                #Change the download directory
                driver.execute_cdp_cmd('Page.setDownloadBehavior',{'behavior':'allow','downloadPath':rf'{os.getcwd()}\{CSVfolder}'})    

                element_click(Webaccess_Data['xpaths']['受注一覧_xpaths']['案件一覧のダウンロード'], locator_type="xpath")
                logging.info("CSV出力/download button clicked.")                                                                                      
                time.sleep(2)
            except (NoSuchElementException, TimeoutException, WebDriverException) as e:
                logging.error(f"Failed during Access: {e}")
                return False
        def CsvMacro(excelFile):
            # Run Excel Macros
            try:
                app = xw.App(visible=False)

                wb = app.books.open(CSVfolder+'\\'+excelFile)
                logging.info("Excel File Opened Successfully")

                wb.macro('Clearalldata')()
                logging.info("cleanupsheets Macro Executed Successfully")

                wb.macro('DataProcess')()
                logging.info("GetFilePath Macro Executed Successfully")
                
                wb.save()
                logging.info("Excel File Saved Successfully")

            except:
                logging.info("ExcelMacroError: Some Macro Didn't execute properly")
                exit()

            finally:
                wb.close()
                app.quit()

                logging.info("Excel File Closed Successfully")

        def Excelformating(Over):
            wb = load_workbook(Over)
            ws = wb["Sheet1"]
            column_widths = {
                    "A": 10,
                    "B": 10,
                    "C": 10,
                }
            for column, width in column_widths.items():
                ws.column_dimensions[column].width = width

            header_border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='medium'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='medium')
            )

            thin_border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )

            for cell in ws[1]:
                cell.border = header_border

            for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
                for cell in row:
                    cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    if cell.value == "OK":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif cell.value == "NG":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            data_range = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
            for row in data_range:
                for cell in row:
                    align = Alignment(horizontal='center', vertical='center')
                    cell.alignment = align
            
            # Adjust the page layout options
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

            # Optional: Adjust margins if needed
            margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
            ws.page_margins = margins

            wb.save(Over)
        def clear_excel_data(Over):
            wb = load_workbook(Over)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=2):  # Assuming the first row has headers; adjust as needed
                    for cell in row:
                        cell.value = None  # Clear cell value
                        cell.fill = PatternFill(fill_type=None)  # Clear cell fill color
            wb.save(Over)
            logging.info("Data cleared from Excel file.")

        def search_anken_folder(anken_number):
            search_url = f"{BASE_URL}/search/query"
            payload = {
                "requests": [
                    {
                        "entityTypes": ["driveItem"],
                        "query": {"queryString": anken_number},
                        "from": 0,
                        "size": 1,
                        "region": "JPN"  # Limit results to only the first hit
                    }
                ]
            }
            headers = {
                "Authorization": f"Bearer {get_access_token()}",
                "Content-Type": "application/json"
            }
            response = requests.post(search_url, headers=headers, json=payload)
            if response.status_code != 200:
                logging.error(f"Search failed with status code {response.status_code}: {response.text}")
                return None
            results = response.json()
            if results['value'][0]['hitsContainers'][0]['total'] == 0:
                return None
            first_hit = results['value'][0]['hitsContainers'][0]['hits'][0]
            return first_hit['resource']['parentReference']['driveId'], first_hit['resource']['id']

        def list_children(drive_id, item_id):
            url = f"{BASE_URL}/drives/{drive_id}/items/{item_id}/children"
            headers = {"Authorization": f"Bearer {get_access_token()}"}
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                logging.error(f"Failed to list children with status code {response.status_code}: {response.text}")
                return []
            return response.json().get('value', [])

        def check_folder_exists(children, folder_name):
            return any(folder_name in child['name'] for child in children)

        def process_folder_checks(案件番号, Over, excelline):

            # Find the drive ID and item ID of the Anken folder
            search_result = search_anken_folder(案件番号)
            if not search_result:
                logging.info(f"No Anken folder found for {案件番号}")
                for col in ['A', 'B', 'C']:
                    sheet[f"{col}{excelline}"].value = "NG"
            else:
                drive_id, item_id = search_result
                children = list_children(drive_id, item_id)
                folder_checks = {
                    '割付図': 'B',
                    '案内図': 'C',
                    '見積': 'A'
                }
                for folder_name, column in folder_checks.items():
                    exists = check_folder_exists(children, folder_name)
                    sheet[f"{column}{excelline}"].value = "OK" if exists else "NG"
                    logging.info(f"Updated Excel {column}{excelline} with {'OK' if exists else 'NG'} for folder {folder_name}")

            wb.save(Over)

        chrome_options = Options()
        chrome_options.add_experimental_option("prefs", {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
        })
        driver = webdriver.Chrome(options=chrome_options)
        driver.maximize_window()
        if not os.path.exists(f'{CSVfolder}'):
                os.makedirs(f'{CSVfolder}')
                print("CSV Folder Created")
        else:
            print("Folder Exists")

        if not os.path.exists(f'{CSVfolder}/{excelFile}'):
            try:
                shutil.move(excelFile, f'{CSVfolder}/{excelFile}')
                print("Excel File Moved to CSV Folder")
            except:
                print("Excel File not found")
                exit()

        for f in os.listdir(CSVfolder):
            fpath = os.path.join(CSVfolder, f)
            if f != excelFile: 
                if os.path.isfile(fpath) or os.path.islink(fpath):
                    os.remove(fpath)  
                elif os.path.isdir(fpath):
                    shutil.rmtree(fpath)  
                print(f'Removed: {fpath}')
            else:
                print(f'Kept: {fpath}')

        clear_excel_data(Over)

        wb = load_workbook(os.path.join(Over))
        sheet = wb["Sheet1"]

        Accesslogin(driver)
        time.sleep(2)

        Access(driver, self.from_date, self.to_date)
        time.sleep(2)

        CsvMacro(excelFile)
        time.sleep(2)

        try:
            df = pd.read_excel(os.path.join(CSVfolder,excelFile), sheet_name='Sheet1', dtype=str).astype(str).dropna()
            logging.info(df)

            案件番号 = df['案件番号']  # Project Number (Column F)
            案件名 = df['物件名']   # Project Name (Column I)
            ビルダー名 = df['得意先名'] # Builder Name (Column G)
            商社名 = df['商社名']
            更新日 = df['更新日']
            確定納期 = df['確定納期']
            追加不足 = df['追加不足']
            軽天有無 = df['軽天有無']
            目地 = df['目地']
            目地数量 = df['目地数量']
            入隅 = df['入隅']
            入隅数量 = df ['入隅数量']
            配送時特記事項 = df['配送時特記事項']
            配送先住所 = df['配送先住所']
            階 = df['階']
            重量 = df ['重量']
            担当者 = df['担当者']
            連絡先 = df['連絡先']
            摘要1 = df['摘要1']
            確未 = df['確未']
            備考 = df['備考']

        
            excellinenumber = 2

            for row_number in range(len(案件名)):
                logging.info(f'{row_number}:Current Run {案件番号[row_number]}, {案件名[row_number]}')

                案件名[row_number] = 案件名[row_number].strip().replace('\t', '').replace('/', '').replace('\\', '').replace(':', '').replace('*', '').replace('?', '').replace('"', '').replace('<', '').replace('>', '').replace('|', '')
                time.sleep(1.5)

                if pd.isna(案件名[row_number]) or pd.isna(ビルダー名[row_number]):
                    break

                sheet[f'E{excellinenumber}'].value = 商社名[row_number]
                sheet[f'F{excellinenumber}'].value = 更新日[row_number]
                sheet[f'G{excellinenumber}'].value = 案件番号[row_number]
                sheet[f'H{excellinenumber}'].value = ビルダー名[row_number]
                sheet[f'I{excellinenumber}'].value = 案件名[row_number]
                sheet[f'J{excellinenumber}'].value = 確未[row_number]
                sheet[f'K{excellinenumber}'].value = 確定納期[row_number]
                sheet[f'L{excellinenumber}'].value = 追加不足[row_number]
                sheet[f'M{excellinenumber}'].value = 軽天有無[row_number]
                sheet[f'N{excellinenumber}'].value = 目地[row_number]
                sheet[f'O{excellinenumber}'].value = 目地数量[row_number]
                sheet[f'P{excellinenumber}'].value = 入隅[row_number]
                sheet[f'Q{excellinenumber}'].value = 入隅数量[row_number]
                sheet[f'R{excellinenumber}'].value = 配送時特記事項[row_number]
                sheet[f'S{excellinenumber}'].value = 配送先住所[row_number]
                sheet[f'T{excellinenumber}'].value = 階[row_number]
                sheet[f'U{excellinenumber}'].value = 重量[row_number]
                sheet[f'V{excellinenumber}'].value = 担当者[row_number]
                sheet[f'W{excellinenumber}'].value = 連絡先[row_number]
                sheet[f'X{excellinenumber}'].value = 摘要1[row_number]
                sheet[f'Y{excellinenumber}'].value = 備考[row_number]

                wb.save(Over)

                process_folder_checks(案件番号[row_number], Over, excellinenumber)

                wb.save(Over)
                excellinenumber+=1
                logging.info(f"current row number: {excellinenumber}\n")
        finally:
            driver.quit()
            time.sleep(2)
            Excelformating(Over)
            time.sleep(2)
            logging.info("Task completed")
class App(tk.Tk):
    def __init__(
        self,
        task_id: str,
        from_date: str,  # Format: yyyy-mm-dd hh:mm:ss.ffffff
        to_date: str,  # Format: yyyy-mm-dd hh:mm:ss.ffffff
    ):
        # ---- Convert Str to Datetime #
        self.from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S.%f")
        self.to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S.%f")
        # --- #
        tk.Tk.__init__(self)
        self.title(f"Yokohama Kakunin [{task_id}]")
        self.geometry("550x400")
        self.resizable(False, False)
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        Label(text="Yokohama Kakunin", font=("Roboto", -18, "bold")).pack(pady=10)

        from_label = Label(text="From:")
        from_label.place(relx=0.3, rely=0.4, anchor=CENTER)
        self.from_date_entry = DateEntry(self, date_pattern='yyyy/mm/dd')
        self.from_date_entry.place(relx=0.5, rely=0.4, anchor=CENTER)
        self.from_date_entry.set_date(self.from_date)

        to_label = Label(text="To:")
        to_label.place(relx=0.3, rely=0.5, anchor=CENTER)
        self.to_date_entry = DateEntry(self, date_pattern='yyyy/mm/dd')  # Initialize to_date_entry properly
        self.to_date_entry.place(relx=0.5, rely=0.5, anchor=CENTER)
        self.from_date_entry.set_date(self.from_date)
        
        start_button = Button(text="Start Bot", command=self.start_script, bg="#3290db", fg="white")
        start_button.place(relx=0.5, rely=0.6, anchor=CENTER)

    def start_script(self):
        self.from_date = self.from_date_entry.get_date()
        self.to_date = self.to_date_entry.get_date()
        threading.Thread(target=self.run_script).start()

    def run_script(self):
        try:
            bot = Yokohama_Kakunin(self.from_date, self.to_date)
            bot.process_data()
        except Exception as e:
            logging.info(e)

DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S.%f"


def parse_datetime(value: str) -> str:
    try:
        datetime.strptime(value, DATETIME_FORMAT)
        return value
    except ValueError as e:
        raise argparse.ArgumentTypeError(
            f"Invalid datetime format: {value}. " "Expected yyyy-mm-dd hh:mm:ss.fff"
        ) from e
    
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--task-id", default="Yokohama Kakunin", help="Task ID")
    parser.add_argument("--from-date", required=True, type=parse_datetime, help="Format: yyyy-mm-dd hh:mm:ss.fff")
    parser.add_argument("--to-date", required=True, type=parse_datetime, help="Format: yyyy-mm-dd hh:mm:ss.fff")
    args = parser.parse_args()

    app = App(
        task_id=args.task_id,
        from_date=args.from_date,
        to_date=args.to_date,
    )
    app.mainloop()