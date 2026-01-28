import logging
import os
import time
from datetime import datetime

import pandas as pd

# ===== DANDOLI =====
# ===== WEBACCESS =====
from Access import Accesslogin, webaccess_update_drawing_status

# ===== SHAREPOINT =====
from config import Download_folder
from dandoli_anken import (
    confirm_notification,
    enter_file_description,
    enter_notification_comment,
    go_to_genba_shiryo,
    move_users_to_receiver,
    open_bulk_upload_single_type,
    select_all_except_nsk,
    select_upload_type_shosetsu_kensetsu,
    send_notification,
    submit_upload,
    upload_single_pdf,
)
from dandoli_genba import (
    ensure_genba_kanri_home,
    enter_anken,
    force_return_to_genba_kanri,
    search_anken_by_name,
    wait_for_loader_to_disappear,  # ✅ use existing function in dandoli_genba
    wait_for_search_results,
)
from dandoli_login import login_dandoli
from dandoli_state import ensure_home_screen, ensure_place

# ===== YAMADA PLACE SWITCH =====
from ensure_place_yamada import ensure_place_yamada
from logging_setup import setup_logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.sync_api import sync_playwright

# ===== CONSTANTS =====
# ==================================================
# INITIALISE LOGGING (ONCE)
# ==================================================
setup_logging()
logger = logging.getLogger(__name__)
logger.info("===== BOT START =====")

EXPECTED_BUILDER = "建新の現場管理（建築）"
EXPECTED_COMPANY = None

INPUT_EXCEL = "健新～ヤマダホームズの図面送付.xlsx"
DOWNLOAD_ROOT = os.path.join(os.getcwd(), "Downloaded_Zumen")
FILE_DESCRIPTION = "軽天割付図"

# ===== REPORT =====
RESULTS = []


# ==================================================
# REPORT HELPERS
# ==================================================
def add_result(anken_no, builder, anken_name, zumen, access, remarks=""):
    RESULTS.append(
        {
            "案件番号": anken_no,
            "ビルダー名": builder,
            "案件名": anken_name,
            "図面UP": zumen,
            "Access変更": access,
            "備考": remarks,
        }
    )


def save_report():
    if not RESULTS:
        logger.warning("⚠ No results to write to report")
        return

    df = pd.DataFrame(RESULTS)

    reports_dir = os.path.join(os.getcwd(), "ProgressReports")
    os.makedirs(reports_dir, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(reports_dir, f"Progress_Report_{ts}.xlsx")

    logger.info("📊 Writing progress report")
    df.to_excel(path, index=False)

    # ---------- Beautification ----------
    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    fill_ok_green = PatternFill("solid", fgColor="C6EFCE")
    fill_ok_blue = PatternFill("solid", fgColor="D9E1F2")
    fill_ng_red = PatternFill("solid", fgColor="F8CBAD")
    fill_na_grey = PatternFill("solid", fgColor="E7E6E6")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column_letter in ("D", "E"):
                cell.alignment = center
                if cell.value == "OK":
                    cell.fill = fill_ok_green if cell.column_letter == "D" else fill_ok_blue
                elif cell.value == "NG":
                    cell.fill = fill_ng_red
                else:
                    cell.fill = fill_na_grey
            else:
                cell.alignment = left

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(path)
    logger.info(f"🎨 Beautified Progress Report saved: {path}")


# ==================================================
# MAIN
# ==================================================
def main():
    logger.info("📂 Loading input Excel")
    df = pd.read_excel(INPUT_EXCEL)

    with sync_playwright() as p:
        logger.info("🌐 Launching browser")
        browser = p.chromium.launch(headless=False, slow_mo=50)
        context = browser.new_context(viewport={"width": 1280, "height": 900})

        dandoli_page = context.new_page()
        access_page = context.new_page()

        # ---------- LOGIN (ONLY AUTH HERE) ----------
        logger.info("🔐 Logging into Dandoli")
        login_dandoli(dandoli_page)
        ensure_home_screen(dandoli_page)

        logger.info("🔐 Logging into WebAccess")
        if not Accesslogin(access_page):
            logger.critical("❌ WebAccess login failed – aborting run")
            return

        # ---------- MAIN LOOP ----------
        for idx, row in df.iterrows():

            # ✅ HARD STOP if Playwright page died
            if dandoli_page.is_closed():
                logger.critical("🔥 Dandoli page closed – aborting remaining ankens")
                break

            anken_no = str(row.get("案件番号", "")).strip()
            builder = str(row.get("ビルダー名", "")).strip()
            anken_name = str(row.get("案件名", "")).strip()
            shiten = str(row.get("支店名", "")).strip()

            logger.info("=" * 60)
            logger.info(f"🚀 START ANKEN [{idx+1}/{len(df)}] : {anken_no} | {anken_name} | {builder} | 支店={shiten}")

            zumen_result = "-"
            access_result = "-"
            remarks = ""

            try:
                # ==================================================
                # 🏢 ENSURE CORRECT PLACE (PER-ANKEN)
                # ==================================================
                try:
                    if builder in ("ヤマダホームズ", "ﾔﾏﾀﾞﾎｰﾑｽﾞ"):
                        logger.info("🏢 Ensuring Yamada place")
                        if not ensure_place_yamada(dandoli_page, shiten):
                            remarks = "支店切替失敗"
                            add_result(anken_no, builder, anken_name, "NG", "-", remarks)
                            logger.warning("⚠ 支店切替失敗")
                            continue

                        # 🚨 IMPORTANT: enter Genba ONLY AFTER Yamada switch
                        ensure_genba_kanri_home(dandoli_page)

                    else:
                        logger.info("🏢 Ensuring Kenshin place")
                        ensure_place(dandoli_page, EXPECTED_BUILDER, EXPECTED_COMPANY)
                        ensure_genba_kanri_home(dandoli_page)

                    # always wait for loader after navigation to avoid click interception
                    wait_for_loader_to_disappear(dandoli_page)

                except Exception:
                    logger.error("❌ Failed during place selection / genba home", exc_info=True)
                    add_result(anken_no, builder, anken_name, "NG", "-", "Place切替/現場管理失敗")
                    continue

                # ----- SharePoint -----
                logger.info("📥 Downloading 図面 from SharePoint")
                download_dir = os.path.join(DOWNLOAD_ROOT, anken_no)
                pdf_path = Download_folder(anken_no, download_dir)

                if not pdf_path:
                    remarks = "図面PDFなし"
                    add_result(anken_no, builder, anken_name, "NG", "-", remarks)
                    logger.warning("⚠ 図面PDFなし")
                    continue

                # ----- Dandoli Search -----
                logger.info("🔍 Searching anken in Dandoli")
                search_anken_by_name(dandoli_page, anken_name)
                time.sleep(1)

                try:
                    # ✅ critical: wait for loader and post-search DOM to reflect
                    wait_for_loader_to_disappear(dandoli_page)
                    wait_for_search_results(dandoli_page, anken_name)

                except Exception:
                    remarks = "Dandoli案件未検出"
                    add_result(anken_no, builder, anken_name, "NG", "-", remarks)
                    logger.warning("⚠ Dandoli案件未検出")
                    continue

                # ----- Upload + Notify -----
                try:
                    logger.info("📤 Uploading 図面 & sending notification")
                    time.sleep(1)

                    # ✅ patched: enter_anken now takes ONLY page
                    enter_anken(dandoli_page)

                    go_to_genba_shiryo(dandoli_page)
                    time.sleep(1)

                    open_bulk_upload_single_type(dandoli_page)
                    select_upload_type_shosetsu_kensetsu(dandoli_page)
                    upload_single_pdf(dandoli_page, pdf_path)
                    enter_file_description(dandoli_page, FILE_DESCRIPTION)

                    submit_upload(dandoli_page)
                    time.sleep(1)

                    confirm_notification(dandoli_page)
                    time.sleep(1)

                    select_all_except_nsk(dandoli_page)
                    move_users_to_receiver(dandoli_page)
                    enter_notification_comment(dandoli_page)
                    send_notification(dandoli_page)

                    zumen_result = "OK"
                    logger.info("✅ 図面UP + 通知 完了")

                except Exception:
                    logger.error("❌ 図面UP失敗", exc_info=True)
                    remarks = "図面UP失敗"
                    add_result(anken_no, builder, anken_name, "NG", "-", remarks)
                    try:
                        force_return_to_genba_kanri(dandoli_page)
                        wait_for_loader_to_disappear(dandoli_page)
                    except Exception:
                        logger.warning("⚠ force_return_to_genba_kanri failed after upload error", exc_info=True)
                    continue

                # ----- WebAccess -----
                logger.info("🔄 Updating WebAccess status")
                wa_result = webaccess_update_drawing_status(access_page, anken_no)

                if wa_result in ("UPDATED", "NO_CHANGE"):
                    access_result = "OK"
                    if wa_result == "NO_CHANGE":
                        if not remarks:
                            remarks = "Access変更不要"
                        logger.info("ℹ Access変更不要")
                    else:
                        logger.info("✅ Access変更完了")
                else:
                    access_result = "NG"
                    if not remarks:
                        remarks = "Access変更失敗"
                    logger.warning("⚠ Access変更失敗")

                try:
                    force_return_to_genba_kanri(dandoli_page)
                    wait_for_loader_to_disappear(dandoli_page)
                except Exception:
                    logger.warning("⚠ force_return_to_genba_kanri failed at end of anken", exc_info=True)

                add_result(anken_no, builder, anken_name, zumen_result, access_result, remarks)

                logger.info(f"🏁 END ANKEN : {anken_no}")

            except Exception:
                logger.error("🔥 Unexpected fatal error in anken loop", exc_info=True)
                add_result(anken_no, builder, anken_name, "NG", "-", "予期せぬエラー")
                try:
                    force_return_to_genba_kanri(dandoli_page)
                    wait_for_loader_to_disappear(dandoli_page)
                except Exception:
                    logger.warning("⚠ force_return_to_genba_kanri failed after fatal error", exc_info=True)

        save_report()
        logger.info("🧹 Closing browser")

        context.close()
        browser.close()

    logger.info("===== BOT END =====")


if __name__ == "__main__":
    main()
